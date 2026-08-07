"""Authenticated firm workspace and secure accountless portal APIs for PBC."""

from __future__ import annotations

import asyncio
import hashlib
import io
import logging
import os
import re
import secrets
import shutil
import subprocess
import tempfile
import uuid
from datetime import date, timedelta
from pathlib import PurePath
from urllib.parse import quote
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from fastapi import APIRouter, Depends, File, Header, HTTPException, Query, Request, Response, UploadFile
from fastapi.encoders import jsonable_encoder
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import load_workbook
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from core.database import get_db
from core.runtime import frontend_base_url, is_local
from dependencies.analytics_rbac import READER_ROLES, require_role
from dependencies.auth import verify_firebase_token
from models.db_models import Client, User
from models.pbc import (
    PbcAccessToken,
    PbcComment,
    PbcContact,
    PbcDocument,
    PbcEngagement,
    PbcEngagementContact,
    PbcFirmSettings,
    PbcNotificationOutbox,
    PbcPortalSession,
    PbcRequest,
    PbcRequestAssignment,
    PbcTemplate,
)
from models.pbc_schemas import (
    PbcAccessLinkCreate,
    PbcAiDraftRequest,
    PbcAiMatchRequest,
    PbcCommentCreate,
    PbcContactAssignment,
    PbcContactCreate,
    PbcEngagementCreate,
    PbcEngagementUpdate,
    PbcImportCommit,
    PbcPortalExchange,
    PbcRequestCreate,
    PbcRequestUpdate,
    PbcSettingsUpdate,
    PbcTemplateCreate,
    PbcTransitionRequest,
    PbcUploadComplete,
    PbcUploadInitiate,
)
from models.tasklytic import TasklyticEntityRecord, TasklyticWorkspaceMember
from services.email_service import email_service
from services.gcs_service import get_storage_service
from services.pbc_ai_service import completeness_flags, draft_request_list, match_document
from services.pbc_service import (
    FIRM_MANAGE_ROLES,
    FIRM_WRITE_ROLES,
    actor_role,
    audit,
    build_metadata_zip,
    build_tracker_xlsx,
    create_access_token,
    exchange_access_token,
    ensure_default_template,
    get_engagement,
    get_pbc_request,
    get_portal_session,
    get_settings,
    portal_request_ids,
    queue_contact_notifications,
    require_actor_role,
    require_portal_request,
    serialize_contact,
    serialize_document,
    serialize_engagement,
    serialize_request,
    token_hash,
    transition_request,
    utcnow,
    validate_tasklytic_link,
)
from services.rate_limit import rate_limiter
from services.analytics.firm_scope import require_firm_id


router = APIRouter(prefix="/api/pbc", tags=["pbc"])
portal_router = APIRouter(prefix="/api/pbc/portal", tags=["pbc-portal"])
COOKIE_NAME = "pbc_portal_session"
MAX_UPLOAD = 100 * 1024 * 1024
BLOCKED_EXTENSIONS = {".app", ".bat", ".cmd", ".com", ".dll", ".dmg", ".exe", ".jar", ".js", ".lnk", ".msi", ".ps1", ".reg", ".scr", ".sh", ".vbs"}
SAFE_MIME_PREFIXES = ("application/pdf", "application/zip", "application/vnd.", "text/", "image/")
logger = logging.getLogger(__name__)


def _commit(db: Session) -> None:
    try:
        db.commit()
    except Exception:
        db.rollback()
        raise


def _firm_id(db: Session, actor: User):
    return require_firm_id(db, actor.id)


def _actor_name(actor: User) -> str:
    return actor.display_name or actor.email


def _request_number(db: Session, engagement_id) -> str:
    existing = {value[0] for value in db.query(PbcRequest.request_number).filter(PbcRequest.engagement_id == engagement_id)}
    index = len(existing) + 1
    while f"PBC-{index:03d}" in existing:
        index += 1
    return f"PBC-{index:03d}"


def _request_assignment_ids(db: Session, engagement_id, request_ids: list[str]) -> list[uuid.UUID]:
    """Validate request membership and retain the UUID type expected by SQLAlchemy."""
    known = {
        str(request_id): request_id
        for (request_id,) in db.query(PbcRequest.id).filter(PbcRequest.engagement_id == engagement_id).all()
    }
    if any(request_id not in known for request_id in request_ids):
        raise HTTPException(status_code=422, detail="Assigned request does not belong to this engagement")
    return [known[request_id] for request_id in dict.fromkeys(request_ids)]


def _safe_upload(payload: PbcUploadInitiate) -> tuple[str, str, int]:
    filename = PurePath(payload.filename.replace("\\", "/")).name
    suffix = PurePath(filename.lower()).suffix
    if not filename or suffix in BLOCKED_EXTENSIONS:
        raise HTTPException(status_code=415, detail="This file type is not allowed")
    mime = payload.content_type.lower().strip()
    if not mime.startswith(SAFE_MIME_PREFIXES):
        raise HTTPException(status_code=415, detail="This file type is not allowed")
    return filename, mime, payload.size


async def _verify_object(document: PbcDocument) -> None:
    storage = get_storage_service()
    if not getattr(storage, "bucket", None):
        raise HTTPException(status_code=503, detail="Direct PBC uploads require configured object storage")
    blob = storage.bucket.blob(document.object_name)
    await asyncio.to_thread(blob.reload)
    if not blob.exists() or int(blob.size or -1) != int(document.size_bytes):
        raise HTTPException(status_code=409, detail="Uploaded object is missing or has an unexpected size")
    content_type = getattr(blob, "content_type", None)
    if content_type and content_type.lower() != document.mime_type.lower():
        raise HTTPException(status_code=409, detail="Uploaded object has an unexpected content type")


async def _scan_document(document: PbcDocument) -> tuple[str, str | None, str]:
    """Run ClamAV when available; production fails closed when it is unavailable."""
    storage = get_storage_service()
    clamscan = shutil.which(os.getenv("PBC_CLAMSCAN_BINARY", "clamscan"))
    with tempfile.NamedTemporaryFile(prefix="pbc-scan-", delete=True) as handle:
        await asyncio.to_thread(storage.bucket.blob(document.object_name).download_to_filename, handle.name)
        def _checksum() -> str:
            with open(handle.name, "rb") as source:
                return hashlib.file_digest(source, "sha256").hexdigest()
        actual_checksum = await asyncio.to_thread(_checksum)
        if not clamscan:
            if is_local():
                return "skipped", "Malware scanner unavailable in local development", actual_checksum
            logger.error("PBC malware scanner binary is unavailable for document %s", document.id)
            return "failed", "Malware scanner unavailable", actual_checksum
        try:
            result = await asyncio.to_thread(
                subprocess.run, [clamscan, "--no-summary", handle.name], capture_output=True, text=True, timeout=180
            )
        except subprocess.TimeoutExpired:
            logger.error("PBC malware scan timed out for document %s", document.id)
            return "failed", "Malware scan timed out", actual_checksum
    detail = (result.stdout or result.stderr or "").strip()[-2000:]
    if result.returncode == 0:
        return "clean", detail or None, actual_checksum
    if result.returncode == 1:
        return "infected", detail or "Malware detected", actual_checksum
    logger.error("PBC malware scan failed for document %s: %s", document.id, detail or f"exit code {result.returncode}")
    return "failed", detail or "Malware scan failed", actual_checksum


def _raise_for_failed_scan(document: PbcDocument) -> None:
    """Do not tell clients a quarantined or infected upload succeeded."""
    if document.state == "available":
        return
    if document.scan_status == "infected":
        raise HTTPException(status_code=422, detail="The uploaded file was rejected because malware was detected")
    raise HTTPException(
        status_code=503,
        detail="The file was uploaded, but its security scan could not be completed. Please try again shortly.",
    )


def _rate_limit(request: Request, key: str, limit: int = 30) -> None:
    forwarded = request.headers.get("x-forwarded-for", "").split(",")[0].strip()
    client = forwarded or (request.client.host if request.client else "unknown")
    if not rate_limiter.check(f"pbc:{key}", client, limit, 60):
        raise HTTPException(status_code=429, detail="Too many requests; try again shortly")


def _portal_context(db: Session, request: Request, csrf: str | None = None, *, mutate: bool = False):
    session, contact, engagement = get_portal_session(db, request.cookies.get(COOKIE_NAME))
    if mutate and (not csrf or not secrets.compare_digest(token_hash(csrf), session.csrf_hash)):
        raise HTTPException(status_code=403, detail="Invalid PBC portal CSRF token")
    return session, contact, engagement


@router.get("/context")
def context(actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    firm_id = _firm_id(db, actor)
    settings = get_settings(db, firm_id)
    ensure_default_template(db, firm_id, actor.id)
    _commit(db)
    return {
        "firm_id": str(firm_id), "role": actor_role(actor),
        "settings": {"timezone": settings.timezone, "portal_name": settings.portal_name, "logo_url": settings.logo_url,
                     "reminder_days_before": settings.reminder_days_before, "overdue_interval_days": settings.overdue_interval_days},
    }


@router.get("/dashboard")
def dashboard(actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    firm_id = _firm_id(db, actor)
    engagements = db.query(PbcEngagement).filter(PbcEngagement.firm_id == firm_id, PbcEngagement.status != "archived").all()
    requests = db.query(PbcRequest).filter(PbcRequest.firm_id == firm_id).all()
    today = date.today()
    return {
        "active_engagements": sum(item.status == "active" for item in engagements),
        "total_requests": len(requests),
        "accepted_requests": sum(item.status == "accepted" for item in requests),
        "awaiting_review": sum(item.status == "submitted" for item in requests),
        "overdue": sum(bool(item.due_date and item.due_date < today and item.status not in {"accepted", "waived"}) for item in requests),
        "due_soon": sum(bool(item.due_date and 0 <= (item.due_date - today).days <= 7 and item.status not in {"accepted", "waived"}) for item in requests),
        "engagements": [serialize_engagement(db, item) for item in sorted(engagements, key=lambda value: value.updated_at, reverse=True)[:10]],
    }


@router.get("/project-links")
def project_links(actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    workspace_ids = [item[0] for item in db.query(TasklyticWorkspaceMember.workspace_id).filter(
        TasklyticWorkspaceMember.user_id == actor.id
    ).all()]
    if not workspace_ids:
        return {"projects": []}
    rows = db.query(TasklyticEntityRecord).filter(
        TasklyticEntityRecord.entity_kind == "projects", TasklyticEntityRecord.workspace_id.in_(workspace_ids)
    ).all()
    projects = []
    for row in rows:
        try:
            validate_tasklytic_link(db, actor.id, row.workspace_id, row.record_id)
        except HTTPException:
            continue
        projects.append({"workspace_id": row.workspace_id, "project_id": row.record_id,
                         "name": str((row.payload or {}).get("name") or "Untitled project")})
    return {"projects": sorted(projects, key=lambda item: item["name"].lower())}


@router.get("/settings")
def read_settings(actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    row = get_settings(db, _firm_id(db, actor))
    return {key: getattr(row, key) for key in ("timezone", "portal_name", "logo_url", "reminder_days_before", "overdue_interval_days")}


@router.put("/settings")
def update_settings(payload: PbcSettingsUpdate, actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    require_actor_role(actor, {"admin"})
    try:
        ZoneInfo(payload.timezone)
    except ZoneInfoNotFoundError:
        raise HTTPException(status_code=422, detail="Unknown IANA timezone")
    row = get_settings(db, _firm_id(db, actor))
    for key, value in payload.model_dump().items():
        setattr(row, key, value)
    _commit(db)
    return read_settings(actor, db)


@router.get("/engagements")
def list_engagements(status: str | None = None, actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    query = db.query(PbcEngagement).filter(PbcEngagement.firm_id == _firm_id(db, actor))
    if status:
        query = query.filter(PbcEngagement.status == status)
    return {"engagements": [serialize_engagement(db, item) for item in query.order_by(PbcEngagement.updated_at.desc()).all()]}


@router.post("/engagements", status_code=201)
def create_engagement(payload: PbcEngagementCreate, actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    require_actor_role(actor, FIRM_WRITE_ROLES)
    if payload.template_id and payload.rollover_from_id:
        raise HTTPException(status_code=422, detail="Choose a template or a rollover source, not both")
    firm_id = _firm_id(db, actor)
    client = db.query(Client).filter(Client.id == payload.client_id, Client.firm_id == firm_id).first()
    if client is None:
        raise HTTPException(status_code=422, detail="Client not found in this firm")
    validate_tasklytic_link(db, actor.id, payload.tasklytic_workspace_id, payload.tasklytic_project_id)
    row = PbcEngagement(
        firm_id=firm_id, client_id=client.id, client_name_snapshot=client.name, name=payload.name,
        engagement_type=payload.engagement_type, period_start=payload.period_start, period_end=payload.period_end,
        due_date=payload.due_date, owner_user_id=payload.owner_user_id or actor.id,
        tasklytic_workspace_id=payload.tasklytic_workspace_id, tasklytic_project_id=payload.tasklytic_project_id,
        created_by_user_id=actor.id,
    )
    db.add(row)
    db.flush()
    source_items: list[dict] = []
    if payload.template_id:
        template = db.query(PbcTemplate).filter(PbcTemplate.id == payload.template_id, PbcTemplate.firm_id == firm_id).first()
        if template is None:
            raise HTTPException(status_code=422, detail="PBC template not found")
        source_items = list(template.items or [])
    elif payload.rollover_from_id:
        source = get_engagement(db, firm_id, payload.rollover_from_id)
        source_items = [{
            "request_number": item.request_number, "category": item.category, "title": item.title,
            "description": item.description, "priority": item.priority, "owner_user_id": item.owner_user_id,
            "expected_filename": item.expected_filename, "expected_formats": item.expected_formats or [],
            "gl_account": item.gl_account, "gl_balance": item.gl_balance, "sensitive": item.sensitive,
            "requires_redaction": item.requires_redaction,
        } for item in db.query(PbcRequest).filter(PbcRequest.engagement_id == source.id).order_by(PbcRequest.sort_order)]
    for index, item in enumerate(source_items):
        db.add(PbcRequest(
            engagement_id=row.id, firm_id=firm_id, request_number=str(item.get("request_number") or f"PBC-{index + 1:03d}"),
            sort_order=index, category=item.get("category"), title=str(item.get("title") or "Untitled request")[:500],
            description=item.get("description"), priority=item.get("priority") if item.get("priority") in {"low","normal","high","urgent"} else "normal",
            due_date=payload.due_date, owner_user_id=item.get("owner_user_id") or actor.id,
            expected_filename=item.get("expected_filename"), expected_formats=item.get("expected_formats") or [],
            gl_account=item.get("gl_account"), gl_balance=str(item.get("gl_balance")) if item.get("gl_balance") is not None else None,
            sensitive=bool(item.get("sensitive")), requires_redaction=bool(item.get("requires_redaction")), created_by_user_id=actor.id,
        ))
    audit(db, row, "engagement_created", actor_kind="firm", actor_id=actor.id,
          details={"source": "template" if payload.template_id else "rollover" if payload.rollover_from_id else "blank"})
    _commit(db)
    return serialize_engagement(db, row, detail=True)


@router.get("/engagements/{engagement_id}")
def read_engagement(engagement_id: str, actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    return serialize_engagement(db, get_engagement(db, _firm_id(db, actor), engagement_id), detail=True)


@router.patch("/engagements/{engagement_id}")
def update_engagement(engagement_id: str, payload: PbcEngagementUpdate, actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    require_actor_role(actor, FIRM_WRITE_ROLES)
    row = get_engagement(db, _firm_id(db, actor), engagement_id, lock=True)
    if row.revision != payload.expected_revision:
        raise HTTPException(status_code=409, detail="Engagement changed; refresh and try again")
    supplied = payload.model_fields_set
    if "tasklytic_workspace_id" in supplied or "tasklytic_project_id" in supplied:
        validate_tasklytic_link(db, actor.id, payload.tasklytic_workspace_id, payload.tasklytic_project_id)
    changes = {}
    for key in supplied - {"expected_revision"}:
        old, new = getattr(row, key), getattr(payload, key)
        if old != new:
            changes[key] = {"from": str(old) if old is not None else None, "to": str(new) if new is not None else None}
            setattr(row, key, new)
    row.revision += 1
    audit(db, row, "engagement_updated", actor_kind="firm", actor_id=actor.id, details={"changes": changes})
    _commit(db)
    return serialize_engagement(db, row, detail=True)


@router.post("/engagements/{engagement_id}/publish")
def publish_engagement(engagement_id: str, actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    require_actor_role(actor, FIRM_MANAGE_ROLES)
    row = get_engagement(db, _firm_id(db, actor), engagement_id, lock=True)
    if row.status != "draft":
        raise HTTPException(status_code=409, detail="Only a draft engagement can be published")
    requests = db.query(PbcRequest).filter(PbcRequest.engagement_id == row.id).all()
    contacts = db.query(PbcEngagementContact).filter(PbcEngagementContact.engagement_id == row.id).count()
    coordinators = db.query(PbcEngagementContact).filter(
        PbcEngagementContact.engagement_id == row.id, PbcEngagementContact.role == "coordinator"
    ).count()
    invalid = []
    for item in requests:
        has_recipient = coordinators > 0 or db.query(PbcRequestAssignment).filter(PbcRequestAssignment.request_id == item.id).count() > 0
        if not item.title or not item.owner_user_id or not item.due_date or not has_recipient:
            invalid.append(item.request_number)
    if not requests or invalid or contacts == 0:
        raise HTTPException(status_code=422, detail={"message": "Publishing requires requests, contacts, owners, and due dates", "invalid_requests": invalid})
    for item in requests:
        if item.status == "draft":
            item.status = "open"
            item.revision += 1
    row.status = "active"
    row.published_at = utcnow()
    row.revision += 1
    audit(db, row, "engagement_published", actor_kind="firm", actor_id=actor.id, details={"request_count": len(requests)})
    queue_contact_notifications(db, row, kind="published", details={"nonce": row.revision})
    _commit(db)
    return serialize_engagement(db, row, detail=True)


@router.post("/engagements/{engagement_id}/actions/{action}")
def engagement_action(engagement_id: str, action: str, actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    require_actor_role(actor, FIRM_MANAGE_ROLES)
    row = get_engagement(db, _firm_id(db, actor), engagement_id, lock=True)
    if action == "complete" and row.status == "active":
        open_count = db.query(PbcRequest).filter(PbcRequest.engagement_id == row.id, PbcRequest.status.notin_(["accepted", "waived"])).count()
        if open_count:
            raise HTTPException(status_code=409, detail="Accept or waive every request before completing the engagement")
        row.status, row.completed_at = "completed", utcnow()
    elif action == "archive" and row.status in {"draft", "completed"}:
        row.status, row.archived_at = "archived", utcnow()
    else:
        raise HTTPException(status_code=409, detail="Engagement action is not allowed in its current status")
    row.revision += 1
    audit(db, row, f"engagement_{action}", actor_kind="firm", actor_id=actor.id)
    _commit(db)
    return serialize_engagement(db, row, detail=True)


@router.post("/engagements/{engagement_id}/requests", status_code=201)
def create_request(engagement_id: str, payload: PbcRequestCreate, actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    require_actor_role(actor, FIRM_WRITE_ROLES)
    engagement = get_engagement(db, _firm_id(db, actor), engagement_id)
    if engagement.status in {"completed", "archived"}:
        raise HTTPException(status_code=409, detail="This engagement is locked")
    count = db.query(PbcRequest).filter(PbcRequest.engagement_id == engagement.id).count()
    row = PbcRequest(
        engagement_id=engagement.id, firm_id=engagement.firm_id,
        request_number=payload.request_number or _request_number(db, engagement.id), sort_order=count,
        category=payload.category, title=payload.title, description=payload.description, period_end=payload.period_end,
        priority=payload.priority, due_date=payload.due_date or engagement.due_date,
        owner_user_id=payload.owner_user_id or actor.id, expected_filename=payload.expected_filename,
        expected_formats=[value.lower().lstrip(".") for value in payload.expected_formats], gl_account=payload.gl_account,
        gl_balance=payload.gl_balance, sensitive=payload.sensitive, requires_redaction=payload.requires_redaction,
        dependency_ids=payload.dependency_ids, external_source_id=payload.external_source_id,
        status="open" if engagement.status == "active" else "draft", created_by_user_id=actor.id,
    )
    db.add(row)
    db.flush()
    audit(db, engagement, "request_created", actor_kind="firm", actor_id=actor.id, request_id=row.id)
    _commit(db)
    return serialize_request(db, row)


@router.patch("/requests/{request_id}")
def update_request(request_id: str, payload: PbcRequestUpdate, actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    require_actor_role(actor, FIRM_WRITE_ROLES)
    row = get_pbc_request(db, _firm_id(db, actor), request_id, lock=True)
    engagement = get_engagement(db, row.firm_id, str(row.engagement_id))
    if engagement.status in {"completed", "archived"} or row.status in {"accepted", "waived"}:
        raise HTTPException(status_code=409, detail="This PBC request is locked")
    if row.revision != payload.expected_revision:
        raise HTTPException(status_code=409, detail="Request changed; refresh and try again")
    changes = {}
    for key in payload.model_fields_set - {"expected_revision"}:
        new = getattr(payload, key)
        if key == "expected_formats" and new is not None:
            new = [value.lower().lstrip(".") for value in new]
        old = getattr(row, key)
        if old != new:
            changes[key] = {"from": str(old) if old is not None else None, "to": str(new) if new is not None else None}
            setattr(row, key, new)
    row.revision += 1
    audit(db, engagement, "request_updated", actor_kind="firm", actor_id=actor.id, request_id=row.id, details={"changes": changes})
    _commit(db)
    return serialize_request(db, row)


@router.delete("/requests/{request_id}", status_code=204)
def delete_request(request_id: str, actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    require_actor_role(actor, FIRM_WRITE_ROLES)
    row = get_pbc_request(db, _firm_id(db, actor), request_id, lock=True)
    engagement = get_engagement(db, row.firm_id, str(row.engagement_id))
    if engagement.status != "draft" or row.status != "draft":
        raise HTTPException(status_code=409, detail="Published PBC requests must be waived rather than deleted")
    audit(db, engagement, "request_deleted", actor_kind="firm", actor_id=actor.id, request_id=row.id,
          details={"request_number": row.request_number, "title": row.title})
    db.delete(row)
    _commit(db)


@router.post("/requests/{request_id}/transition")
def firm_transition(request_id: str, payload: PbcTransitionRequest, actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    row = get_pbc_request(db, _firm_id(db, actor), request_id, lock=True)
    engagement = get_engagement(db, row.firm_id, str(row.engagement_id))
    transition_request(db, row, engagement, payload.action, payload.reason, payload.expected_revision,
                       actor_kind="firm", actor_id=actor.id, actor_role_value=actor_role(actor))
    if payload.action == "return":
        queue_contact_notifications(db, engagement, kind="returned", request_row=row,
                                    details={"reason": payload.reason, "nonce": row.revision})
    _commit(db)
    return serialize_request(db, row)


@router.post("/requests/{request_id}/comments", status_code=201)
def add_firm_comment(request_id: str, payload: PbcCommentCreate, actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    row = get_pbc_request(db, _firm_id(db, actor), request_id)
    comment = PbcComment(request_id=row.id, firm_id=row.firm_id, body=payload.body, visibility=payload.visibility,
                         actor_kind="firm", actor_id=actor.id, actor_name=_actor_name(actor))
    db.add(comment)
    db.flush()
    engagement = get_engagement(db, row.firm_id, str(row.engagement_id))
    audit(db, engagement, "comment_added", actor_kind="firm", actor_id=actor.id, request_id=row.id,
          details={"visibility": payload.visibility})
    if payload.visibility == "client":
        queue_contact_notifications(db, engagement, kind="comment", request_row=row,
                                    details={"nonce": str(comment.id)})
    _commit(db)
    return {"comment": serialize_request(db, row)["comments"][-1]}


@router.get("/contacts")
def list_contacts(client_id: str | None = None, actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    query = db.query(PbcContact).filter(PbcContact.firm_id == _firm_id(db, actor))
    if client_id:
        query = query.filter(PbcContact.client_id == client_id)
    return {"contacts": [serialize_contact(item) for item in query.order_by(PbcContact.name).all()]}


@router.post("/contacts", status_code=201)
def create_contact(payload: PbcContactCreate, actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    require_actor_role(actor, FIRM_MANAGE_ROLES)
    firm_id = _firm_id(db, actor)
    if payload.client_id and not db.query(Client.id).filter(Client.id == payload.client_id, Client.firm_id == firm_id).first():
        raise HTTPException(status_code=422, detail="Client not found in this firm")
    normalized_email = str(payload.email).lower()
    existing = db.query(PbcContact).filter(PbcContact.firm_id == firm_id, PbcContact.email == normalized_email).first()
    if existing:
        if existing.client_id and payload.client_id and str(existing.client_id) != payload.client_id:
            raise HTTPException(status_code=409, detail="This contact email belongs to a different client")
        existing.name = payload.name
        existing.client_id = existing.client_id or payload.client_id
        existing.active = True
        _commit(db)
        return serialize_contact(existing)
    row = PbcContact(firm_id=firm_id, client_id=payload.client_id, name=payload.name,
                     email=normalized_email, created_by_user_id=actor.id)
    db.add(row)
    try:
        _commit(db)
    except IntegrityError:
        raise HTTPException(status_code=409, detail="A PBC contact with this email already exists")
    return serialize_contact(row)


@router.put("/engagements/{engagement_id}/contacts/{contact_id}")
def assign_contact(engagement_id: str, contact_id: str, payload: PbcContactAssignment,
                   actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    require_actor_role(actor, FIRM_MANAGE_ROLES)
    engagement = get_engagement(db, _firm_id(db, actor), engagement_id)
    if payload.contact_id != contact_id:
        raise HTTPException(status_code=422, detail="Contact identifier mismatch")
    contact = db.query(PbcContact).filter(PbcContact.id == contact_id, PbcContact.firm_id == engagement.firm_id).first()
    if contact is None:
        raise HTTPException(status_code=404, detail="PBC contact not found")
    request_ids = _request_assignment_ids(db, engagement.id, payload.request_ids)
    membership = db.get(PbcEngagementContact, (engagement.id, contact.id))
    if membership is None:
        membership = PbcEngagementContact(engagement_id=engagement.id, contact_id=contact.id, role=payload.role)
        db.add(membership)
    else:
        membership.role = payload.role
    db.query(PbcRequestAssignment).filter(
        PbcRequestAssignment.contact_id == contact.id,
        PbcRequestAssignment.request_id.in_(db.query(PbcRequest.id).filter(PbcRequest.engagement_id == engagement.id)),
    ).delete(synchronize_session=False)
    for request_id in request_ids:
        db.add(PbcRequestAssignment(request_id=request_id, contact_id=contact.id))
    audit(db, engagement, "contact_assigned", actor_kind="firm", actor_id=actor.id,
          details={"contact_id": contact_id, "role": payload.role, "request_ids": [str(value) for value in request_ids]})
    _commit(db)
    return serialize_engagement(db, engagement, detail=True)


@router.delete("/engagements/{engagement_id}/contacts/{contact_id}", status_code=204)
def remove_contact(engagement_id: str, contact_id: str,
                   actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    require_actor_role(actor, FIRM_MANAGE_ROLES)
    engagement = get_engagement(db, _firm_id(db, actor), engagement_id)
    contact = db.query(PbcContact).filter(
        PbcContact.id == contact_id, PbcContact.firm_id == engagement.firm_id
    ).first()
    membership = db.get(PbcEngagementContact, (engagement.id, contact.id)) if contact else None
    if membership is None:
        raise HTTPException(status_code=404, detail="Contact is not assigned to this engagement")

    now = utcnow()
    db.query(PbcAccessToken).filter(
        PbcAccessToken.engagement_id == engagement.id,
        PbcAccessToken.contact_id == contact.id,
        PbcAccessToken.revoked_at.is_(None),
    ).update({PbcAccessToken.revoked_at: now}, synchronize_session=False)
    db.query(PbcPortalSession).filter(
        PbcPortalSession.engagement_id == engagement.id,
        PbcPortalSession.contact_id == contact.id,
        PbcPortalSession.revoked_at.is_(None),
    ).update({PbcPortalSession.revoked_at: now}, synchronize_session=False)
    db.query(PbcNotificationOutbox).filter(
        PbcNotificationOutbox.engagement_id == engagement.id,
        PbcNotificationOutbox.contact_id == contact.id,
        PbcNotificationOutbox.status.in_(["pending", "failed"]),
    ).delete(synchronize_session=False)
    db.query(PbcRequestAssignment).filter(
        PbcRequestAssignment.contact_id == contact.id,
        PbcRequestAssignment.request_id.in_(
            db.query(PbcRequest.id).filter(PbcRequest.engagement_id == engagement.id)
        ),
    ).delete(synchronize_session=False)
    db.delete(membership)
    audit(db, engagement, "contact_removed", actor_kind="firm", actor_id=actor.id,
          details={"contact_id": str(contact.id)})
    _commit(db)
    return Response(status_code=204)


@router.post("/engagements/{engagement_id}/access-links")
def issue_access_link(engagement_id: str, payload: PbcAccessLinkCreate,
                      actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    require_actor_role(actor, FIRM_MANAGE_ROLES)
    engagement = get_engagement(db, _firm_id(db, actor), engagement_id)
    contact = db.query(PbcContact).filter(PbcContact.id == payload.contact_id, PbcContact.firm_id == engagement.firm_id).first()
    if contact is None:
        raise HTTPException(status_code=404, detail="PBC contact not found")
    row, raw = create_access_token(db, engagement, contact, purpose=payload.purpose, request_ids=payload.request_ids,
                                   expires_in_days=payload.expires_in_days, actor_user_id=actor.id)
    _commit(db)
    url = f"{frontend_base_url()}/pbc/access?token={quote(raw)}"
    delivered = None
    if payload.send_email:
        delivered = email_service.send_email(
            contact.email, f"Secure PBC request: {engagement.name}",
            f"Hello {contact.name},\n\nUse this secure link to access {engagement.name}:\n{url}\n\nThe link expires on {row.expires_at.date()}.\n\n— CPAAutomation",
        )
    return {"id": str(row.id), "url": url, "expires_at": row.expires_at, "email_delivered": delivered}


@router.get("/templates")
def list_templates(actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    firm_id = _firm_id(db, actor)
    ensure_default_template(db, firm_id, actor.id)
    _commit(db)
    rows = db.query(PbcTemplate).filter(PbcTemplate.firm_id == firm_id).order_by(PbcTemplate.name).all()
    return {"templates": [{"id": str(row.id), "name": row.name, "description": row.description,
                            "engagement_type": row.engagement_type, "items": row.items or [], "updated_at": row.updated_at} for row in rows]}


@router.post("/templates", status_code=201)
def create_template(payload: PbcTemplateCreate, actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    require_actor_role(actor, FIRM_MANAGE_ROLES)
    row = PbcTemplate(firm_id=_firm_id(db, actor), name=payload.name, description=payload.description,
                      engagement_type=payload.engagement_type, items=payload.items, created_by_user_id=actor.id)
    db.add(row)
    _commit(db)
    return {"id": str(row.id), **payload.model_dump()}


def _cell_date(value):
    if value is None or value == "":
        return None
    if hasattr(value, "date"):
        return value.date()
    try:
        return date.fromisoformat(str(value)[:10])
    except ValueError:
        raise HTTPException(status_code=422, detail=f"Invalid import date: {value}")


@router.post("/imports/preview")
async def preview_import(file: UploadFile = File(...), actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    require_actor_role(actor, FIRM_WRITE_ROLES)
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Import file exceeds 10 MB")
    rows = []
    if (file.filename or "").lower().endswith(".csv"):
        import csv
        reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
        source = list(reader)
    else:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            source = []
        else:
            headers = [str(value or "").strip() for value in values[0]]
            source = [dict(zip(headers, row)) for row in values[1:] if any(value not in (None, "") for value in row)]
    aliases = {
        "request_id": "request_number", "number": "request_number", "description": "title",
        "request": "title", "owner": "owner", "due": "due_date", "format": "expected_formats",
        "expected_format": "expected_formats", "source_id": "external_source_id",
    }
    for index, raw in enumerate(source[:5000]):
        normalized = {aliases.get(str(key).strip().lower().replace(" ", "_"), str(key).strip().lower().replace(" ", "_")): value for key, value in raw.items()}
        title = str(normalized.get("title") or "").strip()
        if not title:
            raise HTTPException(status_code=422, detail=f"Import row {index + 2} is missing a description/title")
        formats = normalized.get("expected_formats") or ""
        rows.append({
            "request_number": str(normalized.get("request_number") or "").strip() or None,
            "title": title, "category": str(normalized.get("category") or "").strip() or None,
            "owner": str(normalized.get("owner") or "").strip() or None,
            "due_date": _cell_date(normalized.get("due_date")), "priority": str(normalized.get("priority") or "normal").lower(),
            "period_end": _cell_date(normalized.get("period_end")),
            "expected_filename": str(normalized.get("expected_filename") or "").strip() or None,
            "expected_formats": [part.strip().lower().lstrip(".") for part in re.split(r"[,;/]", str(formats)) if part.strip()],
            "gl_account": str(normalized.get("gl_account") or "").strip() or None,
            "gl_balance": str(normalized.get("gl_balance") or "").strip() or None,
            "external_source_id": str(normalized.get("external_source_id") or "").strip() or None,
        })
    return {"filename": file.filename, "rows": rows, "row_count": len(rows)}


@router.post("/engagements/{engagement_id}/imports")
def commit_import(engagement_id: str, payload: PbcImportCommit, actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    require_actor_role(actor, FIRM_WRITE_ROLES)
    engagement = get_engagement(db, _firm_id(db, actor), engagement_id)
    if engagement.status != "draft":
        raise HTTPException(status_code=409, detail="Import into a draft engagement only")
    existing_numbers = {value[0] for value in db.query(PbcRequest.request_number).filter(PbcRequest.engagement_id == engagement.id)}
    existing_sources = {value[0] for value in db.query(PbcRequest.external_source_id).filter(
        PbcRequest.engagement_id == engagement.id, PbcRequest.external_source_id.isnot(None))}
    users = db.query(User).filter(User.firm_id == engagement.firm_id).all()
    users_by_label = {(user.display_name or user.email).strip().lower(): user.id for user in users}
    created = skipped = 0
    for index, item in enumerate(payload.rows):
        if item.external_source_id and item.external_source_id in existing_sources:
            skipped += 1
            continue
        number = item.request_number or _request_number(db, engagement.id)
        if number in existing_numbers:
            raise HTTPException(status_code=409, detail=f"Duplicate request number: {number}")
        priority = item.priority if item.priority in {"low", "normal", "high", "urgent"} else "normal"
        db.add(PbcRequest(
            engagement_id=engagement.id, firm_id=engagement.firm_id, request_number=number,
            sort_order=db.query(PbcRequest).filter(PbcRequest.engagement_id == engagement.id).count() + index,
            category=item.category, title=item.title, period_end=item.period_end, priority=priority,
            due_date=item.due_date or engagement.due_date,
            owner_user_id=users_by_label.get((item.owner or "").lower(), actor.id),
            expected_filename=item.expected_filename, expected_formats=item.expected_formats,
            gl_account=item.gl_account, gl_balance=item.gl_balance, external_source_id=item.external_source_id,
            created_by_user_id=actor.id,
        ))
        existing_numbers.add(number)
        if item.external_source_id:
            existing_sources.add(item.external_source_id)
        created += 1
    audit(db, engagement, "request_list_imported", actor_kind="firm", actor_id=actor.id,
          details={"created": created, "skipped": skipped})
    _commit(db)
    return {"created": created, "skipped": skipped, "engagement": serialize_engagement(db, engagement, detail=True)}


@router.get("/engagements/{engagement_id}/export.xlsx")
def export_tracker(engagement_id: str, actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    engagement = get_engagement(db, _firm_id(db, actor), engagement_id)
    data = build_tracker_xlsx(db, engagement)
    return StreamingResponse(io.BytesIO(data), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="{engagement.name[:80]}-pbc.xlsx"'})


@router.get("/engagements/{engagement_id}/package.zip")
async def export_package(engagement_id: str, actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    engagement = get_engagement(db, _firm_id(db, actor), engagement_id)
    storage = get_storage_service()
    evidence_files: list[tuple[str, bytes]] = []
    rows = (
        db.query(PbcDocument, PbcRequest)
        .join(PbcRequest, PbcRequest.id == PbcDocument.request_id)
        .filter(PbcRequest.engagement_id == engagement.id, PbcRequest.status != "waived", PbcDocument.state == "available")
        .order_by(PbcRequest.sort_order, PbcDocument.version)
        .all()
    )
    if sum(int(document.size_bytes) for document, _ in rows) > 1024 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Evidence package exceeds the 1 GB synchronous export limit")
    for document, request_row in rows:
        category = re.sub(r"[^A-Za-z0-9._-]+", "_", request_row.category or "Other")[:80]
        number = re.sub(r"[^A-Za-z0-9._-]+", "_", request_row.request_number)[:80]
        filename = PurePath(document.filename).name
        content = await asyncio.to_thread(storage.bucket.blob(document.object_name).download_as_bytes)
        evidence_files.append((f"evidence/{category}/{number}/v{document.version}_{filename}", content))
    data = build_metadata_zip(db, engagement, evidence_files)
    audit(db, engagement, "package_exported", actor_kind="firm", actor_id=actor.id, details={"bytes": len(data)})
    _commit(db)
    return StreamingResponse(io.BytesIO(data), media_type="application/zip",
                             headers={"Content-Disposition": f'attachment; filename="{engagement.name[:80]}-evidence-package.zip"'})


async def _initiate_upload(db: Session, request_row: PbcRequest, payload: PbcUploadInitiate, actor_kind: str, actor_id: str):
    filename, mime, size = _safe_upload(payload)
    version = (db.query(PbcDocument.version).filter(PbcDocument.request_id == request_row.id).order_by(PbcDocument.version.desc()).first() or (0,))[0] + 1
    document_id = uuid.uuid4()
    object_name = f"pbc/{request_row.firm_id}/{request_row.engagement_id}/{request_row.id}/{document_id}/{quote(filename, safe='._-')}"
    row = PbcDocument(id=document_id, request_id=request_row.id, firm_id=request_row.firm_id, object_name=object_name,
                      filename=filename, mime_type=mime, size_bytes=size, version=version,
                      uploaded_by_kind=actor_kind, uploaded_by_id=actor_id)
    db.add(row)
    db.flush()
    try:
        upload_url = await get_storage_service().generate_presigned_put_url(object_name, expiration_minutes=15, content_type=mime)
    except Exception as exc:
        raise HTTPException(status_code=503, detail="Secure object storage is unavailable") from exc
    return row, upload_url


async def _complete_upload(db: Session, document: PbcDocument, checksum: str | None, engagement: PbcEngagement, actor_kind: str, actor_id: str):
    if document.state != "initiated":
        raise HTTPException(status_code=409, detail="Upload is already completed or unavailable")
    await _verify_object(document)
    document.state = "quarantined"
    document.checksum_sha256 = checksum.lower() if checksum else None
    scan_status, detail, actual_checksum = await _scan_document(document)
    if checksum and not secrets.compare_digest(checksum.lower(), actual_checksum):
        scan_status, detail = "failed", "Uploaded file checksum does not match the declared checksum"
    document.scan_status, document.scan_detail = scan_status, detail
    document.checksum_sha256 = actual_checksum
    document.state = "available" if scan_status in {"clean", "skipped"} else "rejected" if scan_status == "infected" else "quarantined"
    document.completed_at = utcnow()
    audit(db, engagement, "document_uploaded", actor_kind=actor_kind, actor_id=actor_id, request_id=document.request_id,
          details={"document_id": str(document.id), "filename": document.filename, "version": document.version,
                   "scan_status": scan_status})
    return document


@router.post("/requests/{request_id}/files:initiate")
async def initiate_firm_upload(request_id: str, payload: PbcUploadInitiate,
                               actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    require_actor_role(actor, FIRM_WRITE_ROLES)
    request_row = get_pbc_request(db, _firm_id(db, actor), request_id)
    row, upload_url = await _initiate_upload(db, request_row, payload, "firm", actor.id)
    _commit(db)
    return {"document": serialize_document(row), "upload_url": upload_url, "content_type": row.mime_type}


@router.post("/files:complete")
async def complete_firm_upload(payload: PbcUploadComplete, actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    document = db.query(PbcDocument).filter(PbcDocument.id == payload.document_id, PbcDocument.firm_id == _firm_id(db, actor)).with_for_update().first()
    if document is None:
        raise HTTPException(status_code=404, detail="PBC document not found")
    request_row = get_pbc_request(db, document.firm_id, str(document.request_id))
    engagement = get_engagement(db, document.firm_id, str(request_row.engagement_id))
    await _complete_upload(db, document, payload.checksum_sha256, engagement, "firm", actor.id)
    _commit(db)
    _raise_for_failed_scan(document)
    return {"document": serialize_document(document)}


@router.get("/documents/{document_id}/download-url")
async def firm_download(document_id: str, actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    row = db.query(PbcDocument).filter(PbcDocument.id == document_id, PbcDocument.firm_id == _firm_id(db, actor),
                                       PbcDocument.state == "available").first()
    if row is None:
        raise HTTPException(status_code=404, detail="PBC document not found")
    url = await get_storage_service().generate_presigned_get_url(row.object_name, expiration_minutes=5, download_filename=row.filename)
    return {"url": url, "expires_in_seconds": 300}


@router.post("/engagements/{engagement_id}/ai/draft")
async def ai_draft(engagement_id: str, payload: PbcAiDraftRequest, actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    require_actor_role(actor, FIRM_WRITE_ROLES)
    engagement = get_engagement(db, _firm_id(db, actor), engagement_id)
    try:
        return await draft_request_list(db, actor.id, engagement, payload.instructions)
    except ValueError as exc:
        raise HTTPException(status_code=502, detail=str(exc))


@router.post("/engagements/{engagement_id}/ai/match")
async def ai_match(engagement_id: str, payload: PbcAiMatchRequest, actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    require_actor_role(actor, FIRM_WRITE_ROLES)
    engagement = get_engagement(db, _firm_id(db, actor), engagement_id)
    document = (
        db.query(PbcDocument)
        .join(PbcRequest, PbcRequest.id == PbcDocument.request_id)
        .filter(PbcDocument.id == payload.document_id, PbcDocument.firm_id == engagement.firm_id,
                PbcRequest.engagement_id == engagement.id)
        .first()
    )
    if document is None:
        raise HTTPException(status_code=404, detail="PBC document not found")
    return await match_document(db, actor.id, engagement, document)


@router.get("/engagements/{engagement_id}/ai/completeness")
def ai_completeness(engagement_id: str, actor: User = Depends(require_role(*READER_ROLES)), db: Session = Depends(get_db)):
    return completeness_flags(db, get_engagement(db, _firm_id(db, actor), engagement_id))


# ---------------------------------------------------------------------------
# External client portal
# ---------------------------------------------------------------------------


@portal_router.post("/exchange")
def portal_exchange(payload: PbcPortalExchange, request: Request, db: Session = Depends(get_db)):
    _rate_limit(request, "exchange", 10)
    session, raw_session, contact, engagement = exchange_access_token(db, payload.token)
    raw_csrf = secrets.token_urlsafe(32)
    session.csrf_hash = token_hash(raw_csrf)
    content = jsonable_encoder({
        "csrf_token": raw_csrf,
        "contact": serialize_contact(contact),
        "engagement": serialize_engagement(db, engagement),
    })
    _commit(db)
    response = JSONResponse(content)
    response.set_cookie(COOKIE_NAME, raw_session, max_age=12 * 60 * 60, httponly=True, secure=not is_local(),
                        samesite="lax", path="/api/pbc/portal")
    return response


@portal_router.get("/workspace")
def portal_workspace(request: Request, db: Session = Depends(get_db)):
    session, contact, engagement = _portal_context(db, request)
    visible = portal_request_ids(db, session)
    requests = db.query(PbcRequest).filter(PbcRequest.engagement_id == engagement.id).order_by(PbcRequest.sort_order).all()
    _commit(db)
    result = serialize_engagement(db, engagement)
    result["requests"] = [serialize_request(db, row, include_internal=False) for row in requests if str(row.id) in visible and row.status != "draft"]
    result["contact"] = serialize_contact(contact)
    settings = get_settings(db, engagement.firm_id)
    result["portal_brand"] = {"portal_name": settings.portal_name or "Client Portal", "logo_url": settings.logo_url}
    _commit(db)
    return result


@portal_router.post("/logout", status_code=204)
def portal_logout(request: Request, response: Response, x_pbc_csrf: str | None = Header(default=None), db: Session = Depends(get_db)):
    session, _, _ = _portal_context(db, request, x_pbc_csrf, mutate=True)
    session.revoked_at = utcnow()
    _commit(db)
    response.delete_cookie(COOKIE_NAME, path="/api/pbc/portal")


@portal_router.post("/requests/{request_id}/comments", status_code=201)
def portal_comment(request_id: str, payload: PbcCommentCreate, request: Request,
                   x_pbc_csrf: str | None = Header(default=None), db: Session = Depends(get_db)):
    session, contact, engagement = _portal_context(db, request, x_pbc_csrf, mutate=True)
    row = require_portal_request(db, session, request_id)
    comment = PbcComment(request_id=row.id, firm_id=row.firm_id, body=payload.body, visibility="client",
                         actor_kind="client", actor_id=str(contact.id), actor_name=contact.name)
    db.add(comment)
    audit(db, engagement, "comment_added", actor_kind="client", actor_id=str(contact.id), request_id=row.id,
          details={"visibility": "client"})
    _commit(db)
    return {"request": serialize_request(db, row, include_internal=False)}


@portal_router.post("/requests/{request_id}/submit")
def portal_submit(request_id: str, payload: PbcTransitionRequest, request: Request,
                  x_pbc_csrf: str | None = Header(default=None), db: Session = Depends(get_db)):
    session, contact, engagement = _portal_context(db, request, x_pbc_csrf, mutate=True)
    row = require_portal_request(db, session, request_id, lock=True)
    if payload.action != "submit":
        raise HTTPException(status_code=422, detail="The portal only supports submit transitions")
    transition_request(db, row, engagement, "submit", payload.reason, payload.expected_revision,
                       actor_kind="client", actor_id=str(contact.id))
    _commit(db)
    return {"request": serialize_request(db, row, include_internal=False)}


@portal_router.post("/requests/{request_id}/files:initiate")
async def portal_upload_init(request_id: str, payload: PbcUploadInitiate, request: Request,
                             x_pbc_csrf: str | None = Header(default=None), db: Session = Depends(get_db)):
    _rate_limit(request, "upload", 20)
    session, contact, _ = _portal_context(db, request, x_pbc_csrf, mutate=True)
    request_row = require_portal_request(db, session, request_id)
    if request_row.status not in {"open", "needs_changes"}:
        raise HTTPException(status_code=409, detail="This request is not accepting uploads")
    row, upload_url = await _initiate_upload(db, request_row, payload, "client", str(contact.id))
    _commit(db)
    return {"document": serialize_document(row), "upload_url": upload_url, "content_type": row.mime_type}


@portal_router.post("/files:complete")
async def portal_upload_complete(payload: PbcUploadComplete, request: Request,
                                 x_pbc_csrf: str | None = Header(default=None), db: Session = Depends(get_db)):
    session, contact, engagement = _portal_context(db, request, x_pbc_csrf, mutate=True)
    document = db.query(PbcDocument).filter(PbcDocument.id == payload.document_id).with_for_update().first()
    if document is None:
        raise HTTPException(status_code=404, detail="PBC document not found")
    require_portal_request(db, session, str(document.request_id))
    if document.uploaded_by_kind != "client" or document.uploaded_by_id != str(contact.id):
        raise HTTPException(status_code=403, detail="This upload belongs to another contributor")
    await _complete_upload(db, document, payload.checksum_sha256, engagement, "client", str(contact.id))
    _commit(db)
    _raise_for_failed_scan(document)
    return {"document": serialize_document(document)}


@portal_router.get("/documents/{document_id}/download-url")
async def portal_download(document_id: str, request: Request, db: Session = Depends(get_db)):
    session, _, _ = _portal_context(db, request)
    document = db.query(PbcDocument).filter(PbcDocument.id == document_id, PbcDocument.state == "available").first()
    if document is None:
        raise HTTPException(status_code=404, detail="PBC document not found")
    require_portal_request(db, session, str(document.request_id))
    url = await get_storage_service().generate_presigned_get_url(document.object_name, expiration_minutes=5, download_filename=document.filename)
    return {"url": url, "expires_in_seconds": 300}
