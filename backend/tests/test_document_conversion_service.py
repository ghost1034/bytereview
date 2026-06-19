from __future__ import annotations

import asyncio
import tempfile
import unittest
import zipfile
from unittest.mock import AsyncMock, patch

from openpyxl import Workbook

from inkwise.services.pdf_extract import ExtractedPage
from inkwise.services.source_normalizer import InkwiseSourceNormalizer, SourceNormalizationError
from services.document_conversion_service import (
    DOCX_MIME,
    PPTX_MIME,
    XLSX_MIME,
    DocumentConversionService,
    normalize_source_mime_type,
)
from services.page_counting_service import PageCountingService
from services.spreadsheet_extraction_service import SpreadsheetExtractionService


class DocumentConversionServiceTests(unittest.TestCase):
    def test_raises_clear_error_when_soffice_missing(self) -> None:
        service = DocumentConversionService()

        with tempfile.NamedTemporaryFile(suffix=".docx") as handle:
            with patch("services.document_conversion_service.shutil.which", return_value=None):
                with self.assertRaisesRegex(
                    RuntimeError,
                    "LibreOffice/soffice is not installed in this runtime",
                ):
                    asyncio.run(service.convert_docx_local_to_pdf_local(handle.name))

    def test_detects_office_mime_types_from_extension_when_browser_type_is_empty(self) -> None:
        self.assertEqual(normalize_source_mime_type("example.docx", "application/octet-stream"), DOCX_MIME)
        self.assertEqual(normalize_source_mime_type("deck.pptx", ""), PPTX_MIME)
        self.assertEqual(normalize_source_mime_type("workbook.xlsx", None), XLSX_MIME)

    def test_xlsx_is_allowed_for_pdf_conversion(self) -> None:
        # Reaching the soffice-missing error (rather than a ValueError) proves the
        # ".xlsx" extension passed the conversion allowlist.
        service = DocumentConversionService()
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as handle:
            with patch("services.document_conversion_service.shutil.which", return_value=None):
                with self.assertRaisesRegex(RuntimeError, "LibreOffice/soffice is not installed"):
                    asyncio.run(service.convert_xlsx_local_to_pdf_local(handle.name))

    def test_unsupported_extension_rejected_for_pdf_conversion(self) -> None:
        service = DocumentConversionService()
        with tempfile.NamedTemporaryFile(suffix=".txt") as handle:
            with self.assertRaisesRegex(ValueError, "Unsupported Office document for PDF conversion"):
                asyncio.run(service.convert_office_local_to_pdf_local(handle.name))

    def test_pptx_gcs_conversion_uses_pptx_converter(self) -> None:
        service = DocumentConversionService()
        storage = type("StorageStub", (), {})()
        storage.download_file = AsyncMock()
        storage.upload_file = AsyncMock()

        async def fake_convert(local_path: str, out_dir: str | None = None) -> str:
            output_path = f"{out_dir}/input.pdf"
            with open(output_path, "wb") as handle:
                handle.write(b"pdf")
            return output_path

        with patch.object(service, "convert_pptx_local_to_pdf_local", side_effect=fake_convert) as convert:
            output, size = asyncio.run(service.convert_pptx_gcs_to_pdf_gcs(storage, "in.pptx", "out.pdf"))

        self.assertEqual(output, "out.pdf")
        self.assertEqual(size, 3)
        storage.download_file.assert_awaited_once()
        storage.upload_file.assert_awaited_once()
        convert.assert_awaited_once()


class PageCountingServiceOfficeTests(unittest.TestCase):
    def test_counts_pptx_slides_from_content(self) -> None:
        with tempfile.NamedTemporaryFile(suffix=".pptx") as handle:
            with zipfile.ZipFile(handle.name, "w") as archive:
                archive.writestr("ppt/slides/slide1.xml", "<slide />")
                archive.writestr("ppt/slides/slide2.xml", "<slide />")
                archive.writestr("ppt/notesSlides/notesSlide1.xml", "<notes />")
            handle.seek(0)
            count = PageCountingService.count_pages_from_content(handle.read(), "deck.pptx")
        self.assertEqual(count, 2)

    def test_counts_non_empty_xlsx_sheets(self) -> None:
        workbook = Workbook()
        active = workbook.active
        active.title = "Data"
        active["A1"] = "Name"
        empty = workbook.create_sheet("Empty")
        self.assertEqual(empty.max_row, 1)

        with tempfile.NamedTemporaryFile(suffix=".xlsx") as handle:
            workbook.save(handle.name)
            count = PageCountingService.count_pages_from_file_path(handle.name, "workbook.xlsx")
        self.assertEqual(count, 1)


class SpreadsheetExtractionServiceTests(unittest.TestCase):
    def test_renders_sheet_names_values_and_formulas(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Trial Balance"
        worksheet["A1"] = "Account"
        worksheet["B1"] = "Amount"
        worksheet["A2"] = "Cash"
        worksheet["B2"] = 125
        worksheet["B3"] = "=SUM(B2:B2)"

        with tempfile.NamedTemporaryFile(suffix=".xlsx") as handle:
            workbook.save(handle.name)
            service = SpreadsheetExtractionService()
            rendered = service.render_xlsx_local_to_text(handle.name, filename="tb.xlsx")

        self.assertIn("Workbook: tb.xlsx", rendered)
        self.assertIn("Sheet 1: Trial Balance", rendered)
        self.assertIn("R1: Account | Amount", rendered)
        self.assertIn("R2: Cash | 125", rendered)
        self.assertIn("[formula: =SUM(B2:B2)]", rendered)

    def test_renders_read_only_sheet_when_dimensions_are_unsized(self) -> None:
        class Cell:
            def __init__(self, value):
                self.value = value

        class Worksheet:
            title = "Unsized Sheet"
            max_row = None
            max_column = None

            def __init__(self, rows):
                self._rows = rows

            def calculate_dimension(self, *args, **kwargs):
                raise ValueError("Worksheet is unsized")

            def iter_rows(self, **kwargs):
                return iter(self._rows)

        class WorkbookStub:
            def __init__(self, worksheet):
                self.worksheets = [worksheet]
                self._worksheet = worksheet

            def __getitem__(self, title):
                return self._worksheet

            def close(self):
                pass

        value_ws = Worksheet([
            [Cell("Date"), Cell("What")],
            [Cell("2026-03-29"), Cell("Taipei")],
        ])
        formula_ws = Worksheet([
            [Cell("Date"), Cell("What")],
            [Cell("2026-03-29"), Cell("Taipei")],
        ])

        with patch(
            "services.spreadsheet_extraction_service.load_workbook",
            side_effect=[WorkbookStub(value_ws), WorkbookStub(formula_ws)],
        ):
            service = SpreadsheetExtractionService()
            rendered = service.render_xlsx_local_to_text("/tmp/unsized.xlsx", filename="unsized.xlsx")

        self.assertIn("dimension=unknown", rendered)
        self.assertIn("rendered_rows=2", rendered)
        self.assertIn("R1: Date | What", rendered)
        self.assertIn("R2: 2026-03-29 | Taipei", rendered)

    def test_empty_workbook_raises_clear_error(self) -> None:
        workbook = Workbook()
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as handle:
            workbook.save(handle.name)
            service = SpreadsheetExtractionService()
            with self.assertRaisesRegex(ValueError, "no readable non-empty rows"):
                service.render_xlsx_local_to_text(handle.name, filename="empty.xlsx")


class InkwiseSourceNormalizerTests(unittest.TestCase):
    def test_wraps_missing_soffice_error_for_docx_sources(self) -> None:
        normalizer = InkwiseSourceNormalizer()
        converter = type("ConverterStub", (), {})()
        converter.convert_docx_local_to_pdf_local = AsyncMock(
            side_effect=RuntimeError("LibreOffice/soffice is not installed in this runtime (expected binary: soffice)")
        )

        with patch("inkwise.services.source_normalizer.get_document_conversion_service", return_value=converter):
            with self.assertRaisesRegex(
                SourceNormalizationError,
                "DOCX conversion failed: LibreOffice/soffice is not installed in this runtime",
            ):
                normalizer._normalize_docx(local_path="/tmp/example.docx", title="Example")

    def test_detect_mime_type_recognizes_office_documents(self) -> None:
        normalizer = InkwiseSourceNormalizer()
        self.assertEqual(normalizer._detect_mime_type(filename="deck.pptx", content_type=""), PPTX_MIME)
        self.assertEqual(normalizer._detect_mime_type(filename="book.xlsx", content_type=""), XLSX_MIME)
        self.assertEqual(normalizer._detect_mime_type(filename="upload.bin", content_type=PPTX_MIME), PPTX_MIME)
        self.assertEqual(normalizer._detect_mime_type(filename="upload.bin", content_type=XLSX_MIME), XLSX_MIME)

    def test_normalize_pptx_produces_pdf_canonical(self) -> None:
        normalizer = InkwiseSourceNormalizer()
        converter = type("ConverterStub", (), {})()
        converter.convert_pptx_local_to_pdf_local = AsyncMock(return_value="/tmp/deck.pdf")
        pages = [
            ExtractedPage(page_number=1, text="Slide one"),
            ExtractedPage(page_number=2, text="Slide two"),
        ]
        with (
            patch("inkwise.services.source_normalizer.get_document_conversion_service", return_value=converter),
            patch.object(
                normalizer,
                "_extract_pdf_pages_with_optional_ocr",
                return_value=(pages, "/tmp/deck.pdf", {"extraction_engine": "pymupdf"}),
            ),
        ):
            result = normalizer._normalize_pptx(local_path="/tmp/deck.pptx", title="Quarterly Deck")

        self.assertEqual(result.source_kind, "pptx")
        self.assertEqual(result.canonical_mime_type, "application/pdf")
        self.assertEqual(result.canonical_local_path, "/tmp/deck.pdf")
        self.assertEqual(result.metadata["normalization"], "pptx_to_pdf")
        self.assertEqual(result.metadata["page_count"], 2)
        self.assertEqual(len(result.text_blocks), 2)
        self.assertEqual({asset.kind for asset in result.assets}, {"original_pptx", "canonical_pdf"})

    def test_wraps_conversion_error_for_pptx_sources(self) -> None:
        normalizer = InkwiseSourceNormalizer()
        converter = type("ConverterStub", (), {})()
        converter.convert_pptx_local_to_pdf_local = AsyncMock(side_effect=RuntimeError("boom"))
        with patch("inkwise.services.source_normalizer.get_document_conversion_service", return_value=converter):
            with self.assertRaisesRegex(SourceNormalizationError, "PPTX conversion failed: boom"):
                normalizer._normalize_pptx(local_path="/tmp/example.pptx", title="Example")

    def test_normalize_xlsx_uses_workbook_text_with_pdf_canonical(self) -> None:
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Summary"
        summary["A1"] = "Account"
        summary["B1"] = "Amount"
        summary["A2"] = "Cash"
        summary["B2"] = 100
        detail = workbook.create_sheet("Detail")
        detail["A1"] = "Item"
        detail["A2"] = "Pencils"

        normalizer = InkwiseSourceNormalizer()
        converter = type("ConverterStub", (), {})()
        converter.convert_xlsx_local_to_pdf_local = AsyncMock(return_value="/tmp/book.pdf")
        pages = [
            ExtractedPage(page_number=1, text="rendered page one"),
            ExtractedPage(page_number=2, text="rendered page two"),
        ]
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as handle:
            workbook.save(handle.name)
            with (
                patch("inkwise.services.source_normalizer.get_document_conversion_service", return_value=converter),
                patch.object(
                    normalizer,
                    "_extract_pdf_pages_with_optional_ocr",
                    return_value=(pages, "/tmp/book.pdf", {"extraction_engine": "pymupdf"}),
                ),
            ):
                result = normalizer._normalize_xlsx(local_path=handle.name, title="book.xlsx")

        self.assertEqual(result.source_kind, "xlsx")
        self.assertEqual(result.canonical_mime_type, "application/pdf")
        self.assertEqual(result.metadata["normalization"], "xlsx_to_pdf_with_text")
        self.assertEqual(result.metadata["sheet_count"], 2)
        self.assertEqual(len(result.text_blocks), 2)
        # Embedded text comes from the openpyxl workbook render, not the converted PDF.
        self.assertIn("Account", result.text_blocks[0].text)
        self.assertIn("Cash | 100", result.text_blocks[0].text)
        self.assertIn("Item", result.text_blocks[1].text)
        self.assertNotIn("rendered page", result.text_blocks[0].text)
        # sheet count == pdf page count -> 1:1 page mapping for citations.
        self.assertEqual(result.text_blocks[0].page_number, 1)
        self.assertEqual(result.text_blocks[1].page_number, 2)
        self.assertEqual(result.text_blocks[0].meta["sheet_title"], "Summary")
        self.assertEqual(result.text_blocks[1].meta["sheet_title"], "Detail")
        self.assertEqual({asset.kind for asset in result.assets}, {"original_xlsx", "canonical_pdf"})

    def test_normalize_xlsx_empty_workbook_raises(self) -> None:
        workbook = Workbook()
        normalizer = InkwiseSourceNormalizer()
        converter = type("ConverterStub", (), {})()
        converter.convert_xlsx_local_to_pdf_local = AsyncMock(return_value="/tmp/empty.pdf")
        pages = [ExtractedPage(page_number=1, text="")]
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as handle:
            workbook.save(handle.name)
            with (
                patch("inkwise.services.source_normalizer.get_document_conversion_service", return_value=converter),
                patch.object(normalizer, "_extract_pdf_pages_with_optional_ocr", return_value=(pages, "/tmp/empty.pdf", {})),
            ):
                with self.assertRaisesRegex(SourceNormalizationError, "Spreadsheet has no readable content"):
                    normalizer._normalize_xlsx(local_path=handle.name, title="empty.xlsx")

    def test_split_xlsx_sheets_parses_headers(self) -> None:
        normalizer = InkwiseSourceNormalizer()
        rendered = (
            "Workbook: book.xlsx\n"
            "\nSheet 1: Summary (dimension=A1:B2, rendered_rows=2, rendered_columns=2)\n"
            "R1: Account | Amount\nR2: Cash | 100\n"
            "\nSheet 2: Detail (dimension=A1:A2, rendered_rows=2, rendered_columns=1)\n"
            "R1: Item\nR2: Pencils"
        )
        sections = normalizer._split_xlsx_sheets(rendered)
        self.assertEqual(len(sections), 2)
        self.assertEqual(sections[0]["sheet_index"], 1)
        self.assertEqual(sections[0]["sheet_title"], "Summary")
        self.assertIn("R1: Account | Amount", sections[0]["text"])
        self.assertEqual(sections[1]["sheet_index"], 2)
        self.assertEqual(sections[1]["sheet_title"], "Detail")
        self.assertIn("R2: Pencils", sections[1]["text"])


if __name__ == "__main__":
    unittest.main()
