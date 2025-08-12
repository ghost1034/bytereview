"""
ARQ worker configuration and task definitions for ByteReview
Background worker system for asynchronous job processing
"""
import os
import asyncio
import logging
import uuid
import mimetypes
import tempfile
import zipfile
import shutil
from typing import Dict, Any
from datetime import datetime, timezone
from pathlib import Path
from dotenv import load_dotenv

# Load environment variables first
load_dotenv()

import sys
import os
from pathlib import Path

# Add the backend directory to Python path
backend_dir = Path(__file__).parent.parent  # Go up to backend/ directory
sys.path.insert(0, str(backend_dir))

from arq import create_pool
from arq.connections import RedisSettings
from arq.cron import cron
from sqlalchemy.orm import Session
from core.database import db_config, get_db
from models.db_models import ExtractionTask, ExtractionResult, SourceFile, JobField, SystemPrompt, SourceFileToTask, ExtractionJob, DataType
from models.job import FileStatus
from services.ai_extraction_service import AIExtractionService
from services.gcs_service import get_storage_service
from services.google_service import google_service
import json
import io
from typing import List

logger = logging.getLogger(__name__)

# Redis configuration
REDIS_URL = os.getenv("REDIS_URL", "redis://localhost:6379")

async def process_extraction_task(ctx: Dict[str, Any], task_id: str, automation_run_id: str = None) -> Dict[str, Any]:
    """
    Process a single extraction task using AI
    This is the core background task for data extraction
    """
    logger.info(f"Processing extraction task: {task_id}")
    
    db = db_config.get_session()
    try:
        # Get the task from database
        task = db.query(ExtractionTask).filter(ExtractionTask.id == task_id).first()
        if not task:
            raise ValueError(f"Task {task_id} not found")
        
        # Get associated source files (ordered by document_order)
        source_files_query = db.query(SourceFile, SourceFileToTask.document_order).join(
            SourceFileToTask, SourceFile.id == SourceFileToTask.source_file_id
        ).filter(
            SourceFileToTask.task_id == task_id
        ).order_by(SourceFileToTask.document_order)
        
        source_files_with_order = source_files_query.all()
        source_files = [sf for sf, _ in source_files_with_order]
        
        # Update task status to processing
        task.status = "processing"
        db.commit()
        
        # Send SSE event for task started
        try:
            from services.sse_service import sse_manager
            await sse_manager.send_task_started(task.job_id, task_id)
        except Exception as e:
            logger.warning(f"Failed to send task_started SSE event: {e}")
        
        if not source_files:
            raise ValueError(f"No source files found for task {task_id}")
        
        # Get job fields (the snapshotted configuration)
        job_fields = db.query(JobField).filter(JobField.job_id == task.job_id).order_by(JobField.display_order).all()
        
        if not job_fields:
            raise ValueError(f"No job fields found for job {task.job_id}")
        
        # Get active system prompt
        system_prompt_record = db.query(SystemPrompt).filter(SystemPrompt.is_active == True).first()
        if not system_prompt_record:
            raise ValueError("No active system prompt found")
        
        # Get data types for JSON schema creation
        data_types = db.query(DataType).all()
        data_types_map = {
            dt.id: {
                "base_json_type": dt.base_json_type,
                "json_format": dt.json_format,
                "display_name": dt.display_name,
                "description": dt.description
            }
            for dt in data_types
        }
        
        # Convert job fields to FieldConfig format for AI service
        from models.extraction import FieldConfig
        field_configs = [
            FieldConfig(
                name=field.field_name,
                data_type=field.data_type_id,
                prompt=field.ai_prompt
            )
            for field in job_fields
        ]
        
        # Initialize services
        ai_service = AIExtractionService()
        storage_service = get_storage_service()
        
        # Download files from GCS and process them locally
        temp_dir = None
        files_data = []
        
        try:
            # Create temporary directory for downloaded files
            temp_dir = tempfile.mkdtemp(prefix=f"extraction_task_{task_id}_")
            logger.info(f"Created temporary directory for extraction: {temp_dir}")
            
            for source_file in source_files:
                logger.info(f"Downloading file for processing: {source_file.original_filename}")
                
                # Download file from GCS to temporary location
                file_extension = os.path.splitext(source_file.original_filename)[1]
                temp_file_path = os.path.join(temp_dir, f"{uuid.uuid4()}{file_extension}")
                
                await storage_service.download_file(source_file.gcs_object_name, temp_file_path)
                logger.info(f"Downloaded {source_file.original_filename} to {temp_file_path}")
                
                # Read file content
                with open(temp_file_path, 'rb') as f:
                    file_content = f.read()
                
                files_data.append({
                    'filename': source_file.original_filename,
                    'content': file_content
                })
                
                logger.info(f"Prepared file data for {source_file.original_filename}, size: {len(file_content)} bytes")
            
            # Process files using downloaded content instead of GCS URIs
            logger.info(f"Processing {len(files_data)} files with AI using downloaded content")
            
            # Extract data using AI service with file content
            logger.info(f"Using processing mode: {task.processing_mode}")
            extraction_result = await ai_service.extract_data_from_files(
                files_data,
                field_configs,
                data_types_map,
                system_prompt_record.template_text,
                processed_files=source_files,  # Pass source files for metadata
                processing_mode=task.processing_mode  # Pass processing mode for routing
            )
            
        finally:
            # Clean up temporary directory
            if temp_dir and os.path.exists(temp_dir):
                try:
                    shutil.rmtree(temp_dir)
                    logger.info(f"Cleaned up temporary directory: {temp_dir}")
                except Exception as e:
                    logger.warning(f"Failed to clean up temporary directory {temp_dir}: {e}")
        
        # Process the extraction result based on processing mode
        if not extraction_result.success:
            raise ValueError(f"AI extraction failed: {extraction_result.error}")
        
        # Convert AI service result to new array-based format with column snapshot
        if extraction_result.data and isinstance(extraction_result.data, list):
            # Get field order from job configuration (snapshot at extraction time)
            field_order = [field.field_name for field in job_fields]
            
            # Convert object-based results to array-based results
            results_arrays = []
            for result_obj in extraction_result.data:
                if isinstance(result_obj, dict):
                    # Convert dict to array using field order
                    result_array = []
                    for field_name in field_order:
                        result_array.append(result_obj.get(field_name))
                    results_arrays.append(result_array)
                else:
                    # Fallback for unexpected format
                    results_arrays.append(result_obj)
            
            # Create new array-based format with column snapshot
            final_result = {
                "results": results_arrays,
                "columns": field_order
            }
        else:
            # Fallback for unexpected format
            final_result = {
                "results": [],
                "columns": [field.field_name for field in job_fields]
            }
        
        # Save results to database
        extraction_result = ExtractionResult(
            task_id=task_id,
            extracted_data=final_result
        )
        db.add(extraction_result)
        
        # Update task status
        task.status = "completed"
        task.processed_at = datetime.now(timezone.utc)
        
        # Record usage for billing (count pages processed)
        try:
            await _record_usage_for_task(db, task, source_files)
        except Exception as e:
            logger.error(f"Failed to record usage for billing: {e}")
            # Don't fail the task for billing errors
        
        # Increment job-level task completion counter
        try:
            from services.job_service import JobService
            job_service = JobService()
            automation_run_id = ctx.get('automation_run_id')
            await job_service.increment_task_completion(task.job_id, success=True, automation_run_id=automation_run_id)
            logger.info(f"Incremented task completion counter for job {task.job_id}")
        except Exception as e:
            logger.error(f"Failed to increment task completion counter: {e}")
        
        
        # Send SSE event for task completed
        try:
            from services.sse_service import sse_manager
            await sse_manager.send_task_completed(task.job_id, task_id, final_result)
        except Exception as e:
            logger.warning(f"Failed to send task_completed SSE event: {e}")
        
        db.commit()
        
        logger.info(f"Successfully completed extraction task: {task_id}")
        return {"success": True, "task_id": task_id}
        
    except Exception as e:
        logger.error(f"Error processing extraction task {task_id}: {e}")
        
        # Update task status to failed only if task exists
        if 'task' in locals() and task is not None:
            task.status = "failed"
            task.error_message = str(e)
            
            # Increment job-level task failure counter
            try:
                from services.job_service import JobService
                job_service = JobService()
                automation_run_id = ctx.get('automation_run_id')
                await job_service.increment_task_completion(task.job_id, success=False, automation_run_id=automation_run_id)
                logger.info(f"Incremented task failure counter for job {task.job_id}")
            except Exception as e:
                logger.error(f"Failed to increment task failure counter: {e}")
            
            db.commit()
            
            # Send SSE event for task failed
            try:
                from services.sse_service import sse_manager
                await sse_manager.send_task_failed(task.job_id, task_id, str(e))
            except Exception as e:
                logger.warning(f"Failed to send task_failed SSE event: {e}")
        else:
            logger.warning(f"Task {task_id} not found in database - likely stale queue item")
        
        # Don't re-raise for missing tasks to avoid infinite retries
        if "not found" in str(e):
            logger.info(f"Skipping missing task {task_id} to prevent retries")
            return {"success": False, "task_id": task_id, "error": "Task not found"}
        
        raise
    finally:
        db.close()

async def unpack_zip_file_task(ctx: Dict[str, Any], source_file_id: str, automation_run_id: str = None) -> Dict[str, Any]:
    """
    Unpack a ZIP file and register its contents as individual source files
    This task runs in the high-memory ZIP worker pool
    """
    logger.info(f"Unpacking ZIP file: {source_file_id}")
    
    db = db_config.get_session()
    try:
        # Get the ZIP source file from database
        zip_file = db.query(SourceFile).filter(SourceFile.id == source_file_id).first()
        if not zip_file:
            raise ValueError(f"Source file {source_file_id} not found")
        
        # Update status to indicate unpacking has started
        zip_file.status = FileStatus.UNPACKING.value
        db.commit()
        
        # Initialize services
        storage_service = get_storage_service()
        
        temp_dir = None
        files_extracted = 0
        
        try:
            # Create temporary directory
            temp_dir = tempfile.mkdtemp(prefix=f"zip_extract_{source_file_id}_")
            logger.info(f"Created temporary directory: {temp_dir}")
            
            # Download ZIP file from GCS to temporary location
            zip_temp_path = os.path.join(temp_dir, "archive.zip")
            await storage_service.download_file(zip_file.gcs_object_name, zip_temp_path)
            logger.info(f"Downloaded ZIP file to: {zip_temp_path}")
            
            # Extract ZIP contents
            extract_dir = os.path.join(temp_dir, "extracted")
            os.makedirs(extract_dir, exist_ok=True)
            
            with zipfile.ZipFile(zip_temp_path, 'r') as zip_ref:
                zip_ref.extractall(extract_dir)
                logger.info(f"Extracted ZIP contents to: {extract_dir}")
            
            # Process extracted files
            for root, dirs, files in os.walk(extract_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    
                    # Skip hidden files and directories
                    if file.startswith('.'):
                        continue
                    
                    # Calculate relative path from extraction root
                    rel_path = os.path.relpath(file_path, extract_dir)
                    
                    # Normalize path separators
                    from services.gcs_service import normalize_path
                    normalized_path = normalize_path(rel_path)
                    
                    # Generate new GCS object name for extracted file
                    job_id = zip_file.job_id
                    file_extension = os.path.splitext(file)[1]
                    new_gcs_name = f"jobs/{job_id}/extracted/{uuid.uuid4()}{file_extension}"
                    
                    # Upload extracted file to GCS
                    await storage_service.upload_file(file_path, new_gcs_name)
                    logger.info(f"Uploaded extracted file: {normalized_path} -> {new_gcs_name}")
                    
                    # Get file size
                    file_size = os.path.getsize(file_path)
                    
                    # Determine MIME type
                    mime_type, _ = mimetypes.guess_type(file)
                    if not mime_type:
                        mime_type = "application/octet-stream"
                    
                    # Count pages in the extracted file
                    from services.page_counting_service import page_counting_service
                    with open(file_path, 'rb') as f:
                        file_content = f.read()
                    page_count = page_counting_service.count_pages_from_content(file_content, file)
                    
                    # Create new SourceFile record for extracted file
                    extracted_file = SourceFile(
                        job_id=job_id,
                        original_filename=file,
                        original_path=normalized_path,
                        gcs_object_name=new_gcs_name,
                        file_type=mime_type,
                        file_size_bytes=file_size,
                        page_count=page_count,
                        status=FileStatus.UPLOADED.value  # Mark as uploaded since it's ready for processing
                    )
                    
                    db.add(extracted_file)
                    files_extracted += 1
            
            # Update original ZIP file status to "unpacked"
            zip_file.status = FileStatus.UNPACKED.value
            db.commit()
            
            # Send SSE events for extracted files
            try:
                from services.sse_service import sse_manager
            except ImportError:
                # Fallback for import issues
                import importlib.util
                spec = importlib.util.spec_from_file_location("sse_service", backend_dir / "services" / "sse_service.py")
                sse_module = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(sse_module)
                sse_manager = sse_module.sse_manager
            
            # Query the extracted files that were just added to the database
            extracted_files = db.query(SourceFile).filter(
                SourceFile.job_id == zip_file.job_id,
                SourceFile.status == "uploaded",
                SourceFile.id != zip_file.id  # Exclude the original ZIP file
            ).all()
            
            # Convert extracted files to dict format for SSE
            files_data = []
            for extracted_file in extracted_files:
                files_data.append({
                    "id": str(extracted_file.id),
                    "filename": extracted_file.original_filename,
                    "original_path": extracted_file.original_path,
                    "file_type": extracted_file.file_type,
                    "file_size": extracted_file.file_size_bytes,
                    "status": extracted_file.status
                })
            
            logger.info(f"Sending files_extracted event for {len(files_data)} files")
            await sse_manager.send_files_extracted(zip_file.job_id, files_data)
            logger.info(f"Sending file_status_changed event for ZIP file")
            await sse_manager.send_file_status_changed(zip_file.job_id, str(zip_file.id), "unpacked")
            logger.info(f"SSE events sent successfully for job {zip_file.job_id}")
            
            logger.info(f"Successfully unpacked ZIP file {source_file_id}: {files_extracted} files extracted")
            
            # If this is part of an automation, count the ZIP file as processed now that it's unpacked
            if automation_run_id:
                from services.google_service import google_service
                await google_service._update_automation_import_tracking(
                    db, automation_run_id, 
                    processed=1
                )
            
        finally:
            # Clean up temporary directory
            if temp_dir and os.path.exists(temp_dir):
                try:
                    shutil.rmtree(temp_dir)
                    logger.info(f"Cleaned up temporary directory: {temp_dir}")
                except Exception as e:
                    logger.warning(f"Failed to clean up temporary directory {temp_dir}: {e}")
        
        return {
            "success": True, 
            "source_file_id": source_file_id, 
            "files_extracted": files_extracted
        }
        
    except Exception as e:
        logger.error(f"Error unpacking ZIP file {source_file_id}: {e}")
        
        # Update status to failed
        if 'zip_file' in locals():
            zip_file.status = FileStatus.FAILED.value
            db.commit()
            
            # If this is part of an automation, count the ZIP file processing as failed
            if automation_run_id:
                from services.google_service import google_service
                await google_service._update_automation_import_tracking(
                    db, automation_run_id, 
                    processing_failed=1
                )
                logger.info(f"Marked ZIP processing as failed for automation run {automation_run_id}")
            
            # Send SSE event for extraction failure
            from services.sse_service import sse_manager
            try:
                from services.sse_service import sse_manager
                await sse_manager.send_extraction_failed(zip_file.job_id, str(zip_file.id), str(e))
            except ImportError as import_err:
                logger.error(f"Could not import SSE service for error notification: {import_err}")
        
        raise
    finally:
        db.close()

async def run_abandoned_cleanup(ctx: Dict[str, Any]) -> Dict[str, Any]:
    """
    Clean up abandoned jobs that were never started
    """
    logger.info("Running abandoned job cleanup")
    
    # TODO: Implement cleanup logic
    # Find jobs in 'pending_configuration' status older than X hours
    # Delete their files from GCS and remove from database
    
    return {"success": True, "jobs_cleaned": 0}

async def run_opt_out_cleanup(ctx: Dict[str, Any]) -> Dict[str, Any]:
    """
    Clean up data for users who opted out of data persistence
    """
    logger.info("Running opt-out data cleanup")
    
    # TODO: Implement cleanup logic
    # Find completed jobs where persist_data=false and older than grace period
    # Delete their files from GCS and remove from database
    
    return {"success": True, "jobs_cleaned": 0}

async def run_artifact_cleanup(ctx: Dict[str, Any]) -> Dict[str, Any]:
    """
    Clean up unpacked ZIP artifacts from GCS
    """
    logger.info("Running artifact cleanup")
    
    # TODO: Implement cleanup logic
    # Find source files with status='unpacked' (old ZIP files)
    # Delete their GCS objects
    
    return {"success": True, "artifacts_cleaned": 0}

async def run_free_user_period_reset(ctx: Dict[str, Any]) -> Dict[str, Any]:
    """Reset billing periods for free users at month boundaries"""
    logger.info("Starting free user period reset")
    
    db = db_config.get_session()
    try:
        from services.billing_service import get_billing_service
        from datetime import datetime, timezone, timedelta
        from models.db_models import BillingAccount, UsageCounter
        
        billing_service = get_billing_service(db)
        now = datetime.now(timezone.utc)
        
        # Find free users whose periods have expired
        expired_accounts = db.query(BillingAccount).filter(
            BillingAccount.plan_code == 'free',
            BillingAccount.current_period_end < now
        ).all()
        
        updated_count = 0
        for account in expired_accounts:
            try:
                # Calculate new period boundaries (current month)
                period_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                period_end = (period_start + timedelta(days=32)).replace(day=1) - timedelta(seconds=1)
                
                # Update billing account period
                account.current_period_start = period_start
                account.current_period_end = period_end
                
                # Create new usage counter for the new period
                new_counter = UsageCounter(
                    user_id=account.user_id,
                    period_start=period_start,
                    period_end=period_end,
                    pages_total=0
                )
                db.merge(new_counter)  # Use merge to handle conflicts
                
                updated_count += 1
                logger.info(f"Reset period for free user {account.user_id}: {period_start} to {period_end}")
                
            except Exception as e:
                logger.error(f"Failed to reset period for user {account.user_id}: {e}")
                continue
        
        db.commit()
        
        logger.info(f"Free user period reset completed: {updated_count} accounts updated")
        return {
            "success": True, 
            "message": f"Period reset completed for {updated_count} free users",
            "updated_count": updated_count
        }
        
    except Exception as e:
        db.rollback()
        logger.error(f"Free user period reset failed: {e}")
        return {"success": False, "error": str(e)}
    finally:
        db.close()

async def run_stripe_usage_reconciliation(ctx: Dict[str, Any]) -> Dict[str, Any]:
    """Retry failed Stripe usage reports and reconcile usage data"""
    logger.info("Starting Stripe usage reconciliation")
    
    db = db_config.get_session()
    try:
        from services.billing_service import get_billing_service
        from datetime import datetime, timezone, timedelta
        from models.db_models import BillingAccount, UsageEvent
        
        # Find usage events that failed to report to Stripe
        unreported_events = db.query(UsageEvent).filter(
            UsageEvent.stripe_reported == False,
            UsageEvent.occurred_at > datetime.now(timezone.utc) - timedelta(days=7)  # Only retry recent events
        ).all()
        
        retry_count = 0
        success_count = 0
        
        for event in unreported_events:
            try:
                # Get billing account to check if user is on paid plan
                billing_account = db.query(BillingAccount).filter(
                    BillingAccount.user_id == event.user_id
                ).first()
                
                if not billing_account or billing_account.plan_code == 'free':
                    # Mark as reported for free users (no Stripe reporting needed)
                    event.stripe_reported = True
                    continue
                
                if not billing_account.stripe_customer_id:
                    logger.warning(f"No Stripe customer ID for paid user {event.user_id}")
                    continue
                
                # Retry Stripe reporting
                billing_service = get_billing_service(db)
                billing_service._report_usage_to_stripe_async(
                    event.user_id, 
                    event.pages, 
                    str(event.id)
                )
                
                retry_count += 1
                success_count += 1
                logger.info(f"Successfully retried Stripe reporting for event {event.id}")
                
            except Exception as e:
                logger.error(f"Failed to retry Stripe reporting for event {event.id}: {e}")
                retry_count += 1
                continue
        
        db.commit()
        
        logger.info(f"Stripe usage reconciliation completed: {success_count}/{retry_count} events processed")
        return {
            "success": True,
            "message": f"Reconciliation completed: {success_count}/{retry_count} events processed",
            "retry_count": retry_count,
            "success_count": success_count
        }
        
    except Exception as e:
        db.rollback()
        logger.error(f"Stripe usage reconciliation failed: {e}")
        return {"success": False, "error": str(e)}
    finally:
        db.close()

async def run_usage_counter_cleanup(ctx: Dict[str, Any]) -> Dict[str, Any]:
    """Clean up old usage counters to prevent database bloat"""
    logger.info("Starting usage counter cleanup")
    
    db = db_config.get_session()
    try:
        from datetime import datetime, timezone, timedelta
        from models.db_models import UsageCounter
        
        # Remove usage counters older than 13 months (keep 1 year + current month)
        cutoff_date = datetime.now(timezone.utc) - timedelta(days=395)  # ~13 months
        
        deleted_count = db.query(UsageCounter).filter(
            UsageCounter.period_start < cutoff_date
        ).delete()
        
        db.commit()
        
        logger.info(f"Usage counter cleanup completed: {deleted_count} old counters removed")
        return {
            "success": True,
            "message": f"Cleanup completed: {deleted_count} old usage counters removed",
            "deleted_count": deleted_count
        }
        
    except Exception as e:
        db.rollback()
        logger.error(f"Usage counter cleanup failed: {e}")
        return {"success": False, "error": str(e)}
    finally:
        db.close()

# ===================================================================
# Scheduled/Cron Task Wrappers
# ===================================================================

async def schedule_free_user_period_reset(ctx: Dict[str, Any]) -> Dict[str, Any]:
    """
    Scheduled task to reset free user billing periods
    Runs daily at 00:30 UTC to catch month rollovers
    """
    logger.info("Scheduled: Free user period reset")
    try:
        result = await run_free_user_period_reset(ctx)
        logger.info(f"Scheduled free user period reset completed: {result}")
        return result
    except Exception as e:
        logger.error(f"Scheduled free user period reset failed: {e}")
        return {"success": False, "error": str(e)}

async def schedule_stripe_usage_reconciliation(ctx: Dict[str, Any]) -> Dict[str, Any]:
    """
    Scheduled task to retry failed Stripe usage reports
    Runs every 2 hours to ensure usage is properly reported
    """
    logger.info("Scheduled: Stripe usage reconciliation")
    try:
        result = await run_stripe_usage_reconciliation(ctx)
        logger.info(f"Scheduled Stripe reconciliation completed: {result}")
        return result
    except Exception as e:
        logger.error(f"Scheduled Stripe reconciliation failed: {e}")
        return {"success": False, "error": str(e)}

async def schedule_usage_counter_cleanup(ctx: Dict[str, Any]) -> Dict[str, Any]:
    """
    Scheduled task to clean up old usage counters
    Runs weekly on Sundays at 02:00 UTC
    """
    logger.info("Scheduled: Usage counter cleanup")
    try:
        result = await run_usage_counter_cleanup(ctx)
        logger.info(f"Scheduled usage counter cleanup completed: {result}")
        return result
    except Exception as e:
        logger.error(f"Scheduled usage counter cleanup failed: {e}")
        return {"success": False, "error": str(e)}

async def schedule_abandoned_cleanup(ctx: Dict[str, Any]) -> Dict[str, Any]:
    """
    Scheduled task to clean up abandoned jobs
    Runs daily at 01:00 UTC
    """
    logger.info("Scheduled: Abandoned job cleanup")
    try:
        result = await run_abandoned_cleanup(ctx)
        logger.info(f"Scheduled abandoned cleanup completed: {result}")
        return result
    except Exception as e:
        logger.error(f"Scheduled abandoned cleanup failed: {e}")
        return {"success": False, "error": str(e)}

async def schedule_artifact_cleanup(ctx: Dict[str, Any]) -> Dict[str, Any]:
    """
    Scheduled task to clean up old artifacts
    Runs daily at 03:00 UTC
    """
    logger.info("Scheduled: Artifact cleanup")
    try:
        result = await run_artifact_cleanup(ctx)
        logger.info(f"Scheduled artifact cleanup completed: {result}")
        return result
    except Exception as e:
        logger.error(f"Scheduled artifact cleanup failed: {e}")
        return {"success": False, "error": str(e)}

async def schedule_opt_out_cleanup(ctx: Dict[str, Any]) -> Dict[str, Any]:
    """
    Scheduled task to clean up opt-out user data
    Runs weekly on Saturdays at 04:00 UTC
    """
    logger.info("Scheduled: Opt-out data cleanup")
    try:
        result = await run_opt_out_cleanup(ctx)
        logger.info(f"Scheduled opt-out cleanup completed: {result}")
        return result
    except Exception as e:
        logger.error(f"Scheduled opt-out cleanup failed: {e}")
        return {"success": False, "error": str(e)}

# ARQ worker settings
class WorkerSettings:
    """ARQ worker configuration for AI extraction tasks (default queue)"""
    
    redis_settings = RedisSettings.from_dsn(REDIS_URL)
    
    # Task functions that the worker can execute
    functions = [
        process_extraction_task,
    ]
    
    # Worker configuration for AI tasks (low memory, high concurrency)
    max_jobs = 5  # Reduce concurrent jobs to prevent Redis overload
    job_timeout = 300  # 5 minutes timeout per job
    keep_result = 3600  # Keep results for 1 hour
    
    # Logging
    log_results = True

class ZipWorkerSettings:
    """ARQ worker configuration for ZIP unpacking tasks (zip_queue)"""
    
    redis_settings = RedisSettings.from_dsn(REDIS_URL)
    queue_name = "zip_queue"  # Dedicated queue for ZIP tasks
    
    # Task functions for ZIP processing
    functions = [
        unpack_zip_file_task,
    ]
    
    # Worker configuration for ZIP tasks (high memory, low concurrency)
    max_jobs = 1  # Low concurrency to prevent memory issues
    job_timeout = 1800  # 30 minutes timeout for large ZIP files
    keep_result = 3600  # Keep results for 1 hour
    
    # Logging
    log_results = True

# ImportWorkerSettings will be defined after the functions

# ===================================================================
# Import Worker Functions (Drive, Gmail)
# ===================================================================

async def import_startup(ctx: Dict[str, Any]) -> None:
    """Initialize import worker resources"""
    logger.info("Import worker starting up...")

async def import_shutdown(ctx: Dict[str, Any]) -> None:
    """Cleanup import worker resources"""
    logger.info("Import worker shutting down...")

async def import_drive_files(
    ctx: Dict[str, Any],
    job_id: str, 
    user_id: str, 
    drive_file_ids: List[str]
) -> Dict[str, Any]:
    """
    Import files from Google Drive (manual imports)
    
    Args:
        job_id: Extraction job ID
        user_id: User ID for OAuth credentials
        drive_file_ids: List of Google Drive file IDs to import
        
    Returns:
        Dict with import results and status
    """
    logger.info(f"Starting Drive import for job {job_id}, {len(drive_file_ids)} files")
    
    with next(get_db()) as db:
        try:
            # Verify job exists and user has access
            job = db.query(ExtractionJob).filter(
                ExtractionJob.id == job_id,
                ExtractionJob.user_id == user_id
            ).first()
            
            if not job:
                raise ValueError(f"Job {job_id} not found or access denied")
            
            # Use service layer for actual import logic
            from services.google_service import google_service
            import_result = await google_service.import_drive_files(db, job_id, user_id, drive_file_ids)
            
            # Convert service result to worker result format
            results = {
                'job_id': job_id,
                'total_files': import_result.get('total', len(drive_file_ids)),
                'successful': import_result.get('successful', 0),
                'failed': import_result.get('failed', 0),
                'errors': import_result.get('errors', [])
            }
            
            # Update job status
            if results['failed'] == 0:
                logger.info(f"Drive import completed successfully for job {job_id}")
            else:
                logger.warning(f"Drive import completed with {results['failed']} failures for job {job_id}")
            
            return results
            
        except Exception as e:
            logger.error(f"Drive import failed for job {job_id}: {e}")
            raise

async def import_gmail_attachments(
    ctx: Dict[str, Any],
    job_id: str,
    user_id: str,
    attachment_data: List[Dict[str, str]],
    automation_run_id: str = None
) -> Dict[str, Any]:
    """
    Import attachments from Gmail (manual imports)
    
    Args:
        job_id: Extraction job ID
        user_id: User ID for OAuth credentials
        attachment_data: List of dicts with messageId, attachmentId, filename
        
    Returns:
        Dict with import results and status
    """
    logger.info(f"Starting Gmail import for job {job_id}, {len(attachment_data)} attachments")
    
    with next(get_db()) as db:
        try:
            # Verify job exists and user has access
            job = db.query(ExtractionJob).filter(
                ExtractionJob.id == job_id,
                ExtractionJob.user_id == user_id
            ).first()
            
            if not job:
                raise ValueError(f"Job {job_id} not found or access denied")
            
            # Use service layer for actual import logic
            from services.google_service import google_service
            logger.info(f"Gmail import worker: automation_run_id parameter = {automation_run_id}")
            import_result = await google_service.import_gmail_attachments(db, job_id, attachment_data, automation_run_id)
            
            # Convert service result to worker result format
            results = {
                'job_id': job_id,
                'total_files': import_result.get('total', len(attachment_data)),
                'successful': import_result.get('successful', 0),
                'failed': import_result.get('failed', 0),
                'errors': import_result.get('errors', [])
            }
            
            # Update job status
            if results['failed'] == 0:
                logger.info(f"Gmail import completed successfully for job {job_id}")
            else:
                logger.warning(f"Gmail import completed with {results['failed']} failures for job {job_id}")
            
            return results
            
        except Exception as e:
            logger.error(f"Gmail import failed for job {job_id}: {e}")
            raise



async def _handle_zip_file(
    db: Session,
    source_file: SourceFile,
    zip_content: bytes
) -> None:
    """Handle ZIP file extraction and create individual source file records"""
    logger.info(f"Processing ZIP file: {source_file.original_filename}")
    
    try:
        with zipfile.ZipFile(io.BytesIO(zip_content), 'r') as zip_ref:
            for file_info in zip_ref.infolist():
                # Skip directories and hidden files
                if file_info.is_dir() or file_info.filename.startswith('.'):
                    continue
                
                # Extract individual file
                extracted_content = zip_ref.read(file_info.filename)
                extracted_filename = os.path.basename(file_info.filename)
                
                # Determine MIME type
                if extracted_filename.lower().endswith('.pdf'):
                    mime_type = 'application/pdf'
                else:
                    mime_type = 'application/octet-stream'
                
                # Create source file record for extracted file
                extracted_source_file = SourceFile(
                    job_id=source_file.job_id,
                    original_filename=extracted_filename,
                    original_path=file_info.filename,  # Full path within ZIP
                    gcs_object_name=f"imports/{source_file.job_id}/extracted_{extracted_filename}",
                    file_type=mime_type,
                    file_size_bytes=len(extracted_content),
                    status='importing',
                    source_type=source_file.source_type,
                    external_id=f"{source_file.external_id}:{file_info.filename}"
                )
                db.add(extracted_source_file)
                db.flush()  # Get ID without committing
                
                try:
                    # Upload extracted file to GCS
                    storage_service = get_storage_service()
                    await storage_service.upload_file_content(
                        file_content=extracted_content,
                        gcs_object_name=extracted_source_file.gcs_object_name
                    )
                    
                    extracted_source_file.status = 'unpacked'
                    logger.info(f"Extracted and uploaded: {extracted_filename}")
                    
                except Exception as e:
                    extracted_source_file.status = 'failed'
                    logger.error(f"Failed to upload extracted file {extracted_filename}: {e}")
            
            db.commit()
            logger.info(f"ZIP processing completed for {source_file.original_filename}")
            
    except Exception as e:
        logger.error(f"Failed to process ZIP file {source_file.original_filename}: {e}")
        raise


class ImportWorkerSettings:
    """ARQ worker configuration for import tasks (imports queue)"""
    
    redis_settings = RedisSettings.from_dsn(REDIS_URL)
    queue_name = "imports"  # Dedicated queue for import tasks
    
    # Task functions for import processing
    functions = [
        import_drive_files,
        import_gmail_attachments
    ]
    
    # Worker configuration for import tasks (moderate concurrency)
    max_jobs = 10  # Higher concurrency for I/O bound tasks
    job_timeout = 300  # 5 minutes timeout for import operations
    keep_result = 3600  # Keep results for 1 hour
    
    # Logging
    log_results = True

# ===================================================================
# Export Worker Functions (Google Drive, etc.)
# ===================================================================

async def export_startup(ctx: Dict[str, Any]) -> None:
    """Initialize export worker resources"""
    logger.info("Export worker starting up...")

async def export_shutdown(ctx: Dict[str, Any]) -> None:
    """Cleanup export worker resources"""
    logger.info("Export worker shutting down...")

async def export_job_to_google_drive(
    ctx: Dict[str, Any],
    job_id: str,
    user_id: str,
    file_type: str,  # 'csv' or 'xlsx'
    folder_id: str = None,
    automation_run_id: str = None
) -> Dict[str, Any]:
    """
    Export job results to Google Drive as CSV or Excel
    
    Args:
        job_id: Extraction job ID
        user_id: User ID for OAuth credentials
        file_type: Export format ('csv' or 'xlsx')
        folder_id: Optional Google Drive folder ID
        
    Returns:
        Dict with export results and Google Drive file info
    """
    logger.info(f"Starting Google Drive export for job {job_id}, format: {file_type}")
    
    with next(get_db()) as db:
        try:
            # Import here to avoid circular imports
            from models.db_models import JobExport
            from services.job_service import JobService
            from services.google_service import google_service
            
            # Verify job exists and user has access
            job = db.query(ExtractionJob).filter(
                ExtractionJob.id == job_id,
                ExtractionJob.user_id == user_id
            ).first()
            
            if not job:
                raise ValueError(f"Job {job_id} not found or access denied")
            
            # Create JobExport record to track the export
            job_export = JobExport(
                job_id=job_id,
                dest_type="gdrive",
                file_type=file_type,
                status="processing"
            )
            db.add(job_export)
            db.commit()
            db.refresh(job_export)
            
            # Send export started event
            from services.sse_service import sse_manager
            await sse_manager.send_export_started(job_id, "Google Drive", file_type)
            
            # Get job results
            job_service = JobService()
            results_response = await job_service.get_job_results(user_id, job_id)
            
            if not results_response.results:
                raise ValueError("No results found for this job")
            
            # Generate export content based on file type
            if file_type == "csv":
                content = _generate_csv_content(results_response)
                filename = f"job_{job_id}_results.csv"
                mime_type = "text/csv"
                content_bytes = content.encode('utf-8')
            elif file_type == "xlsx":
                content_bytes = _generate_excel_content(results_response)
                filename = f"job_{job_id}_results.xlsx"
                mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            else:
                raise ValueError(f"Unsupported file type: {file_type}")
            
            # Upload to Google Drive
            drive_file = google_service.upload_to_drive(
                db=db,
                user_id=user_id,
                file_content=content_bytes,
                filename=filename,
                mime_type=mime_type,
                folder_id=folder_id
            )
            
            if not drive_file:
                raise ValueError("Failed to upload to Google Drive")
            
            # Update JobExport record with success
            job_export.status = "completed"
            job_export.external_id = drive_file.get('id')
            db.commit()
            
            # Send export completed event
            await sse_manager.send_export_completed(
                job_id, 
                "Google Drive", 
                file_type,
                drive_file.get('webViewLink')
            )
            
            # Check if this export was triggered by an automation
            if automation_run_id:
                # This export was triggered by a specific automation run
                from services.automation_service import automation_service
                await automation_service.update_automation_run_status(
                    db, automation_run_id, 'completed'
                )
                logger.info(f"Marked automation run {automation_run_id} as completed after export")
            
            logger.info(f"Successfully exported job {job_id} to Google Drive: {drive_file.get('id')}")
            
            return {
                "success": True,
                "job_id": job_id,
                "file_type": file_type,
                "drive_file_id": drive_file.get('id'),
                "drive_file_name": drive_file.get('name'),
                "web_view_link": drive_file.get('webViewLink'),
                "web_content_link": drive_file.get('webContentLink')
            }
            
        except Exception as e:
            logger.error(f"Google Drive export failed for job {job_id}: {e}")
            
            # Update JobExport record with failure
            if 'job_export' in locals():
                job_export.status = "failed"
                job_export.error_message = str(e)
                db.commit()
            
            # Send export failed event
            try:
                from services.sse_service import sse_manager
                await sse_manager.send_export_failed(job_id, "Google Drive", file_type, str(e))
            except Exception as sse_error:
                logger.warning(f"Failed to send export_failed SSE event: {sse_error}")
            
            raise

def _generate_csv_content(results_response) -> str:
    """Generate CSV content from job results (sync version for worker)"""
    import csv
    from io import StringIO
    
    if not results_response.results:
        raise ValueError("No results found for this job")
    
    # Create CSV content
    output = StringIO()
    
    # Determine field names from the first result
    first_result = results_response.results[0]
    if not first_result.extracted_data:
        raise ValueError("No extracted data found")
    
    # Get field names from the columns snapshot in extracted_data
    if "columns" not in first_result.extracted_data:
        raise ValueError("Invalid extracted data format - missing columns")
    
    field_names = first_result.extracted_data["columns"]
    
    # Process array-based results
    writer = csv.DictWriter(output, fieldnames=field_names)
    writer.writeheader()
    
    for result in results_response.results:
        if result.extracted_data and "results" in result.extracted_data:
            for result_array in result.extracted_data["results"]:
                row = {}
                for i, field_name in enumerate(field_names):
                    if i < len(result_array):
                        value = result_array[i]
                        row[field_name] = str(value) if value is not None else ""
                    else:
                        row[field_name] = ""
                writer.writerow(row)
    
    # Get CSV content
    csv_content = output.getvalue()
    output.close()
    return csv_content

def _generate_excel_content(results_response) -> bytes:
    """Generate Excel content from job results (sync version for worker)"""
    import openpyxl
    from io import BytesIO
    
    if not results_response.results:
        raise ValueError("No results found for this job")
    
    # Create Excel workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Extraction Results"
    
    # Determine field names from the first result
    first_result = results_response.results[0]
    if not first_result.extracted_data:
        raise ValueError("No extracted data found")
    
    # Get field names from the columns snapshot in extracted_data
    if "columns" not in first_result.extracted_data:
        raise ValueError("Invalid extracted data format - missing columns")
    
    field_names = first_result.extracted_data["columns"]
    
    # Write headers
    for col_idx, field_name in enumerate(field_names, 1):
        worksheet.cell(row=1, column=col_idx, value=field_name)
    
    # Process array-based results
    row_idx = 2
    for result in results_response.results:
        if result.extracted_data and "results" in result.extracted_data:
            for result_array in result.extracted_data["results"]:
                for col_idx, field_name in enumerate(field_names, 1):
                    if col_idx - 1 < len(result_array):
                        value = result_array[col_idx - 1]
                        cell_value = str(value) if value is not None else ""
                    else:
                        cell_value = ""
                    worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                row_idx += 1
    
    # Save to BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()

class ExportWorkerSettings:
    """ARQ worker configuration for export tasks (exports queue)"""
    
    redis_settings = RedisSettings.from_dsn(REDIS_URL)
    queue_name = "exports"  # Dedicated queue for export tasks
    
    # Task functions for export processing
    functions = [
        export_job_to_google_drive
    ]
    
    # Worker configuration for export tasks (moderate concurrency)
    max_jobs = 5  # Moderate concurrency for export operations
    job_timeout = 600  # 10 minutes timeout for export operations (longer for large datasets)
    keep_result = 3600  # Keep results for 1 hour
    
    # Logging
    log_results = True

# ===================================================================
# Automation Worker Functions
# ===================================================================

async def automation_startup(ctx: Dict[str, Any]) -> None:
    """Initialize automation worker resources"""
    logger.info("Automation worker starting up...")

async def automation_shutdown(ctx: Dict[str, Any]) -> None:
    """Cleanup automation worker resources"""
    logger.info("Automation worker shutting down...")

async def automation_trigger_worker(
    ctx: Dict[str, Any],
    user_id: str,
    message_data: Dict[str, Any]
) -> Dict[str, Any]:
    """
    Process Gmail push notifications and trigger automations
    
    Args:
        user_id: User ID for the Gmail account
        message_data: Gmail Pub/Sub message payload
        
    Returns:
        Dict with processing results
    """
    logger.info(f"Processing Gmail trigger for user {user_id}")
    
    with next(get_db()) as db:
        try:
            from services.automation_service import automation_service
            from services.google_service import google_service
            
            # Get enabled automations for this user
            automations = await automation_service.get_enabled_automations(db, user_id)
            gmail_automations = [auto for auto in automations if auto.trigger_type == 'gmail_attachment']
            
            if not gmail_automations:
                logger.info(f"No Gmail automations found for user {user_id}")
                return {"processed": 0, "message": "No Gmail automations configured"}
            
            # Double-check automation limits before processing (in case plan was downgraded)
            try:
                from services.billing_service import get_billing_service
                billing_service = get_billing_service(db)
                can_run_automations = billing_service.check_automation_limit(user_id)
                
                if not can_run_automations:
                    billing_info = billing_service.get_billing_info(user_id)
                    logger.warning(f"User {user_id} has exceeded automation limits ({billing_info['automations_count']}/{billing_info['automations_limit']}), skipping automation processing")
                    return {"processed": 0, "message": "Automation limit exceeded"}
            except Exception as e:
                logger.warning(f"Could not check automation limits for user {user_id}: {e}")
                # Continue processing if we can't check limits
            
            processed_count = 0
            
            # Use Gmail History API to get new messages since last processed history ID
            # Extract history ID from push notification (prefer snake_case, fallback to camelCase)
            current_history_id = message_data.get('history_id') or message_data.get('raw_data', {}).get('historyId')
            
            if not current_history_id:
                logger.warning("No history_id in push notification")
                return {"processed": 0, "message": "No history_id in notification"}
            
            logger.info(f"Processing Gmail push notification with history ID: {current_history_id}")
            
            # Get the last processed history ID for this user (stored in database)
            last_processed_history_id = google_service.get_last_processed_history_id(db, user_id)
            
            if last_processed_history_id and int(current_history_id) <= int(last_processed_history_id):
                logger.info(f"History ID {current_history_id} already processed (last: {last_processed_history_id}), skipping")
                return {"processed": 0, "message": "History already processed"}
            
            # Get new messages since the last processed history ID
            start_history_id = last_processed_history_id or str(int(current_history_id) - 1)
            message_ids_to_check = google_service.get_messages_since_history(db, user_id, start_history_id)
            
            if not message_ids_to_check:
                logger.info(f"No messages to check for user {user_id}")
                return {"processed": 0, "message": "No messages to check"}
            
            for automation in gmail_automations:
                try:
                    # Extract Gmail query from trigger config
                    gmail_query = automation.trigger_config.get('query', '')
                    if not gmail_query:
                        logger.warning(f"Automation {automation.id} has no Gmail query configured")
                        continue
                    
                    # Check each message against the automation query
                    for message_id in message_ids_to_check:
                        logger.info(f"Checking message {message_id} against automation {automation.id} query: '{gmail_query}'")
                        
                        # Check if this specific message matches the automation query
                        if google_service.message_matches_query(db, user_id, message_id, gmail_query):
                            logger.info(f"Message {message_id} matches automation {automation.id} query")
                            
                            # Check if we've already processed this message for this automation
                            from models.db_models import AutomationProcessedMessage
                            existing_processed = db.query(AutomationProcessedMessage).filter(
                                AutomationProcessedMessage.automation_id == automation.id,
                                AutomationProcessedMessage.message_id == message_id
                            ).first()
                            
                            if existing_processed:
                                logger.info(f"Message {message_id} already processed by automation {automation.id} at {existing_processed.processed_at}, skipping")
                                continue
                            
                            # Get message attachments
                            attachments = google_service.get_gmail_message_attachments(db, user_id, message_id)
                            
                            if not attachments:
                                logger.info(f"Message {message_id} matches query but has no attachments")
                                continue
                            
                            logger.info(f"Message {message_id} has {len(attachments)} attachments")
                            
                            # Mark message as processed BEFORE creating automation run
                            from models.db_models import AutomationProcessedMessage
                            processed_message = AutomationProcessedMessage(
                                automation_id=automation.id,
                                message_id=message_id
                            )
                            db.add(processed_message)
                            db.commit()
                            
                            logger.info(f"Marked message {message_id} as processed for automation {automation.id}")
                            
                            # Check if job is currently running before proceeding
                            job = db.query(ExtractionJob).filter(ExtractionJob.id == automation.job_id).first()
                            if job and job.status == 'in_progress':
                                logger.warning(f"Job {automation.job_id} is currently running, skipping automation trigger")
                                continue
                            
                            # Clear existing data for automation runs to support multiple runs
                            logger.info(f"Clearing existing data for automation job {automation.job_id}")
                            await _clear_job_data_for_automation(db, automation.job_id)
                            
                            # Create automation run with import tracking
                            automation_run = await automation_service.create_automation_run(
                                db, automation.id, automation.job_id
                            )
                            
                            # Initialize import tracking
                            automation_run.imports_total = len(attachments)
                            automation_run.imports_successful = 0
                            automation_run.imports_failed = 0
                            automation_run.imports_processed = 0
                            automation_run.imports_processing_failed = 0
                            db.commit()
                            
                            # Enqueue Gmail import worker (now using consolidated function)
                            from arq import create_pool
                            redis = await create_pool(ImportWorkerSettings.redis_settings)
                            
                            await redis.enqueue_job(
                                'import_gmail_attachments',
                                job_id=automation.job_id,
                                user_id=user_id,
                                attachment_data=attachments,
                                automation_run_id=str(automation_run.id),
                                _queue_name='imports'
                            )
                            
                            await redis.close()
                            processed_count += 1
                            
                            logger.info(f"Triggered automation {automation.id} for new message {message_id}")
                            processed_count += 1
                
                except Exception as e:
                    logger.error(f"Failed to process automation {automation.id}: {e}")
                    continue
            
            # Update the last processed history ID after processing
            google_service.update_last_processed_history_id(db, user_id, str(current_history_id))
            logger.info(f"Updated last processed history ID to {current_history_id}")
            
            return {
                "processed": processed_count,
                "message": f"Processed {processed_count} automation triggers"
            }
            
        except Exception as e:
            logger.error(f"Automation trigger worker failed for user {user_id}: {e}")
            raise

async def run_initializer_worker(
    ctx: Dict[str, Any],
    job_id: str,
    automation_run_id: str = None
) -> Dict[str, Any]:
    """
    Initialize job execution (now just delegates to consolidated job service)
    
    Args:
        job_id: Extraction job ID to initialize
        automation_run_id: Optional automation run ID if triggered by automation
        
    Returns:
        Dict with initialization results
    """
    logger.info(f"Initializing job execution for job {job_id}, automation_run_id={automation_run_id}")
    
    with next(get_db()) as db:
        try:
            from services.job_service import JobService
            
            # Get user_id from job
            job = db.query(ExtractionJob).filter(ExtractionJob.id == job_id).first()
            if not job:
                raise ValueError(f"Job {job_id} not found")
            
            # Use automation-specific job service method
            job_service = JobService()
            await job_service.submit_automation_job(job_id, job.user_id, automation_run_id)
            
            # Count tasks for response
            tasks_count = db.query(ExtractionTask).filter(
                ExtractionTask.job_id == job_id
            ).count()
            
            logger.info(f"Initialized job {job_id} with {tasks_count} tasks")
            
            return {
                "message": "Job initialization successful",
                "tasks_queued": tasks_count,
                "automation_run_id": automation_run_id
            }
            
        except Exception as e:
            logger.error(f"Job initialization failed for job {job_id}: {e}")
            
            # Update automation run status if this is from automation
            if automation_run_id:
                try:
                    from services.automation_service import automation_service
                    await automation_service.update_automation_run_status(
                        db, automation_run_id, 'failed', str(e)
                    )
                except Exception as sse_error:
                    logger.warning(f"Failed to update automation run status: {sse_error}")
            
            raise


class AutomationWorkerSettings:
    """ARQ worker configuration for automation tasks (automation queue)"""
    
    redis_settings = RedisSettings.from_dsn(REDIS_URL)
    queue_name = "automation"  # Dedicated queue for automation tasks
    
    # Task functions for automation processing
    functions = [
        automation_trigger_worker,
        run_initializer_worker
    ]
    
    # Worker configuration for automation tasks (lower concurrency)
    max_jobs = 3  # Lower concurrency for automation triggers
    job_timeout = 600  # 10 minutes timeout for automation operations
    keep_result = 3600  # Keep results for 1 hour
    
    # Logging
    log_results = True

async def main():
    """For testing the worker locally"""
    redis = await create_pool(WorkerSettings.redis_settings)
    
    # Example: enqueue a test task
    job = await redis.enqueue_job('process_extraction_task', 'test-task-id')
    print(f"Enqueued job: {job}")
    
    await redis.close()

# These functions are now moved to job_service.py for better organization


async def _record_usage_for_task(db: Session, task: ExtractionTask, source_files: List[SourceFile]):
    """
    Record usage for billing when an extraction task completes successfully
    """
    try:
        from services.billing_service import get_billing_service, PlanLimitExceeded
        
        # Calculate total pages processed
        total_pages = 0
        for source_file in source_files:
            if source_file.page_count:
                total_pages += source_file.page_count
            else:
                # Fallback: estimate 1 page per file if page_count not available
                total_pages += 1
                logger.warning(f"No page_count for file {source_file.id}, using 1 page estimate")
        
        if total_pages <= 0:
            logger.info("No pages to record for billing")
            return
        
        # Get user ID from the job
        job = db.query(ExtractionJob).filter(ExtractionJob.id == task.job_id).first()
        if not job:
            logger.error(f"Job {task.job_id} not found for usage recording")
            return
        
        # Record usage through billing service
        billing_service = get_billing_service(db)
        event_id = billing_service.record_usage(
            user_id=job.user_id,
            pages=total_pages,
            source="extraction_task",
            task_id=str(task.id),
            notes=f"Processed {len(source_files)} files"
        )
        
        logger.info(f"Recorded {total_pages} pages usage for user {job.user_id}, task {task.id}, event {event_id}")
        
    except PlanLimitExceeded as e:
        # This shouldn't happen since we check limits before starting tasks
        logger.error(f"Plan limit exceeded during task completion: {e}")
        raise  # Re-raise to fail the task
    except Exception as e:
        logger.error(f"Unexpected error recording usage: {e}")
        # Don't re-raise for other errors to avoid failing the extraction task

# Removed per-task limit checking - now handled at job start

class CronWorkerSettings:
    """Settings for cron/scheduled tasks worker"""
    redis_settings = RedisSettings.from_dsn(REDIS_URL)
    queue_name = "cron_queue"
    
    # Worker functions (both manual and scheduled versions)
    functions = [
        # Manual cleanup functions (can be called directly)
        run_abandoned_cleanup,
        run_opt_out_cleanup,
        run_artifact_cleanup,
        run_free_user_period_reset,
        run_stripe_usage_reconciliation,
        run_usage_counter_cleanup,
        # Scheduled wrapper functions
        schedule_free_user_period_reset,
        schedule_stripe_usage_reconciliation,
        schedule_usage_counter_cleanup,
        schedule_abandoned_cleanup,
        schedule_artifact_cleanup,
        schedule_opt_out_cleanup,
    ]
    
    # Cron jobs schedule
    cron_jobs = [
        # Free user period reset - Daily at 00:30 UTC (after month rollover)
        cron(schedule_free_user_period_reset, hour=0, minute=30, run_at_startup=False),
        
        # Stripe usage reconciliation - Every 2 hours
        cron(schedule_stripe_usage_reconciliation, hour={0, 2, 4, 6, 8, 10, 12, 14, 16, 18, 20, 22}, minute=15, run_at_startup=False),
        
        # Usage counter cleanup - Weekly on Sundays at 02:00 UTC
        cron(schedule_usage_counter_cleanup, weekday=6, hour=2, minute=0, run_at_startup=False),
        
        # Abandoned job cleanup - Daily at 01:00 UTC
        cron(schedule_abandoned_cleanup, hour=1, minute=0, run_at_startup=False),
        
        # Artifact cleanup - Daily at 03:00 UTC
        cron(schedule_artifact_cleanup, hour=3, minute=0, run_at_startup=False),
        
        # Opt-out data cleanup - Weekly on Saturdays at 04:00 UTC
        cron(schedule_opt_out_cleanup, weekday=5, hour=4, minute=0, run_at_startup=False),
    ]
    
    # Worker configuration
    max_jobs = 5  # Lower concurrency for maintenance tasks
    job_timeout = 1800  # 30 minutes timeout for cleanup tasks
    keep_result = 86400  # Keep results for 24 hours
    
    # Health check
    health_check_interval = 300  # 5 minutes
    
    # Logging
    log_results = True

if __name__ == "__main__":
    asyncio.run(main())