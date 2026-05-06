"""
Service for counting pages in uploaded files
"""
import os
import tempfile
import logging
import io
import zipfile
from typing import Optional
import PyPDF2
import fitz  # PyMuPDF - fallback for complex PDFs
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

class PageCountingService:
    """Service for counting pages in various file types"""
    
    @staticmethod
    def count_pages_from_content(file_content: bytes, filename: str) -> Optional[int]:
        """
        Count pages in a file from its content
        
        Args:
            file_content: Raw file content as bytes
            filename: Original filename for type detection
            
        Returns:
            Number of pages, or None if unable to count
        """
        try:
            # Determine file type from extension
            file_ext = os.path.splitext(filename)[1].lower()
            
            if file_ext == '.pdf':
                return PageCountingService._count_pdf_pages(file_content)
            elif file_ext == '.pptx':
                return PageCountingService._count_pptx_slides(file_content)
            elif file_ext == '.xlsx':
                return PageCountingService._count_xlsx_sheets_from_content(file_content)
            else:
                # For non-PDF files, assume 1 page
                logger.info(f"Non-PDF file {filename}, assuming 1 page")
                return 1
                
        except Exception as e:
            logger.error(f"Error counting pages for {filename}: {e}")
            return None
    
    @staticmethod
    def _count_pdf_pages(pdf_content: bytes) -> Optional[int]:
        """
        Count pages in a PDF file using multiple methods
        
        Args:
            pdf_content: PDF file content as bytes
            
        Returns:
            Number of pages, or None if unable to count
        """
        # Method 1: Try PyPDF2 first (faster)
        try:
            import io
            pdf_reader = PyPDF2.PdfReader(io.BytesIO(pdf_content))
            page_count = len(pdf_reader.pages)
            logger.debug(f"PyPDF2 counted {page_count} pages")
            return page_count
        except Exception as e:
            logger.warning(f"PyPDF2 failed to count pages: {e}")
        
        # Method 2: Try PyMuPDF as fallback (more robust)
        try:
            import io
            pdf_doc = fitz.open(stream=pdf_content, filetype="pdf")
            page_count = pdf_doc.page_count
            pdf_doc.close()
            logger.debug(f"PyMuPDF counted {page_count} pages")
            return page_count
        except Exception as e:
            logger.warning(f"PyMuPDF failed to count pages: {e}")
        
        # Method 3: Try with temporary file (last resort)
        try:
            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_file:
                temp_file.write(pdf_content)
                temp_file.flush()
                
                try:
                    pdf_doc = fitz.open(temp_file.name)
                    page_count = pdf_doc.page_count
                    pdf_doc.close()
                    logger.debug(f"PyMuPDF (temp file) counted {page_count} pages")
                    return page_count
                finally:
                    os.unlink(temp_file.name)
                    
        except Exception as e:
            logger.error(f"All PDF page counting methods failed: {e}")
        
        return None

    @staticmethod
    def count_pages_from_file_path(file_path: str, filename: str) -> Optional[int]:
        """Count pages from a local file path.

        This avoids loading large PDFs into memory when possible.
        """
        try:
            file_ext = os.path.splitext(filename)[1].lower()
            if file_ext == '.pptx':
                return PageCountingService._count_pptx_slides_from_path(file_path)
            if file_ext == '.xlsx':
                return PageCountingService._count_xlsx_sheets_from_path(file_path)
            if file_ext != '.pdf':
                logger.info(f"Non-PDF file {filename}, assuming 1 page")
                return 1

            # Method 1: Try PyPDF2 from file (fast)
            try:
                with open(file_path, 'rb') as f:
                    pdf_reader = PyPDF2.PdfReader(f)
                    page_count = len(pdf_reader.pages)
                    logger.debug(f"PyPDF2 counted {page_count} pages (file path)")
                    return page_count
            except Exception as e:
                logger.warning(f"PyPDF2 failed to count pages from file path: {e}")

            # Method 2: Try PyMuPDF from file path (more robust)
            try:
                pdf_doc = fitz.open(file_path)
                page_count = pdf_doc.page_count
                pdf_doc.close()
                logger.debug(f"PyMuPDF counted {page_count} pages (file path)")
                return page_count
            except Exception as e:
                logger.warning(f"PyMuPDF failed to count pages from file path: {e}")

            return None
        except Exception as e:
            logger.error(f"Error counting pages from file path for {filename}: {e}")
            return None

    @staticmethod
    def _count_pptx_slides(file_content: bytes) -> Optional[int]:
        try:
            with zipfile.ZipFile(io.BytesIO(file_content), 'r') as archive:
                slide_names = [
                    name for name in archive.namelist()
                    if name.startswith('ppt/slides/slide') and name.endswith('.xml')
                ]
            return max(1, len(slide_names))
        except Exception as e:
            logger.warning(f"Failed to count PPTX slides: {e}")
            return None

    @staticmethod
    def _count_pptx_slides_from_path(file_path: str) -> Optional[int]:
        try:
            with zipfile.ZipFile(file_path, 'r') as archive:
                slide_names = [
                    name for name in archive.namelist()
                    if name.startswith('ppt/slides/slide') and name.endswith('.xml')
                ]
            return max(1, len(slide_names))
        except Exception as e:
            logger.warning(f"Failed to count PPTX slides from file path: {e}")
            return None

    @staticmethod
    def _count_xlsx_sheets_from_content(file_content: bytes) -> Optional[int]:
        try:
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                temp_file.write(file_content)
                temp_file.flush()
                temp_path = temp_file.name
            try:
                return PageCountingService._count_xlsx_sheets_from_path(temp_path)
            finally:
                os.unlink(temp_path)
        except Exception as e:
            logger.warning(f"Failed to count XLSX sheets from content: {e}")
            return None

    @staticmethod
    def _count_xlsx_sheets_from_path(file_path: str) -> Optional[int]:
        workbook = None
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            non_empty = 0
            for worksheet in workbook.worksheets:
                has_value = False
                for row in worksheet.iter_rows(values_only=True):
                    if any(value is not None and str(value).strip() for value in row):
                        has_value = True
                        break
                if has_value:
                    non_empty += 1
            return max(1, non_empty)
        except Exception as e:
            logger.warning(f"Failed to count XLSX sheets from file path: {e}")
            return None
        finally:
            if workbook is not None:
                workbook.close()

# Global instance
page_counting_service = PageCountingService()
