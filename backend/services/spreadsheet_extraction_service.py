from __future__ import annotations

import os
from itertools import zip_longest
from typing import Any

from openpyxl import load_workbook


class SpreadsheetExtractionService:
    """Render XLSX workbooks into deterministic text for Gemini extraction."""

    def __init__(self) -> None:
        self.max_rows_per_sheet = int(os.getenv("EXTRACTION_XLSX_MAX_ROWS_PER_SHEET", "1000"))
        self.max_cols_per_sheet = int(os.getenv("EXTRACTION_XLSX_MAX_COLS_PER_SHEET", "100"))
        self.max_chars = int(os.getenv("EXTRACTION_XLSX_MAX_CHARS", "200000"))

    def _format_value(self, value: Any) -> str:
        if value is None:
            return ""
        return str(value).replace("\n", " ").replace("\r", " ").strip()

    def _worksheet_dimension(self, worksheet: Any) -> str:
        try:
            return str(worksheet.calculate_dimension(force=True))
        except TypeError:
            try:
                return str(worksheet.calculate_dimension())
            except Exception:
                return "unknown"
        except Exception:
            return "unknown"

    def render_xlsx_local_to_text(self, local_path: str, filename: str | None = None) -> str:
        values_wb = load_workbook(local_path, read_only=True, data_only=True)
        formulas_wb = load_workbook(local_path, read_only=True, data_only=False)
        parts: list[str] = []
        workbook_non_empty_rows = 0
        try:
            title = filename or os.path.basename(local_path)
            parts.append(f"Workbook: {title}")

            for sheet_index, values_ws in enumerate(values_wb.worksheets, start=1):
                formulas_ws = formulas_wb[values_ws.title]
                dimension = self._worksheet_dimension(values_ws)
                sheet_header_index = len(parts)
                parts.append(f"\nSheet {sheet_index}: {values_ws.title} (dimension={dimension})")

                value_rows = values_ws.iter_rows(
                    max_row=self.max_rows_per_sheet + 1,
                    max_col=self.max_cols_per_sheet + 1,
                )
                formula_rows = formulas_ws.iter_rows(
                    max_row=self.max_rows_per_sheet + 1,
                    max_col=self.max_cols_per_sheet + 1,
                )
                rows_added = 0
                max_rendered_cols = 0
                row_limit_reached = False
                column_limit_reached = False

                for row_number, (value_row, formula_row) in enumerate(
                    zip_longest(value_rows, formula_rows, fillvalue=()),
                    start=1,
                ):
                    if row_number > self.max_rows_per_sheet:
                        row_limit_reached = True
                        break

                    cells: list[str] = []
                    row_has_value = False

                    for col_number, (value_cell, formula_cell) in enumerate(
                        zip_longest(value_row or (), formula_row or (), fillvalue=None),
                        start=1,
                    ):
                        display = self._format_value(getattr(value_cell, "value", None))
                        formula = self._format_value(getattr(formula_cell, "value", None))
                        if col_number > self.max_cols_per_sheet:
                            if display or formula:
                                column_limit_reached = True
                            break

                        if formula.startswith("=") and formula != display:
                            display = f"{display} [formula: {formula}]" if display else f"[formula: {formula}]"
                        if display:
                            row_has_value = True
                        cells.append(display)

                    if row_has_value:
                        while cells and not cells[-1]:
                            cells.pop()
                        rows_added += 1
                        workbook_non_empty_rows += 1
                        max_rendered_cols = max(max_rendered_cols, len(cells))
                        parts.append(f"R{row_number}: " + " | ".join(cells).rstrip(" |"))

                    current = "\n".join(parts)
                    if len(current) >= self.max_chars:
                        return current[: self.max_chars] + "\n[TRUNCATED]"

                parts[sheet_header_index] = (
                    f"\nSheet {sheet_index}: {values_ws.title} "
                    f"(dimension={dimension}, rendered_rows={rows_added}, rendered_columns={max_rendered_cols})"
                )

                if row_limit_reached:
                    parts.append(f"[TRUNCATED rows after {self.max_rows_per_sheet}]")
                if column_limit_reached:
                    parts.append(f"[TRUNCATED columns after {self.max_cols_per_sheet}]")
                if rows_added == 0:
                    parts.append("(no non-empty rows found in inspected range)")

            if workbook_non_empty_rows == 0:
                raise ValueError("XLSX workbook contains no readable non-empty rows")

            return "\n".join(parts)[: self.max_chars]
        finally:
            values_wb.close()
            formulas_wb.close()


spreadsheet_extraction_service = SpreadsheetExtractionService()
