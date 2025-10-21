"""
Export service for generating CSV and Excel files from job results
"""
import csv
import openpyxl
from io import StringIO, BytesIO
from typing import List
from datetime import datetime
import re


def generate_csv_content(results_response) -> str:
    """Generate CSV content from job results with source file paths"""
    if not results_response.results:
        raise ValueError("No results found for this job")
    
    # Create CSV content
    output = StringIO()
    
    # Determine field names from the first result
    first_result = results_response.results[0]
    if not first_result.extracted_data:
        raise ValueError("No extracted data found")
    
    # Get field names from the columns snapshot in extracted_data
    if "columns" not in first_result.extracted_data:
        raise ValueError("Invalid extracted data format - missing columns")
    
    field_names = first_result.extracted_data["columns"]
    
    # Add source file paths column as first column
    field_names_with_source = ["Source File Path(s)"] + field_names
    
    # Process array-based results
    writer = csv.DictWriter(output, fieldnames=field_names_with_source)
    writer.writeheader()
    
    for result in results_response.results:
        if result.extracted_data and "results" in result.extracted_data:
            # Get source file paths for this result
            source_paths = _get_source_file_paths(result)
            
            for result_array in result.extracted_data["results"]:
                row = {}
                # Add source file paths as first column
                row["Source File Path(s)"] = source_paths
                
                for i, field_name in enumerate(field_names):
                    if i < len(result_array):
                        value = result_array[i]
                        row[field_name] = str(value) if value is not None else ""
                    else:
                        row[field_name] = ""
                
                writer.writerow(row)
    
    # Get CSV content
    csv_content = output.getvalue()
    output.close()
    return csv_content


def generate_excel_content(results_response) -> bytes:
    """Generate Excel content from job results with source file paths"""
    if not results_response.results:
        raise ValueError("No results found for this job")
    
    # Create Excel workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Extraction Results"
    
    # Determine field names from the first result
    first_result = results_response.results[0]
    if not first_result.extracted_data:
        raise ValueError("No extracted data found")
    
    # Get field names from the columns snapshot in extracted_data
    if "columns" not in first_result.extracted_data:
        raise ValueError("Invalid extracted data format - missing columns")
    
    field_names = first_result.extracted_data["columns"]
    
    # Add source file paths column as first column
    field_names_with_source = ["Source File Path(s)"] + field_names
    
    # Write headers
    for col_num, field_name in enumerate(field_names_with_source, 1):
        worksheet.cell(row=1, column=col_num, value=field_name)
    
    # Write data
    row_num = 2
    for result in results_response.results:
        if result.extracted_data and "results" in result.extracted_data:
            # Get source file paths for this result
            source_paths = _get_source_file_paths(result)
            
            for result_array in result.extracted_data["results"]:
                # Add source file paths in the first column
                worksheet.cell(row=row_num, column=1, value=source_paths)
                
                # Add the rest of the data starting from column 2
                for col_num, field_name in enumerate(field_names, 2):
                    if col_num - 2 < len(result_array):
                        value = result_array[col_num - 2]
                        worksheet.cell(row=row_num, column=col_num, value=str(value) if value is not None else "")
                    else:
                        worksheet.cell(row=row_num, column=col_num, value="")
                
                row_num += 1
    
    # Save to bytes
    output = BytesIO()
    workbook.save(output)
    excel_content = output.getvalue()
    output.close()
    return excel_content


def _get_source_file_paths(result) -> str:
    """Extract source file paths from a result and return as comma-separated string"""
    source_paths = []
    
    # Check if result has source_files information
    if hasattr(result, 'source_files') and result.source_files:
        # If source_files is a list of file objects
        if isinstance(result.source_files, list):
            for source_file in result.source_files:
                if hasattr(source_file, 'original_filename'):
                    source_paths.append(source_file.original_filename)
                elif hasattr(source_file, 'filename'):
                    source_paths.append(source_file.filename)
                elif isinstance(source_file, str):
                    source_paths.append(source_file)
        else:
            # If source_files is a single file object
            if hasattr(result.source_files, 'original_filename'):
                source_paths.append(result.source_files.original_filename)
            elif hasattr(result.source_files, 'filename'):
                source_paths.append(result.source_files.filename)
    
    # Check if result has a single source_file
    elif hasattr(result, 'source_file') and result.source_file:
        if hasattr(result.source_file, 'original_filename'):
            source_paths.append(result.source_file.original_filename)
        elif hasattr(result.source_file, 'filename'):
            source_paths.append(result.source_file.filename)
    
    # Check if result has file_path or similar attribute
    elif hasattr(result, 'file_path') and result.file_path:
        source_paths.append(result.file_path)
    
    # Check if result has original_filename directly
    elif hasattr(result, 'original_filename') and result.original_filename:
        source_paths.append(result.original_filename)
    
    # If no source paths found, return empty string
    if not source_paths:
        return ""
    
    # Return comma-separated paths
    return ", ".join(source_paths)


def _slugify_filename_component(name: str) -> str:
    """Create a filesystem-safe component for filenames from a job name."""
    if not name:
        return "job"
    # Replace whitespace with underscore
    name = re.sub(r"\s+", "_", name.strip())
    # Remove characters not allowed in filenames on common OS
    name = re.sub(r"[^A-Za-z0-9._-]", "", name)
    # Collapse multiple underscores
    name = re.sub(r"_+", "_", name)
    # Trim to reasonable length
    return name[:80] if name else "job"


def generate_export_filename(job_name: str, export_time: datetime, ext: str) -> str:
    """
    Generate export filename using job name and export timestamp.
    Example: my_job_20250130_142355Z.csv
    """
    safe_job = _slugify_filename_component(job_name or "job")
    ts = export_time.strftime("%Y%m%d_%H%M%SZ")
    return f"{safe_job}_{ts}.{ext}"
