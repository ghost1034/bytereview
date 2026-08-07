"""Core business rules, serialization, reporting, and scheduled work for PBC."""

from __future__ import annotations

import csv
import hashlib
import io
import secrets
import uuid
import zipfile
from collections import Counter
from datetime import date, datetime, time, timedelta, timezone
from typing import Any
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from fastapi import HTTPException
from openpyxl import Workbook
from sqlalchemy import func
from sqlalchemy.orm import Session

from core.runtime import frontend_base_url
from models.db_models import AnalyticsUserRole, Client, Firm, User
from models.pbc import (
    PbcAccessToken,
    PbcAuditEvent,
    PbcComment,
    PbcContact,
    PbcDocument,
    PbcEngagement,
    PbcEngagementContact,
    PbcFirmSettings,
    PbcNotificationOutbox,
    PbcPortalSession,
    PbcRequest,
    PbcRequestAssignment,
    PbcTemplate,
)
from models.tasklytic import TasklyticEntityRecord, TasklyticWorkspaceMember
from services.email_service import email_service
from services.tasklytic_service import authorize_record


FIRM_WRITE_ROLES = {"admin", "manager", "analyst"}
FIRM_REVIEW_ROLES = {"admin", "manager", "reviewer"}
FIRM_MANAGE_ROLES = {"admin", "manager"}
REQUEST_TRANSITIONS = {
    "submit": ({"open", "needs_changes"}, "submitted"),
    "accept": ({"submitted"}, "accepted"),
    "return": ({"submitted"}, "needs_changes"),
    "waive": ({"open", "needs_changes"}, "waived"),
    "reopen": ({"accepted", "waived"}, "open"),
}
DEFAULT_AUDIT_TEMPLATE_ITEMS = [
    {"category": "General", "title": "Year-end trial balance", "priority": "high", "expected_formats": ["xlsx", "csv"]},
    {"category": "Cash", "title": "Bank reconciliations and bank statements for all accounts", "priority": "high", "expected_formats": ["pdf", "xlsx"]},
    {"category": "Accounts Receivable", "title": "Accounts receivable aging as of period end", "priority": "normal", "expected_formats": ["xlsx"]},
    {"category": "Fixed Assets", "title": "Fixed asset rollforward and additions detail", "priority": "normal", "expected_formats": ["xlsx"]},
    {"category": "Debt", "title": "Debt agreements and year-end lender statements", "priority": "normal", "expected_formats": ["pdf"]},
    {"category": "Revenue", "title": "Listing of significant customer contracts", "priority": "normal", "expected_formats": ["xlsx", "pdf"]},
]


def utcnow() -> datetime:
    return datetime.now(timezone.utc)


def token_hash(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def _spreadsheet_safe(value: Any) -> Any:
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def actor_role(actor: User) -> str:
    return str(getattr(actor.role, "value", actor.role))


def require_actor_role(actor: User, allowed: set[str]) -> None:
    if actor_role(actor) not in allowed:
        raise HTTPException(status_code=403, detail="Insufficient PBC permission for this action")


def get_settings(db: Session, firm_id) -> PbcFirmSettings:
    row = db.get(PbcFirmSettings, firm_id)
    if row is None:
        firm = db.get(Firm, firm_id)
        row = PbcFirmSettings(firm_id=firm_id, portal_name=firm.name if firm else "Client Portal")
        db.add(row)
        db.flush()
    return row


def ensure_default_template(db: Session, firm_id, actor_user_id: str | None = None) -> PbcTemplate:
    row = db.query(PbcTemplate).filter(
        PbcTemplate.firm_id == firm_id, PbcTemplate.name == "Annual financial statement audit"
    ).first()
    if row is None:
        row = PbcTemplate(
            firm_id=firm_id,
            name="Annual financial statement audit",
            description="Audit-first starter request list. Review and tailor before publishing.",
            engagement_type="audit",
            items=DEFAULT_AUDIT_TEMPLATE_ITEMS,
            created_by_user_id=actor_user_id,
        )
        db.add(row)
        db.flush()
    return row


def get_engagement(db: Session, firm_id, engagement_id: str, *, lock: bool = False) -> PbcEngagement:
    query = db.query(PbcEngagement).filter(PbcEngagement.id == engagement_id, PbcEngagement.firm_id == firm_id)
    if lock:
        query = query.with_for_update()
    row = query.first()
    if row is None:
        raise HTTPException(status_code=404, detail="PBC engagement not found")
    return row


def get_pbc_request(db: Session, firm_id, request_id: str, *, lock: bool = False) -> PbcRequest:
    query = db.query(PbcRequest).filter(PbcRequest.id == request_id, PbcRequest.firm_id == firm_id)
    if lock:
        query = query.with_for_update()
    row = query.first()
    if row is None:
        raise HTTPException(status_code=404, detail="PBC request not found")
    return row


def audit(
    db: Session,
    engagement: PbcEngagement,
    event_type: str,
    *,
    actor_kind: str,
    actor_id: str | None,
    request_id=None,
    details: dict[str, Any] | None = None,
) -> None:
    db.add(PbcAuditEvent(
        firm_id=engagement.firm_id,
        engagement_id=engagement.id,
        request_id=request_id,
        actor_kind=actor_kind,
        actor_id=actor_id,
        event_type=event_type,
        details=details or {},
    ))


def _user_label(user: User | None) -> str | None:
    if user is None:
        return None
    return user.display_name or user.email


def serialize_contact(contact: PbcContact, role: str | None = None) -> dict[str, Any]:
    return {
        "id": str(contact.id), "client_id": str(contact.client_id) if contact.client_id else None,
        "name": contact.name, "email": contact.email, "active": contact.active, "role": role,
        "created_at": contact.created_at,
    }


def serialize_document(row: PbcDocument) -> dict[str, Any]:
    return {
        "id": str(row.id), "request_id": str(row.request_id), "filename": row.filename,
        "mime_type": row.mime_type, "size_bytes": row.size_bytes, "checksum_sha256": row.checksum_sha256,
        "version": row.version, "state": row.state, "scan_status": row.scan_status,
        "scan_detail": row.scan_detail, "uploaded_by_kind": row.uploaded_by_kind,
        "uploaded_by_id": row.uploaded_by_id, "ai_metadata": row.ai_metadata,
        "completed_at": row.completed_at, "created_at": row.created_at,
    }


def serialize_request(db: Session, row: PbcRequest, *, include_internal: bool = True) -> dict[str, Any]:
    owner = db.get(User, row.owner_user_id) if row.owner_user_id else None
    assignments = (
        db.query(PbcContact)
        .join(PbcRequestAssignment, PbcRequestAssignment.contact_id == PbcContact.id)
        .filter(PbcRequestAssignment.request_id == row.id)
        .order_by(PbcContact.name)
        .all()
    )
    docs = db.query(PbcDocument).filter(
        PbcDocument.request_id == row.id, PbcDocument.state != "deleted"
    ).order_by(PbcDocument.version.desc()).all()
    comments_query = db.query(PbcComment).filter(PbcComment.request_id == row.id)
    if not include_internal:
        comments_query = comments_query.filter(PbcComment.visibility == "client")
    comments = comments_query.order_by(PbcComment.created_at).all()
    return {
        "id": str(row.id), "engagement_id": str(row.engagement_id), "request_number": row.request_number,
        "sort_order": row.sort_order, "category": row.category, "title": row.title,
        "description": row.description, "period_end": row.period_end, "priority": row.priority,
        "due_date": row.due_date, "owner_user_id": row.owner_user_id, "owner_name": _user_label(owner),
        "expected_filename": row.expected_filename, "expected_formats": row.expected_formats or [],
        "gl_account": row.gl_account, "gl_balance": row.gl_balance, "sensitive": row.sensitive,
        "requires_redaction": row.requires_redaction, "dependency_ids": row.dependency_ids or [],
        "status": row.status, "status_reason": row.status_reason, "revision": row.revision,
        "submitted_at": row.submitted_at, "accepted_at": row.accepted_at,
        "assignments": [serialize_contact(item) for item in assignments],
        "documents": [serialize_document(item) for item in docs if include_internal or item.state == "available"],
        "comments": [{
            "id": str(item.id), "body": item.body, "visibility": item.visibility,
            "actor_kind": item.actor_kind, "actor_name": item.actor_name, "created_at": item.created_at,
        } for item in comments],
        "created_at": row.created_at, "updated_at": row.updated_at,
    }


def serialize_engagement(db: Session, row: PbcEngagement, *, detail: bool = False, include_internal: bool = True) -> dict[str, Any]:
    requests = db.query(PbcRequest).filter(PbcRequest.engagement_id == row.id).order_by(
        PbcRequest.sort_order, PbcRequest.request_number
    ).all()
    statuses = Counter(item.status for item in requests)
    fulfilled = statuses["accepted"] + statuses["waived"]
    progress = round((fulfilled / len(requests)) * 100) if requests else 0
    contacts = []
    if detail:
        contact_rows = (
            db.query(PbcContact, PbcEngagementContact.role)
            .join(PbcEngagementContact, PbcEngagementContact.contact_id == PbcContact.id)
            .filter(PbcEngagementContact.engagement_id == row.id)
            .order_by(PbcContact.name)
            .all()
        )
        contacts = [serialize_contact(contact, role) for contact, role in contact_rows]
    result = {
        "id": str(row.id), "firm_id": str(row.firm_id), "client_id": str(row.client_id) if row.client_id else None,
        "name": row.name, "client_name": row.client_name_snapshot, "engagement_type": row.engagement_type,
        "period_start": row.period_start, "period_end": row.period_end, "due_date": row.due_date,
        "owner_user_id": row.owner_user_id, "status": row.status, "reminders_paused": row.reminders_paused,
        "tasklytic_workspace_id": row.tasklytic_workspace_id, "tasklytic_project_id": row.tasklytic_project_id,
        "revision": row.revision, "published_at": row.published_at, "completed_at": row.completed_at,
        "created_at": row.created_at, "updated_at": row.updated_at,
        "progress": progress, "request_count": len(requests), "status_counts": dict(statuses),
    }
    if detail:
        result["contacts"] = contacts
        result["requests"] = [serialize_request(db, item, include_internal=include_internal) for item in requests]
        result["activity"] = [{
            "id": str(event.id), "event_type": event.event_type, "actor_kind": event.actor_kind,
            "actor_id": event.actor_id, "request_id": str(event.request_id) if event.request_id else None,
            "details": event.details or {}, "created_at": event.created_at,
        } for event in db.query(PbcAuditEvent).filter(PbcAuditEvent.engagement_id == row.id).order_by(PbcAuditEvent.created_at.desc()).limit(200)]
    return result


def validate_tasklytic_link(db: Session, user_id: str, workspace_id: str | None, project_id: str | None) -> None:
    if not workspace_id and not project_id:
        return
    if not workspace_id or not project_id:
        raise HTTPException(status_code=422, detail="Both Tasklytic workspace and project are required")
    membership = db.get(TasklyticWorkspaceMember, (workspace_id, user_id))
    project = db.query(TasklyticEntityRecord).filter_by(
        workspace_id=workspace_id, entity_kind="projects", record_id=project_id
    ).first()
    if membership is None or project is None:
        raise HTTPException(status_code=403, detail="Project Management link is not accessible")
    authorize_record(db, "projects", project.payload or {}, workspace_id, user_id)


def transition_request(
    db: Session,
    row: PbcRequest,
    engagement: PbcEngagement,
    action: str,
    reason: str | None,
    expected_revision: int,
    *,
    actor_kind: str,
    actor_id: str,
    actor_role_value: str | None = None,
) -> PbcRequest:
    if row.revision != expected_revision:
        raise HTTPException(status_code=409, detail="Request changed; refresh and try again")
    if action not in REQUEST_TRANSITIONS:
        raise HTTPException(status_code=422, detail="Unsupported PBC transition")
    if actor_kind == "client" and action != "submit":
        raise HTTPException(status_code=403, detail="Clients may only submit requests")
    if actor_kind == "firm":
        allowed = FIRM_REVIEW_ROLES if action in {"accept", "return"} else FIRM_MANAGE_ROLES if action in {"waive", "reopen"} else FIRM_WRITE_ROLES
        if actor_role_value not in allowed:
            raise HTTPException(status_code=403, detail="Insufficient PBC permission for this transition")
    allowed_from, target = REQUEST_TRANSITIONS[action]
    if row.status not in allowed_from:
        raise HTTPException(status_code=409, detail=f"Cannot {action} a request in {row.status} status")
    if action == "submit":
        available = db.query(PbcDocument).filter(
            PbcDocument.request_id == row.id, PbcDocument.state == "available"
        ).count()
        if available == 0:
            raise HTTPException(status_code=409, detail="Upload at least one available document before submitting")
        row.submitted_at = utcnow()
    if action == "accept":
        row.accepted_at = utcnow()
    elif action in {"return", "reopen"}:
        row.accepted_at = None
    previous = row.status
    row.status = target
    row.status_reason = reason
    row.revision += 1
    audit(db, engagement, f"request_{action}", actor_kind=actor_kind, actor_id=actor_id,
          request_id=row.id, details={"from": previous, "to": target, "reason": reason})
    return row


def create_access_token(
    db: Session,
    engagement: PbcEngagement,
    contact: PbcContact,
    *,
    purpose: str,
    request_ids: list[str],
    expires_in_days: int,
    actor_user_id: str,
) -> tuple[PbcAccessToken, str]:
    membership = db.get(PbcEngagementContact, (engagement.id, contact.id))
    if membership is None:
        raise HTTPException(status_code=422, detail="Contact is not assigned to this engagement")
    known_ids = {str(value[0]) for value in db.query(PbcRequest.id).filter(PbcRequest.engagement_id == engagement.id)}
    if any(item not in known_ids for item in request_ids):
        raise HTTPException(status_code=422, detail="A scoped request does not belong to this engagement")
    raw = secrets.token_urlsafe(48)
    row = PbcAccessToken(
        engagement_id=engagement.id, contact_id=contact.id, token_hash=token_hash(raw), purpose=purpose,
        request_ids=request_ids if purpose == "scoped_link" else [], expires_at=utcnow() + timedelta(days=expires_in_days),
        one_time=purpose == "portal_login", created_by_user_id=actor_user_id,
    )
    db.add(row)
    audit(db, engagement, "access_link_created", actor_kind="firm", actor_id=actor_user_id,
          details={"contact_id": str(contact.id), "purpose": purpose, "request_count": len(request_ids)})
    return row, raw


def exchange_access_token(db: Session, raw: str) -> tuple[PbcPortalSession, str, PbcContact, PbcEngagement]:
    row = db.query(PbcAccessToken).filter(PbcAccessToken.token_hash == token_hash(raw)).with_for_update().first()
    now = utcnow()
    if row is None or row.revoked_at or row.expires_at <= now or (row.one_time and row.used_at):
        raise HTTPException(status_code=401, detail="This PBC access link is invalid or expired")
    contact = db.get(PbcContact, row.contact_id)
    engagement = db.get(PbcEngagement, row.engagement_id)
    if contact is None or engagement is None or not contact.active or engagement.status not in {"active", "completed"}:
        raise HTTPException(status_code=403, detail="PBC portal access is unavailable")
    raw_session = secrets.token_urlsafe(48)
    raw_csrf = secrets.token_urlsafe(32)
    session = PbcPortalSession(
        engagement_id=row.engagement_id, contact_id=row.contact_id, session_hash=token_hash(raw_session),
        csrf_hash=token_hash(raw_csrf), request_ids=row.request_ids or [], expires_at=now + timedelta(hours=12),
    )
    db.add(session)
    row.used_at = now
    audit(db, engagement, "portal_session_created", actor_kind="client", actor_id=str(contact.id),
          details={"purpose": row.purpose})
    return session, raw_session, contact, engagement


def get_portal_session(db: Session, raw_session: str | None) -> tuple[PbcPortalSession, PbcContact, PbcEngagement]:
    if not raw_session:
        raise HTTPException(status_code=401, detail="PBC portal session required")
    row = db.query(PbcPortalSession).filter(PbcPortalSession.session_hash == token_hash(raw_session)).first()
    now = utcnow()
    if row is None or row.revoked_at or row.expires_at <= now:
        raise HTTPException(status_code=401, detail="PBC portal session expired")
    contact = db.get(PbcContact, row.contact_id)
    engagement = db.get(PbcEngagement, row.engagement_id)
    if contact is None or engagement is None or not contact.active:
        raise HTTPException(status_code=401, detail="PBC portal session unavailable")
    row.last_seen_at = now
    return row, contact, engagement


def portal_request_ids(db: Session, session: PbcPortalSession) -> set[str]:
    if session.request_ids:
        return {str(item) for item in session.request_ids}
    membership = db.get(PbcEngagementContact, (session.engagement_id, session.contact_id))
    if membership and membership.role == "coordinator":
        return {str(item[0]) for item in db.query(PbcRequest.id).filter(PbcRequest.engagement_id == session.engagement_id)}
    return {str(item[0]) for item in db.query(PbcRequestAssignment.request_id).filter(
        PbcRequestAssignment.contact_id == session.contact_id
    )}


def require_portal_request(db: Session, session: PbcPortalSession, request_id: str, *, lock: bool = False) -> PbcRequest:
    if request_id not in portal_request_ids(db, session):
        raise HTTPException(status_code=404, detail="PBC request not found")
    query = db.query(PbcRequest).filter(PbcRequest.id == request_id, PbcRequest.engagement_id == session.engagement_id)
    if lock:
        query = query.with_for_update()
    row = query.first()
    if row is None or row.status == "draft":
        raise HTTPException(status_code=404, detail="PBC request not found")
    return row


def queue_reminders(db: Session, *, as_of: date | None = None) -> int:
    queued = 0
    active_requests = (
        db.query(PbcRequest, PbcEngagement)
        .join(PbcEngagement, PbcEngagement.id == PbcRequest.engagement_id)
        .filter(PbcEngagement.status == "active", PbcEngagement.reminders_paused.is_(False),
                PbcRequest.status.in_(["open", "needs_changes"]), PbcRequest.due_date.isnot(None))
        .all()
    )
    for request_row, engagement in active_requests:
        settings = get_settings(db, engagement.firm_id)
        if as_of is not None:
            today = as_of
        else:
            try:
                today = utcnow().astimezone(ZoneInfo(settings.timezone)).date()
            except ZoneInfoNotFoundError:
                today = utcnow().date()
        delta = (request_row.due_date - today).days
        kind = None
        if delta == settings.reminder_days_before:
            kind = "due_soon"
        elif delta == 0:
            kind = "due_today"
        elif delta < 0 and abs(delta) % settings.overdue_interval_days == 0:
            kind = "overdue"
        if not kind:
            continue
        contacts = (
            db.query(PbcContact)
            .join(PbcRequestAssignment, PbcRequestAssignment.contact_id == PbcContact.id)
            .filter(PbcRequestAssignment.request_id == request_row.id, PbcContact.active.is_(True))
            .all()
        )
        if not contacts:
            contacts = (
                db.query(PbcContact)
                .join(PbcEngagementContact, PbcEngagementContact.contact_id == PbcContact.id)
                .filter(PbcEngagementContact.engagement_id == engagement.id,
                        PbcEngagementContact.role == "coordinator", PbcContact.active.is_(True)).all()
            )
        for contact in contacts:
            key = f"{request_row.id}:{contact.id}:{kind}:{today.isoformat()}"
            if db.query(PbcNotificationOutbox.id).filter(PbcNotificationOutbox.dedupe_key == key).first():
                continue
            db.add(PbcNotificationOutbox(
                engagement_id=engagement.id, request_id=request_row.id, contact_id=contact.id,
                dedupe_key=key, kind=kind, recipient_email=contact.email,
                payload={"contact_name": contact.name, "engagement": engagement.name,
                         "request_number": request_row.request_number, "request_title": request_row.title,
                         "due_date": request_row.due_date.isoformat()},
            ))
            queued += 1
    return queued


def queue_contact_notifications(
    db: Session,
    engagement: PbcEngagement,
    *,
    kind: str,
    request_row: PbcRequest | None = None,
    details: dict[str, Any] | None = None,
) -> int:
    contacts: list[PbcContact]
    if request_row is not None:
        contacts = (
            db.query(PbcContact)
            .join(PbcRequestAssignment, PbcRequestAssignment.contact_id == PbcContact.id)
            .filter(PbcRequestAssignment.request_id == request_row.id, PbcContact.active.is_(True))
            .all()
        )
        coordinators = (
            db.query(PbcContact)
            .join(PbcEngagementContact, PbcEngagementContact.contact_id == PbcContact.id)
            .filter(PbcEngagementContact.engagement_id == engagement.id,
                    PbcEngagementContact.role == "coordinator", PbcContact.active.is_(True)).all()
        )
        contacts = list({str(item.id): item for item in contacts + coordinators}.values())
    else:
        contacts = (
            db.query(PbcContact)
            .join(PbcEngagementContact, PbcEngagementContact.contact_id == PbcContact.id)
            .filter(PbcEngagementContact.engagement_id == engagement.id, PbcContact.active.is_(True)).all()
        )
    nonce = str((details or {}).get("nonce") or int(utcnow().timestamp() * 1_000_000))
    queued = 0
    for contact in contacts:
        key = f"event:{engagement.id}:{request_row.id if request_row else 'engagement'}:{contact.id}:{kind}:{nonce}"
        db.add(PbcNotificationOutbox(
            engagement_id=engagement.id,
            request_id=request_row.id if request_row else None,
            contact_id=contact.id,
            dedupe_key=key,
            kind=kind,
            recipient_email=contact.email,
            payload={
                "contact_name": contact.name,
                "engagement": engagement.name,
                "request_number": request_row.request_number if request_row else None,
                "request_title": request_row.title if request_row else None,
                **(details or {}),
            },
        ))
        queued += 1
    return queued


def deliver_notifications(db: Session, *, limit: int = 100) -> dict[str, int]:
    now = utcnow()
    rows = db.query(PbcNotificationOutbox).filter(
        PbcNotificationOutbox.status.in_(["pending", "failed"]), PbcNotificationOutbox.next_attempt_at <= now
    ).order_by(PbcNotificationOutbox.created_at).limit(limit).all()
    sent = failed = 0
    for row in rows:
        data = row.payload or {}
        if row.kind == "published":
            subject = f"PBC request list is ready: {data.get('engagement')}"
            body = f"Hello {data.get('contact_name') or ''},\n\nThe PBC request list for {data.get('engagement')} is ready. Open your secure portal link to begin.\n\n— CPAAutomation"
        elif row.kind == "returned":
            subject = f"Changes requested: {data.get('request_number')} {data.get('request_title')}"
            body = f"Hello {data.get('contact_name') or ''},\n\nThe firm requested changes to {data.get('request_number')} — {data.get('request_title')}.\n\nReviewer note: {data.get('reason') or 'Please review the portal message.'}\n\nOpen your secure portal link to respond.\n\n— CPAAutomation"
        elif row.kind == "comment":
            subject = f"New PBC message: {data.get('request_number')} {data.get('request_title')}"
            body = f"Hello {data.get('contact_name') or ''},\n\nThere is a new message on {data.get('request_number')} — {data.get('request_title')}.\n\nOpen your secure portal link to view it.\n\n— CPAAutomation"
        else:
            subject = f"PBC reminder: {data.get('request_number')} {data.get('request_title')}"
            body = (
                f"Hello {data.get('contact_name') or ''},\n\n"
                f"This is a reminder for {data.get('engagement')}:\n"
                f"{data.get('request_number')} — {data.get('request_title')}\n"
                f"Due: {data.get('due_date')}\n\n"
                "Open your most recent secure PBC portal link to respond.\n\n— CPAAutomation"
            )
        row.attempt_count += 1
        if email_service.send_email(row.recipient_email, subject, body):
            row.status = "sent"
            row.sent_at = now
            row.last_error = None
            sent += 1
        else:
            row.status = "failed"
            row.last_error = "Email provider returned a delivery failure"
            row.next_attempt_at = now + timedelta(minutes=min(60, 2 ** row.attempt_count))
            failed += 1
    return {"sent": sent, "failed": failed}


def build_tracker_xlsx(db: Session, engagement: PbcEngagement) -> bytes:
    requests = db.query(PbcRequest).filter(PbcRequest.engagement_id == engagement.id).order_by(PbcRequest.sort_order).all()
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    counts = Counter(row.status for row in requests)
    summary.append(["PBC Engagement", engagement.name])
    summary.append(["Client", engagement.client_name_snapshot])
    summary.append(["Period End", engagement.period_end.isoformat() if engagement.period_end else ""])
    summary.append(["Total Requests", len(requests)])
    summary.append(["Accepted / Waived", counts["accepted"] + counts["waived"]])
    summary.append(["Completion %", round(((counts["accepted"] + counts["waived"]) / len(requests)) * 100, 1) if requests else 0])
    status = wb.create_sheet("Status")
    headers = ["Request ID", "Number", "Category", "Description", "Owner", "Due Date", "Priority", "Status", "Files", "Last Updated"]
    status.append(headers)
    for row in requests:
        owner = db.get(User, row.owner_user_id) if row.owner_user_id else None
        file_count = db.query(PbcDocument).filter(PbcDocument.request_id == row.id, PbcDocument.state == "available").count()
        status.append([str(row.id), _spreadsheet_safe(row.request_number), _spreadsheet_safe(row.category or ""),
                       _spreadsheet_safe(row.title), _spreadsheet_safe(_user_label(owner) or ""),
                       row.due_date.isoformat() if row.due_date else "", row.priority, row.status, file_count,
                       row.updated_at.isoformat() if row.updated_at else ""])
    overdue = wb.create_sheet("Overdue")
    overdue.append(headers)
    for values in status.iter_rows(min_row=2, values_only=True):
        due = values[5]
        if due and date.fromisoformat(due) < utcnow().date() and values[7] not in {"accepted", "waived"}:
            overdue.append(list(values))
    audit_ws = wb.create_sheet("Audit Trail")
    audit_ws.append(["Timestamp", "Event", "Actor Type", "Actor ID", "Request ID", "Details"])
    for event in db.query(PbcAuditEvent).filter(PbcAuditEvent.engagement_id == engagement.id).order_by(PbcAuditEvent.created_at):
        audit_ws.append([event.created_at.isoformat(), event.event_type, event.actor_kind, event.actor_id or "",
                         str(event.request_id) if event.request_id else "", str(event.details or {})])
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def build_metadata_zip(db: Session, engagement: PbcEngagement, evidence_files: list[tuple[str, bytes]] | None = None) -> bytes:
    """Build a traceable evidence package with tracker, manifest, and available binaries."""
    requests = db.query(PbcRequest).filter(PbcRequest.engagement_id == engagement.id).order_by(PbcRequest.sort_order).all()
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("pbc_readiness.xlsx", build_tracker_xlsx(db, engagement))
        index_io = io.StringIO()
        writer = csv.writer(index_io)
        writer.writerow(["request_id", "request_number", "category", "status", "filename", "version", "checksum_sha256", "scan_status"])
        for request_row in requests:
            if request_row.status == "waived":
                continue
            docs = db.query(PbcDocument).filter(
                PbcDocument.request_id == request_row.id, PbcDocument.state == "available"
            ).order_by(PbcDocument.version).all()
            if not docs:
                writer.writerow([str(request_row.id), _spreadsheet_safe(request_row.request_number),
                                 _spreadsheet_safe(request_row.category or ""), request_row.status, "", "", "", ""])
            for document in docs:
                writer.writerow([str(request_row.id), _spreadsheet_safe(request_row.request_number),
                                 _spreadsheet_safe(request_row.category or ""), request_row.status,
                                 _spreadsheet_safe(document.filename), document.version, document.checksum_sha256 or "", document.scan_status])
        archive.writestr("index.csv", index_io.getvalue())
        for path, content in evidence_files or []:
            archive.writestr(path, content)
        archive.writestr("README.txt", "This package contains the PBC tracker, immutable evidence manifest, and available evidence versions at the time of export.\n")
    return output.getvalue()


def cleanup_expired(db: Session) -> dict[str, int]:
    now = utcnow()
    sessions = db.query(PbcPortalSession).filter(PbcPortalSession.expires_at <= now, PbcPortalSession.revoked_at.is_(None)).update(
        {PbcPortalSession.revoked_at: now}, synchronize_session=False
    )
    abandoned = db.query(PbcDocument).filter(
        PbcDocument.state == "initiated", PbcDocument.created_at < now - timedelta(hours=24)
    ).update({PbcDocument.state: "abandoned"}, synchronize_session=False)
    return {"expired_sessions": sessions, "abandoned_uploads": abandoned}
