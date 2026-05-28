"""Close Orchestration Dashboard — dependency-aware close tracking and critical path.

Implements the workflow in ../SKILL.md.
"""
from __future__ import annotations

import argparse
import logging
import sys
from collections import defaultdict, deque
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

LOG = logging.getLogger("close_orchestration")

HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN = Side(border_style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

REQUIRED = ("task_id", "task_name", "owner", "due_date", "status")
COMPLETE_STATUSES = {"complete", "done", "closed"}


@dataclass
class CloseResult:
    total: int
    completed: int
    overdue: int
    blocked: int
    percent_complete: float
    critical_path: list[str] = field(default_factory=list)
    overdue_tasks: pd.DataFrame = field(default_factory=pd.DataFrame)
    blocked_tasks: pd.DataFrame = field(default_factory=pd.DataFrame)
    entity_summary: pd.DataFrame = field(default_factory=pd.DataFrame)


def _read_checklist(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path)
    else:
        df = pd.read_excel(path)
    df.columns = [c.strip().lower() for c in df.columns]
    missing = [c for c in REQUIRED if c not in df.columns]
    if missing:
        raise ValueError(f"Checklist missing required columns: {missing}")
    df = df.copy()
    df["task_id"] = df["task_id"].astype(str)
    df["due_date"] = pd.to_datetime(df["due_date"], errors="coerce")
    df["status_norm"] = df["status"].astype(str).str.strip().str.lower()
    if "dependencies" not in df.columns:
        df["dependencies"] = ""
    else:
        df["dependencies"] = df["dependencies"].fillna("").astype(str)
    if "entity" not in df.columns:
        df["entity"] = "All"
    if "day_target" not in df.columns:
        df["day_target"] = ""
    return df


def _deps(raw: str) -> list[str]:
    if not raw or raw.strip() in ("", "nan"):
        return []
    return [d.strip() for d in raw.split(",") if d.strip()]


def _is_complete(status: str) -> bool:
    return status in COMPLETE_STATUSES


def _build_graph(df: pd.DataFrame) -> tuple[dict[str, list[str]], dict[str, int]]:
    graph: dict[str, list[str]] = defaultdict(list)
    indegree: dict[str, int] = defaultdict(int)
    ids = set(df["task_id"])
    for _, row in df.iterrows():
        tid = row["task_id"]
        for dep in _deps(row["dependencies"]):
            if dep in ids:
                graph[dep].append(tid)
                indegree[tid] += 1
        indegree.setdefault(tid, indegree.get(tid, 0))
    return graph, indegree


def _critical_path(df: pd.DataFrame) -> list[str]:
    graph, indegree = _build_graph(df)
    duration = {row["task_id"]: int(row["day_target"]) if str(row["day_target"]).isdigit() else 1
                for _, row in df.iterrows()}
    # Longest path in DAG via topological DP
    dist: dict[str, int] = {tid: duration.get(tid, 1) for tid in df["task_id"]}
    prev: dict[str, str | None] = {tid: None for tid in df["task_id"]}
    q = deque([tid for tid, deg in indegree.items() if deg == 0])
    while q:
        u = q.popleft()
        for v in graph.get(u, []):
            cand = dist[u] + duration.get(v, 1)
            if cand >= dist.get(v, 0):
                dist[v] = cand
                prev[v] = u
            indegree[v] -= 1
            if indegree[v] == 0:
                q.append(v)
    if not dist:
        return []
    end = max(dist, key=dist.get)
    path: list[str] = []
    cur: str | None = end
    while cur is not None:
        path.append(cur)
        cur = prev.get(cur)
    path.reverse()
    return path


def analyze(df: pd.DataFrame, as_of: datetime) -> CloseResult:
    total = len(df)
    completed = int(df["status_norm"].apply(_is_complete).sum())
    incomplete = df[~df["status_norm"].apply(_is_complete)].copy()
    overdue_mask = incomplete["due_date"] < pd.Timestamp(as_of)
    overdue_df = incomplete[overdue_mask].copy()

    status_map = {row["task_id"]: row["status_norm"] for _, row in df.iterrows()}
    blocked_rows = []
    for _, row in incomplete.iterrows():
        deps = _deps(row["dependencies"])
        unmet = [d for d in deps if not _is_complete(status_map.get(d, ""))]
        if unmet:
            blocked_rows.append({
                "task_id": row["task_id"],
                "task_name": row["task_name"],
                "owner": row["owner"],
                "blocked_by": ", ".join(unmet),
                "due_date": row["due_date"],
            })
    blocked_df = pd.DataFrame(blocked_rows)

    entity_summary = (
        df.groupby("entity", dropna=False)
        .apply(lambda g: pd.Series({
            "total": len(g),
            "completed": g["status_norm"].apply(_is_complete).sum(),
            "pct_complete": round(g["status_norm"].apply(_is_complete).mean() * 100, 1),
        }), include_groups=False)
        .reset_index()
    )

    return CloseResult(
        total=total,
        completed=completed,
        overdue=len(overdue_df),
        blocked=len(blocked_df),
        percent_complete=round(completed / total * 100, 1) if total else 0.0,
        critical_path=_critical_path(df),
        overdue_tasks=overdue_df,
        blocked_tasks=blocked_df,
        entity_summary=entity_summary,
    )


def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX


def _autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(45, max_len + 2)


def write_dashboard(df: pd.DataFrame, result: CloseResult, as_of: datetime, output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Close Orchestration Dashboard"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["As Of", as_of.strftime("%Y-%m-%d")])
    ws.append(["Total Tasks", result.total])
    ws.append(["Completed", result.completed])
    ws.append(["Percent Complete", f"{result.percent_complete:.1f}%"])
    ws.append(["Overdue", result.overdue])
    ws.append(["Blocked", result.blocked])
    ws.append(["Critical Path", " → ".join(result.critical_path)])
    _autosize(ws)

    ws_t = wb.create_sheet("Tasks")
    cols = ["task_id", "task_name", "owner", "entity", "due_date", "status", "dependencies", "day_target"]
    ws_t.append(cols)
    _style_header(ws_t)
    for _, row in df.iterrows():
        ws_t.append([row.get(c, "") for c in cols])
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        status = str(getattr(row, "status", "")).lower()
        if status not in COMPLETE_STATUSES and getattr(row, "due_date") < pd.Timestamp(as_of):
            for cell in ws_t[r_idx]:
                cell.fill = RED_FILL
        elif status not in COMPLETE_STATUSES:
            for cell in ws_t[r_idx]:
                cell.fill = YELLOW_FILL
    _autosize(ws_t)

    ws_o = wb.create_sheet("Overdue")
    ws_o.append(["task_id", "task_name", "owner", "due_date", "status"])
    _style_header(ws_o)
    if not result.overdue_tasks.empty:
        for _, r in result.overdue_tasks.iterrows():
            ws_o.append([r["task_id"], r["task_name"], r["owner"],
                         r["due_date"].strftime("%Y-%m-%d"), r["status"]])
    else:
        ws_o.append(["(none)"])
    _autosize(ws_o)

    ws_b = wb.create_sheet("Blocked")
    ws_b.append(["task_id", "task_name", "owner", "blocked_by", "due_date"])
    _style_header(ws_b)
    if not result.blocked_tasks.empty:
        for _, r in result.blocked_tasks.iterrows():
            ws_b.append([
                r["task_id"], r["task_name"], r["owner"], r["blocked_by"],
                r["due_date"].strftime("%Y-%m-%d") if pd.notna(r["due_date"]) else "",
            ])
    else:
        ws_b.append(["(none)"])
    _autosize(ws_b)

    ws_e = wb.create_sheet("EntitySummary")
    ws_e.append(["entity", "total", "completed", "pct_complete"])
    _style_header(ws_e)
    for _, r in result.entity_summary.iterrows():
        ws_e.append([r["entity"], int(r["total"]), int(r["completed"]), f"{r['pct_complete']:.1f}%"])
    _autosize(ws_e)

    ws_cp = wb.create_sheet("CriticalPath")
    ws_cp.append(["Sequence", "Task ID"])
    _style_header(ws_cp)
    for i, tid in enumerate(result.critical_path, start=1):
        name = df.loc[df["task_id"] == tid, "task_name"]
        ws_cp.append([i, f"{tid} — {name.iloc[0] if not name.empty else ''}"])
    _autosize(ws_cp)

    ws_at = wb.create_sheet("AuditTrail")
    ws_at.append(["Key", "Value"])
    _style_header(ws_at)
    ws_at.append(["Generated", datetime.now().isoformat(timespec="seconds")])
    ws_at.append(["As Of", as_of.strftime("%Y-%m-%d")])
    _autosize(ws_at)

    ws_so = wb.create_sheet("SignOff")
    ws_so.append(["Role", "Name", "Title", "Date", "Notes"])
    _style_header(ws_so)
    for role in ("Close Manager", "Controller", "CFO"):
        ws_so.append([role, "", "", "", ""])
    _autosize(ws_so)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    LOG.info("Dashboard written: %s", output)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--checklist", required=True, type=Path)
    parser.add_argument("--as-of", required=True, help="YYYY-MM-DD")
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--quiet", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.WARNING if args.quiet else logging.INFO,
        format="%(levelname)s %(name)s: %(message)s",
    )

    try:
        df = _read_checklist(args.checklist)
    except ValueError as exc:
        LOG.error("%s", exc)
        return 1

    as_of = datetime.strptime(args.as_of, "%Y-%m-%d")
    result = analyze(df, as_of)
    write_dashboard(df, result, as_of, args.output)

    print(f"Progress:        {result.percent_complete:.1f}% ({result.completed}/{result.total})")
    print(f"Overdue:         {result.overdue}")
    print(f"Blocked:         {result.blocked}")
    print(f"Critical Path:   {' → '.join(result.critical_path)}")
    print(f"Dashboard:       {args.output}")
    return 2 if result.overdue else 0


if __name__ == "__main__":
    sys.exit(main())
