"""Reserves & Allowance Estimator (CECL ASC 326) — pooled-aging method with Q-factors and forecast overlay.

Implements the workflow in ../SKILL.md.
"""
from __future__ import annotations

import argparse
import logging
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

LOG = logging.getLogger("reserves_estimator")

ACCOUNTING_FMT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN = Side(border_style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STANDARD_BUCKETS = ("Current", "31-60", "61-90", "Over 90")


def _read_table(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path)
    else:
        df = pd.read_excel(path)
    return df


def _detect_buckets(df: pd.DataFrame) -> list[str]:
    return [c for c in df.columns if c in STANDARD_BUCKETS or c in (
        "0-30", "1-30", "91-180", "Over 180", "121-150", "151-180")]


def aggregate_aging(df: pd.DataFrame, buckets: list[str]) -> dict[str, float]:
    pools = {}
    for b in buckets:
        pools[b] = float(pd.to_numeric(df[b], errors="coerce").fillna(0).clip(lower=0).sum())
    return pools


def build_rates(buckets: list[str], loss_rates: pd.DataFrame,
                q_factors: pd.DataFrame | None,
                forecast: pd.DataFrame | None) -> dict[str, dict[str, float]]:
    rates: dict[str, dict[str, float]] = {}
    lr_map = {str(r["bucket"]): float(r["rate"]) for _, r in loss_rates.iterrows()}
    qf_map = {str(r["bucket"]): float(r["q_factor_pct"]) for _, r in q_factors.iterrows()} if q_factors is not None else {}
    fc_map = {str(r["bucket"]): float(r["multiplier"]) for _, r in forecast.iterrows()} if forecast is not None else {}
    for b in buckets:
        hist = lr_map.get(b, 0.0)
        qf = qf_map.get(b, 0.0)
        fc = fc_map.get(b, 1.0)
        final = (hist + qf) * fc
        rates[b] = {"historical": hist, "q_factor": qf, "forecast_multiplier": fc, "final": final}
    return rates


def compute_acl(pools: dict[str, float], rates: dict[str, dict[str, float]]) -> tuple[float, list[dict]]:
    detail = []
    total = 0.0
    for b, pool in pools.items():
        rate = rates[b]["final"]
        acl = pool * rate
        total += acl
        detail.append({
            "Bucket": b, "Pool Amount": pool,
            "Historical Rate": rates[b]["historical"],
            "Q-Factor": rates[b]["q_factor"],
            "Forecast Multiplier": rates[b]["forecast_multiplier"],
            "Final Rate": rates[b]["final"],
            "ACL": round(acl, 2),
        })
    return round(total, 2), detail


def scenario_acl(pools: dict[str, float], rates: dict[str, dict[str, float]],
                 scenario_label: str, multiplier: float) -> dict:
    rows = []
    total = 0.0
    for b, pool in pools.items():
        rate = rates[b]["final"] * multiplier
        acl = pool * rate
        total += acl
        rows.append({"Bucket": b, "Pool": pool, "Final Rate": rate, "ACL": round(acl, 2)})
    return {"label": scenario_label, "multiplier": multiplier, "total": round(total, 2), "detail": rows}


def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX


def _autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(40, max_len + 2)


def write_workpaper(
    pools: dict[str, float], rates: dict[str, dict[str, float]],
    acl_detail: list[dict], total_acl: float,
    scenarios: list[dict],
    prior_acl: float | None, writeoffs: float, recoveries: float,
    customer_specific: pd.DataFrame | None,
    aging_path: Path, period_end: datetime, output: Path,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Allowance for Credit Losses (CECL) Workpaper"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    total_ar = sum(pools.values())
    ws.append(["Period End", period_end.strftime("%Y-%m-%d")])
    ws.append(["Total AR", total_ar])
    ws.append(["Total ACL (Base)", total_acl])
    ws.append(["ACL / AR %", f"{(total_acl/total_ar*100) if total_ar else 0:.2f}%"])
    for cell in ("B4", "B5"):
        ws[cell].number_format = ACCOUNTING_FMT
    if prior_acl is not None:
        provision = total_acl - prior_acl + writeoffs - recoveries
        ws.append(["Prior ACL", prior_acl])
        ws.append(["Write-offs", writeoffs])
        ws.append(["Recoveries", recoveries])
        ws.append(["Period Provision", round(provision, 2)])
        for cell in ("B7", "B8", "B9", "B10"):
            ws[cell].number_format = ACCOUNTING_FMT
    _autosize(ws)

    # Aging Pools
    ws_a = wb.create_sheet("AgingPools")
    ws_a.append(["Bucket", "Pool Amount", "Historical Rate", "Q-Factor", "Forecast Mult", "Final Rate", "ACL"])
    _style_header(ws_a)
    for r in acl_detail:
        ws_a.append([r["Bucket"], r["Pool Amount"], r["Historical Rate"], r["Q-Factor"],
                     r["Forecast Multiplier"], r["Final Rate"], r["ACL"]])
    for letter in ("B", "G"):
        for cell in ws_a[letter][1:]:
            cell.number_format = ACCOUNTING_FMT
    for letter in ("C", "D", "F"):
        for cell in ws_a[letter][1:]:
            cell.number_format = "0.0000"
    _autosize(ws_a)

    # Loss Rate Build (showing the formula assembly)
    ws_lr = wb.create_sheet("LossRateBuild")
    ws_lr.append(["Bucket", "Historical (Pillar 1)", "+ Q-Factor (Pillar 2)", "× Forecast Mult (Pillar 3)", "= Final Rate"])
    _style_header(ws_lr)
    for b, r in rates.items():
        ws_lr.append([b, r["historical"], r["q_factor"], r["forecast_multiplier"], r["final"]])
    for letter in ("B", "C", "E"):
        for cell in ws_lr[letter][1:]:
            cell.number_format = "0.0000"
    _autosize(ws_lr)

    # Scenarios
    ws_s = wb.create_sheet("Scenarios")
    if scenarios:
        cols = ["Scenario", "Multiplier", "Total ACL", "Δ vs Base"]
        ws_s.append(cols)
        _style_header(ws_s)
        for sc in scenarios:
            ws_s.append([sc["label"], sc["multiplier"], sc["total"],
                         round(sc["total"] - total_acl, 2)])
        for letter in ("C", "D"):
            for cell in ws_s[letter][1:]:
                cell.number_format = ACCOUNTING_FMT
    _autosize(ws_s)

    # RollForward
    ws_rf = wb.create_sheet("RollForward")
    ws_rf.append(["Line", "Amount"])
    _style_header(ws_rf)
    if prior_acl is not None:
        provision = total_acl - prior_acl + writeoffs - recoveries
        ws_rf.append(["Beginning ACL", prior_acl])
        ws_rf.append(["Provision (computed)", round(provision, 2)])
        ws_rf.append(["Write-offs", -writeoffs])
        ws_rf.append(["Recoveries", recoveries])
        ws_rf.append(["Ending ACL", total_acl])
        for cell in ws_rf["B"][1:]:
            cell.number_format = ACCOUNTING_FMT
    else:
        ws_rf.append(["(no prior ACL supplied — provide --prior-acl-balance)"])
    _autosize(ws_rf)

    # DraftJE
    ws_je = wb.create_sheet("DraftJE")
    ws_je.append(["Date", "Account", "Description", "Debit", "Credit"])
    _style_header(ws_je)
    period_str = period_end.strftime("%Y-%m-%d")
    if prior_acl is not None:
        provision = total_acl - prior_acl + writeoffs - recoveries
    else:
        provision = total_acl
    ws_je.append([period_str, "7200 Provision for Credit Losses", "CECL provision", round(provision, 2), 0])
    ws_je.append([period_str, "1199 Allowance for Credit Losses (Contra)", "CECL provision", 0, round(provision, 2)])
    for letter in ("D", "E"):
        for cell in ws_je[letter][1:]:
            cell.number_format = ACCOUNTING_FMT
    _autosize(ws_je)

    # CustomerSpecific
    ws_cs = wb.create_sheet("CustomerSpecific")
    if customer_specific is not None and not customer_specific.empty:
        ws_cs.append(list(customer_specific.columns))
        _style_header(ws_cs)
        for _, r in customer_specific.iterrows():
            ws_cs.append([r[c] for c in customer_specific.columns])
    else:
        ws_cs.append(["(no customer-specific reserves)"])
    _autosize(ws_cs)

    # AuditTrail
    ws_at = wb.create_sheet("AuditTrail")
    ws_at.append(["Key", "Value"])
    _style_header(ws_at)
    ws_at.append(["Aging Input", str(aging_path)])
    ws_at.append(["Period End", period_str])
    ws_at.append(["Total AR", round(total_ar, 2)])
    ws_at.append(["Total ACL", total_acl])
    ws_at.append(["Generated", datetime.now().isoformat(timespec="seconds")])
    _autosize(ws_at)

    # SignOff
    ws_so = wb.create_sheet("SignOff")
    ws_so.append(["Role", "Name", "Title", "Date", "Notes"])
    _style_header(ws_so)
    ws_so.append(["Preparer (AR Accountant)", "", "", "", ""])
    ws_so.append(["Reviewer (Senior Accountant)", "", "", "", ""])
    ws_so.append(["Approver (Controller)", "", "", "", ""])
    _autosize(ws_so)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    LOG.info("Workpaper written: %s", output)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--aging", required=True, type=Path)
    parser.add_argument("--loss-rates", required=True, type=Path)
    parser.add_argument("--q-factors", type=Path, default=None)
    parser.add_argument("--forecast", type=Path, default=None)
    parser.add_argument("--prior-acl-balance", type=float, default=None)
    parser.add_argument("--writeoffs", type=float, default=0.0)
    parser.add_argument("--recoveries", type=float, default=0.0)
    parser.add_argument("--scenarios", default="stress_up_25,stress_down_25")
    parser.add_argument("--period-end", required=True)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--quiet", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.WARNING if args.quiet else logging.INFO,
        format="%(levelname)s %(name)s: %(message)s",
    )

    aging_df = _read_table(args.aging)
    buckets = _detect_buckets(aging_df)
    if not buckets:
        LOG.error("Could not detect aging-bucket columns; expected one of: %s", STANDARD_BUCKETS)
        return 1
    loss_rates = _read_table(args.loss_rates)
    if "bucket" not in loss_rates.columns or "rate" not in loss_rates.columns:
        LOG.error("--loss-rates must have columns: bucket, rate")
        return 1
    q_factors = _read_table(args.q_factors) if args.q_factors else None
    forecast = _read_table(args.forecast) if args.forecast else None

    period_end = datetime.strptime(args.period_end, "%Y-%m-%d")
    pools = aggregate_aging(aging_df, buckets)
    rates = build_rates(buckets, loss_rates, q_factors, forecast)
    total_acl, acl_detail = compute_acl(pools, rates)

    # Scenarios
    scenarios: list[dict] = []
    for label in [s.strip() for s in args.scenarios.split(",") if s.strip()]:
        if label.startswith("stress_up_"):
            pct = int(label.split("_")[-1]) / 100.0
            mult = 1.0 + pct
        elif label.startswith("stress_down_"):
            pct = int(label.split("_")[-1]) / 100.0
            mult = max(1.0 - pct, 0.0)
        else:
            mult = 1.0
        scenarios.append(scenario_acl(pools, rates, label, mult))

    # Customer-specific reserves
    customer_specific = None
    if "customer_specific_reserve" in aging_df.columns:
        cs = aging_df[["customer", "customer_specific_reserve"]].copy()
        cs["customer_specific_reserve"] = pd.to_numeric(cs["customer_specific_reserve"], errors="coerce").fillna(0)
        cs = cs[cs["customer_specific_reserve"] > 0]
        if not cs.empty:
            customer_specific = cs
            total_acl += float(cs["customer_specific_reserve"].sum())

    write_workpaper(
        pools=pools, rates=rates, acl_detail=acl_detail, total_acl=total_acl,
        scenarios=scenarios, prior_acl=args.prior_acl_balance,
        writeoffs=args.writeoffs, recoveries=args.recoveries,
        customer_specific=customer_specific,
        aging_path=args.aging, period_end=period_end, output=args.output,
    )

    print(f"Buckets:       {buckets}")
    print(f"Total AR:      {sum(pools.values()):>15,.2f}")
    print(f"Total ACL:     {total_acl:>15,.2f}")
    print(f"Workpaper:     {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
