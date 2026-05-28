"""Tax Provision Calculator — ASC 740 current/deferred tax, ETR reconciliation, DTA/DTL roll-forward.

Implements the workflow in ../SKILL.md.
"""
from __future__ import annotations

import argparse
import logging
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

LOG = logging.getLogger("tax_provision")

ACCOUNTING_FMT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
PCT_FMT = "0.00%"
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN = Side(border_style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


@dataclass
class ProvisionResult:
    pre_tax_income: float
    federal_rate: float
    state_rate: float
    blended_rate: float
    permanent_diffs: list[dict[str, Any]]
    temporary_diffs: list[dict[str, Any]]
    taxable_income: float
    current_tax: float
    deferred_tax: float
    total_tax_expense: float
    etr: float
    etr_lines: list[dict[str, Any]] = field(default_factory=list)
    dta_dtl: list[dict[str, Any]] = field(default_factory=list)
    valuation_allowance: float = 0.0
    net_dta: float = 0.0
    utp_flagged: list[str] = field(default_factory=list)


def _load_inputs(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as fh:
        data = yaml.safe_load(fh)
    if not isinstance(data, dict):
        raise ValueError("Provision inputs must be a YAML mapping")
    return data


def _normalize_diffs(items: Any, diff_type: str) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    if not items:
        return out
    for item in items:
        if isinstance(item, dict):
            for desc, amount in item.items():
                out.append({"description": str(desc), "amount": float(amount), "type": diff_type})
        elif isinstance(item, str):
            out.append({"description": item, "amount": 0.0, "type": diff_type})
    return out


def calculate(data: dict[str, Any]) -> ProvisionResult:
    pre_tax = float(data.get("pre_tax_income", 0))
    federal = float(data.get("statutory_rate", data.get("federal_rate", 0.21)))
    state = float(data.get("state_rate", 0.0))
    apportionment = float(data.get("state_apportionment_pct", 1.0))
    blended = federal + state * apportionment

    perm = _normalize_diffs(data.get("permanent_differences"), "permanent")
    temp = _normalize_diffs(data.get("temporary_differences"), "temporary")
    perm_total = sum(d["amount"] for d in perm)
    temp_total = sum(d["amount"] for d in temp)

    taxable = pre_tax + perm_total + temp_total
    current_tax = max(0.0, taxable * blended)
    deferred_tax = -temp_total * blended
    total = current_tax + deferred_tax
    etr = total / pre_tax if pre_tax else 0.0

    etr_lines = [
        {"line": "Federal statutory rate", "rate": federal, "amount": pre_tax * federal},
        {"line": "State tax (net of federal benefit)", "rate": state * apportionment * (1 - federal),
         "amount": pre_tax * state * apportionment * (1 - federal)},
    ]
    for d in perm:
        tax_impact = d["amount"] * blended
        etr_lines.append({
            "line": f"Permanent — {d['description']}",
            "rate": tax_impact / pre_tax if pre_tax else 0,
            "amount": tax_impact,
        })

    dta_dtl = []
    for d in temp:
        deferred = -d["amount"] * blended
        dta_dtl.append({
            "description": d["description"],
            "temporary_difference": d["amount"],
            "deferred_tax": deferred,
            "classification": "DTA" if deferred < 0 else "DTL",
        })

    gross_dta = sum(abs(x["deferred_tax"]) for x in dta_dtl if x["classification"] == "DTA")
    va_pct = float(data.get("valuation_allowance_pct", 0.0))
    va = gross_dta * va_pct
    net_dta = gross_dta - va

    utp = [str(x) for x in (data.get("uncertain_tax_positions") or [])]

    return ProvisionResult(
        pre_tax_income=pre_tax, federal_rate=federal, state_rate=state,
        blended_rate=blended, permanent_diffs=perm, temporary_diffs=temp,
        taxable_income=taxable, current_tax=current_tax, deferred_tax=deferred_tax,
        total_tax_expense=total, etr=etr, etr_lines=etr_lines, dta_dtl=dta_dtl,
        valuation_allowance=va, net_dta=net_dta, utp_flagged=utp,
    )


def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX


def _autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(45, max_len + 2)


def write_workpaper(result: ProvisionResult, period_end: str, output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Income Tax Provision Workpaper (ASC 740)"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Period End", period_end])
    ws.append(["Pre-Tax Book Income", result.pre_tax_income])
    ws.append(["Blended Tax Rate", result.blended_rate])
    ws.append(["Current Tax Expense", result.current_tax])
    ws.append(["Deferred Tax Benefit/(Expense)", result.deferred_tax])
    ws.append(["Total Tax Expense", result.total_tax_expense])
    ws.append(["Effective Tax Rate", result.etr])
    ws.append(["Valuation Allowance", result.valuation_allowance])
    ws.append(["Net DTA", result.net_dta])
    for cell in ("B4", "B6", "B7", "B8", "B9", "B11", "B12"):
        ws[cell].number_format = ACCOUNTING_FMT
    ws["B5"].number_format = PCT_FMT
    ws["B10"].number_format = PCT_FMT
    _autosize(ws)

    ws_etr = wb.create_sheet("ETR_Reconciliation")
    ws_etr.append(["Line", "Rate Impact", "Tax Amount"])
    _style_header(ws_etr)
    for line in result.etr_lines:
        ws_etr.append([line["line"], line["rate"], line["amount"]])
    ws_etr.append(["Effective Tax Rate", result.etr, result.total_tax_expense])
    for letter in ("B",):
        for cell in ws_etr[letter][1:]:
            cell.number_format = PCT_FMT
    for letter in ("C",):
        for cell in ws_etr[letter][1:]:
            cell.number_format = ACCOUNTING_FMT
    _autosize(ws_etr)

    ws_d = wb.create_sheet("DTA_DTL")
    ws_d.append(["Description", "Temporary Difference", "Deferred Tax", "Classification"])
    _style_header(ws_d)
    for row in result.dta_dtl:
        ws_d.append([row["description"], row["temporary_difference"], row["deferred_tax"], row["classification"]])
    for letter in ("B", "C"):
        for cell in ws_d[letter][1:]:
            cell.number_format = ACCOUNTING_FMT
    _autosize(ws_d)

    ws_rf = wb.create_sheet("DeferredRollForward")
    ws_rf.append(["Item", "Amount"])
    _style_header(ws_rf)
    ws_rf.append(["Opening DTA/DTL (input assumed zero)", 0])
    ws_rf.append(["Current-period movement", result.deferred_tax])
    ws_rf.append(["Valuation allowance", -result.valuation_allowance])
    ws_rf.append(["Closing net deferred tax", result.deferred_tax - result.valuation_allowance])
    for cell in ws_rf["B"][1:]:
        cell.number_format = ACCOUNTING_FMT
    _autosize(ws_rf)

    ws_je = wb.create_sheet("DraftJE")
    ws_je.append(["Account", "Debit", "Credit", "Memo"])
    _style_header(ws_je)
    ws_je.append(["7400 Income Tax Expense", result.total_tax_expense, 0, "Total provision"])
    ws_je.append(["2400 Income Taxes Payable", 0, result.current_tax, "Current tax"])
    if result.deferred_tax >= 0:
        ws_je.append(["2410 Deferred Tax Liability", 0, result.deferred_tax, "Deferred"])
    else:
        ws_je.append(["1410 Deferred Tax Asset", abs(result.deferred_tax), 0, "Deferred"])
    for letter in ("B", "C"):
        for cell in ws_je[letter][1:]:
            cell.number_format = ACCOUNTING_FMT
    _autosize(ws_je)

    ws_utp = wb.create_sheet("UTP_Flags")
    ws_utp.append(["Uncertain Tax Position"])
    _style_header(ws_utp)
    if result.utp_flagged:
        for u in result.utp_flagged:
            ws_utp.append([u])
    else:
        ws_utp.append(["(none flagged)"])
    _autosize(ws_utp)

    ws_at = wb.create_sheet("AuditTrail")
    ws_at.append(["Key", "Value"])
    _style_header(ws_at)
    ws_at.append(["Generated", datetime.now().isoformat(timespec="seconds")])
    _autosize(ws_at)

    ws_so = wb.create_sheet("SignOff")
    ws_so.append(["Role", "Name", "Title", "Date", "Notes"])
    _style_header(ws_so)
    for role in ("Preparer (Tax Accountant)", "Reviewer (Tax Manager)", "Approver (Controller)"):
        ws_so.append([role, "", "", "", ""])
    _autosize(ws_so)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    LOG.info("Workpaper written: %s", output)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--inputs", required=True, type=Path)
    parser.add_argument("--period-end", required=True)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--quiet", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.WARNING if args.quiet else logging.INFO,
        format="%(levelname)s %(name)s: %(message)s",
    )

    try:
        data = _load_inputs(args.inputs)
        result = calculate(data)
    except (OSError, yaml.YAMLError, ValueError) as exc:
        LOG.error("%s", exc)
        return 1

    write_workpaper(result, args.period_end, args.output)
    print(f"Pre-Tax Income:  {result.pre_tax_income:>15,.2f}")
    print(f"Total Tax Exp:   {result.total_tax_expense:>15,.2f}")
    print(f"ETR:             {result.etr:>15.2%}")
    print(f"Workpaper:       {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
