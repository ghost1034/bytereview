"""Fixed Asset Lifecycle Manager — Book + Tax depreciation, disposals, roll-forward.

Implements the workflow in ../SKILL.md. Supports Straight-Line, 200% Declining Balance
switching to SL, and MACRS half-year convention (3/5/7/10/15/20-year class lives).
Produces a multi-sheet XLSX workpaper plus a consolidated depreciation JE.
"""
from __future__ import annotations

import argparse
import logging
import sys
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Iterable

import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

LOG = logging.getLogger("fixed_asset_manager")

DEFAULT_CAP_THRESHOLD = 5_000.0
ACCOUNTING_FMT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN = Side(border_style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# MACRS half-year percentages (IRS Pub 946 Appendix A, Table A-1)
# Each row is a recovery year; total per column sums to 100%.
MACRS_HALF_YEAR = {
    3: [33.33, 44.45, 14.81, 7.41],
    5: [20.00, 32.00, 19.20, 11.52, 11.52, 5.76],
    7: [14.29, 24.49, 17.49, 12.49, 8.93, 8.92, 8.93, 4.46],
    10: [10.00, 18.00, 14.40, 11.52, 9.22, 7.37, 6.55, 6.55, 6.56, 6.55, 3.28],
    15: [5.00, 9.50, 8.55, 7.70, 6.93, 6.23, 5.90, 5.90, 5.91, 5.90, 5.91, 5.90, 5.91, 5.90, 5.91, 2.95],
    20: [3.750, 7.219, 6.677, 6.177, 5.713, 5.285, 4.888, 4.522, 4.462, 4.461, 4.462, 4.461,
         4.462, 4.461, 4.462, 4.461, 4.462, 4.461, 4.462, 4.461, 2.231],
}


@dataclass
class AssetCalc:
    asset_id: str
    description: str
    asset_class: str
    method: str
    cost: float
    salvage: float
    useful_life_years: int
    acquisition_date: datetime
    accumulated_book: float
    period_expense_book: float
    nbv: float
    fully_depreciated: bool
    is_disposed: bool
    disposal_date: datetime | None = None
    disposal_proceeds: float = 0.0
    gain_loss: float = 0.0
    accumulated_tax: float = 0.0
    book_tax_diff: float = 0.0
    impairment_flag: str = ""
    notes: list[str] = field(default_factory=list)


# --- I/O ------------------------------------------------------------------

def _read_table(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path)
    else:
        df = pd.read_excel(path)
    df.columns = [c.strip().lower() for c in df.columns]
    return df


def _validate_far(df: pd.DataFrame) -> pd.DataFrame:
    required = ("asset_id", "description", "acquisition_date", "cost", "useful_life_years", "asset_class")
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"FAR missing required columns: {missing}")
    df = df.copy()
    df["acquisition_date"] = pd.to_datetime(df["acquisition_date"], errors="coerce")
    df["cost"] = pd.to_numeric(df["cost"], errors="coerce")
    df["useful_life_years"] = pd.to_numeric(df["useful_life_years"], errors="coerce").fillna(0).astype(int)
    bad = df[df["acquisition_date"].isna() | (df["cost"] <= 0) | (df["useful_life_years"] <= 0)]
    if not bad.empty:
        LOG.warning("Dropping %d invalid FAR rows (zero/negative cost or life)", len(bad))
        df = df[~df.index.isin(bad.index)]
    for opt, default in [
        ("salvage_value", 0.0), ("method", "SL"), ("convention", "half-year"),
        ("disposal_date", pd.NaT), ("disposal_proceeds", 0.0), ("is_disposed", False),
        ("entity", ""), ("cost_center", ""), ("tax_method", ""), ("tax_life_years", 0),
        ("bonus_eligible", False),
    ]:
        if opt not in df.columns:
            df[opt] = default
    df["salvage_value"] = pd.to_numeric(df["salvage_value"], errors="coerce").fillna(0.0)
    df["disposal_date"] = pd.to_datetime(df["disposal_date"], errors="coerce")
    df["disposal_proceeds"] = pd.to_numeric(df["disposal_proceeds"], errors="coerce").fillna(0.0)
    df["is_disposed"] = df["is_disposed"].fillna(False).astype(bool)
    df["method"] = df["method"].fillna("SL").astype(str).str.upper().str.strip()
    df["asset_class"] = df["asset_class"].astype(str).str.strip()
    df["tax_life_years"] = pd.to_numeric(df["tax_life_years"], errors="coerce").fillna(0).astype(int)
    df["tax_method"] = df["tax_method"].fillna("").astype(str).str.strip()
    df["convention"] = df["convention"].fillna("half-year").astype(str).str.strip()
    return df.reset_index(drop=True)


# --- Depreciation engines -------------------------------------------------

def _months_in_service(acq_date: datetime, report_date: datetime, convention: str) -> float:
    """Return decimal months in service through report_date applying the convention."""
    if convention.lower() == "half-year":
        # Half-year convention: 6 months in year 1 regardless of placed-in-service date
        # Then 12 months per subsequent year, ending with 6 months in the final recovery year.
        years_elapsed = (report_date - acq_date).days / 365.25
        if years_elapsed <= 0.5:
            return min(years_elapsed * 12, 6)
        return 6 + (years_elapsed - 0.5) * 12
    if convention.lower() == "mid-month":
        # Real estate: assume placed in service mid of acquisition month
        acq_eff = acq_date.replace(day=15)
        delta = relativedelta(report_date, acq_eff)
        return delta.years * 12 + delta.months + delta.days / 30.0
    # Full-month default
    delta = relativedelta(report_date, acq_date)
    return delta.years * 12 + delta.months + delta.days / 30.0


def _sl_depreciation(cost: float, salvage: float, useful_life_years: int,
                     months_elapsed: float, report_date: datetime,
                     prior_month_elapsed: float | None = None) -> tuple[float, float]:
    """Return (accumulated_through_report, current_month_expense)."""
    basis = cost - salvage
    total_months = useful_life_years * 12
    per_month = basis / total_months if total_months > 0 else 0.0
    capped = min(months_elapsed, total_months)
    accumulated = per_month * capped
    # Current month expense = depreciation for the month ending at report_date
    if capped > 0:
        prior_capped = min(max(months_elapsed - 1, 0), total_months)
        period_expense = (capped - prior_capped) * per_month
    else:
        period_expense = 0.0
    return accumulated, period_expense


def _ddb_depreciation(cost: float, salvage: float, useful_life_years: int,
                      months_elapsed: float) -> tuple[float, float]:
    """200%-DB switching to SL when SL >= DB. Returns (accumulated, current_month)."""
    basis_start = cost
    salvage = max(salvage, 0.0)
    rate = 2.0 / useful_life_years if useful_life_years > 0 else 0.0
    total_months = useful_life_years * 12
    capped_months = min(months_elapsed, total_months)
    # Walk month by month
    book_value = cost
    accum_prev = 0.0
    accum_now = 0.0
    last_month_exp = 0.0
    for m in range(1, int(round(capped_months)) + 1):
        years_in = m / 12.0
        years_remaining = max(useful_life_years - (m - 1) / 12.0, 1e-9)
        # SL on remaining basis (book_value - salvage) / remaining months
        remaining_months = max(total_months - (m - 1), 1)
        sl_month = max(book_value - salvage, 0.0) / remaining_months
        ddb_month = book_value * rate / 12.0
        # Switch when SL >= DDB
        period = max(sl_month, ddb_month)
        # Don't depreciate below salvage
        period = max(min(period, book_value - salvage), 0.0)
        book_value -= period
        accum_prev = accum_now
        accum_now += period
        last_month_exp = period
        if book_value <= salvage + 1e-6:
            break
    # If months_elapsed has a fractional part, prorate the last month
    frac = capped_months - int(round(capped_months))
    return accum_now, last_month_exp


def _macrs_depreciation(cost: float, life_years: int, report_date: datetime,
                        acq_date: datetime) -> tuple[float, float]:
    """Half-year convention MACRS. Returns (accumulated_through_year, current_year_remaining)."""
    table = MACRS_HALF_YEAR.get(life_years)
    if table is None:
        return 0.0, 0.0
    # Year index: 1 = acquisition year
    year_index = report_date.year - acq_date.year + 1
    if year_index <= 0:
        return 0.0, 0.0
    accumulated = 0.0
    for i in range(min(year_index, len(table))):
        accumulated += cost * table[i] / 100.0
    current_year_total = cost * table[min(year_index, len(table)) - 1] / 100.0 if year_index <= len(table) else 0.0
    return accumulated, current_year_total / 12.0


# --- Per-asset orchestration ---------------------------------------------

def compute_asset(row: pd.Series, report_date: datetime) -> AssetCalc:
    asset_class = str(row["asset_class"])
    is_land = "land" in asset_class.lower()
    method = (row.get("method") or "SL").upper()
    if is_land:
        method = "NONE"
    salvage = float(row.get("salvage_value", 0))
    convention = (row.get("convention") or "half-year").lower()
    months_elapsed = _months_in_service(row["acquisition_date"], report_date, convention)
    months_elapsed = max(months_elapsed, 0)

    # Disposal handling
    is_disposed = bool(row.get("is_disposed", False)) or (
        pd.notna(row.get("disposal_date")) and row["disposal_date"] <= report_date
    )
    disposal_date = row["disposal_date"] if pd.notna(row.get("disposal_date")) else None
    disposal_proceeds = float(row.get("disposal_proceeds", 0))

    accum = 0.0
    period_expense = 0.0
    if method == "NONE":
        accum = 0.0
        period_expense = 0.0
    elif method == "SL":
        # Use mid-month convention for real estate; otherwise honor convention.
        accum, period_expense = _sl_depreciation(
            float(row["cost"]), salvage, int(row["useful_life_years"]), months_elapsed, report_date
        )
    elif method == "DDB":
        accum, period_expense = _ddb_depreciation(
            float(row["cost"]), salvage, int(row["useful_life_years"]), months_elapsed
        )
    elif method.startswith("MACRS-"):
        life = int(method.split("-")[1])
        accum, period_expense = _macrs_depreciation(float(row["cost"]), life, report_date, row["acquisition_date"])
    else:
        LOG.warning("Unknown method %s for asset %s; falling back to SL", method, row["asset_id"])
        accum, period_expense = _sl_depreciation(
            float(row["cost"]), salvage, int(row["useful_life_years"]), months_elapsed, report_date
        )

    # If disposed before report date, cut depreciation off
    notes: list[str] = []
    if is_disposed and disposal_date is not None and disposal_date <= report_date:
        disp_months = _months_in_service(row["acquisition_date"], disposal_date, convention)
        if method == "SL":
            accum, _ = _sl_depreciation(
                float(row["cost"]), salvage, int(row["useful_life_years"]), disp_months, disposal_date
            )
            period_expense = 0.0  # disposed; no May expense on a disposed asset
        notes.append(f"Disposed {disposal_date.date()}")

    nbv = float(row["cost"]) - accum
    # Don't go below salvage
    if nbv < salvage:
        accum = float(row["cost"]) - salvage
        nbv = salvage

    fully_dep = abs(nbv - salvage) < 0.01 and not is_land
    gain_loss = 0.0
    if is_disposed:
        gain_loss = disposal_proceeds - nbv
        nbv = 0.0  # off the books after disposal

    # Tax accumulated (book-tax difference)
    tax_method = (row.get("tax_method") or "").upper()
    tax_life = int(row.get("tax_life_years") or 0)
    tax_accum = 0.0
    if tax_method.startswith("MACRS-") and tax_life > 0:
        tax_accum, _ = _macrs_depreciation(float(row["cost"]), tax_life, report_date, row["acquisition_date"])
    elif tax_method == "SL" and tax_life > 0:
        tax_accum, _ = _sl_depreciation(float(row["cost"]), salvage, tax_life, months_elapsed, report_date)
    else:
        tax_accum = accum  # if no separate tax method specified, book == tax (no DTA/DTL)
    book_tax_diff = accum - tax_accum

    if float(row["cost"]) > 100000 and abs(disposal_proceeds) > 0 and gain_loss < -1000:
        notes.append("Large loss on disposal — consider impairment review")

    return AssetCalc(
        asset_id=str(row["asset_id"]),
        description=str(row["description"]),
        asset_class=asset_class,
        method=method,
        cost=float(row["cost"]),
        salvage=salvage,
        useful_life_years=int(row["useful_life_years"]),
        acquisition_date=row["acquisition_date"],
        accumulated_book=round(accum, 2),
        period_expense_book=round(period_expense, 2),
        nbv=round(nbv, 2),
        fully_depreciated=fully_dep,
        is_disposed=is_disposed,
        disposal_date=disposal_date,
        disposal_proceeds=disposal_proceeds,
        gain_loss=round(gain_loss, 2),
        accumulated_tax=round(tax_accum, 2),
        book_tax_diff=round(book_tax_diff, 2),
        notes=notes,
    )


# --- XLSX writer ----------------------------------------------------------

def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX


def _autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(50, max_len + 2)


def write_workpaper(
    assets: list[AssetCalc],
    far_df: pd.DataFrame,
    report_date: datetime,
    cap_policy: dict[str, float],
    dep_account: str,
    accum_account: str,
    disposal_account: str,
    prior_far: pd.DataFrame | None,
    output: Path,
    far_path: Path,
) -> None:
    wb = Workbook()
    # Summary
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Fixed Asset Workpaper"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Report Date", report_date.strftime("%Y-%m-%d")])
    ws.append(["Asset Count", len(assets)])
    total_cost = sum(a.cost for a in assets)
    total_accum = sum(a.accumulated_book for a in assets)
    total_nbv = sum(a.nbv for a in assets)
    total_exp = sum(a.period_expense_book for a in assets)
    total_disp = sum(a.gain_loss for a in assets if a.is_disposed)
    ws.append(["Total Cost (Book)", total_cost])
    ws.append(["Total Accumulated (Book)", total_accum])
    ws.append(["Total NBV (Book)", total_nbv])
    ws.append(["Period Depreciation Expense", total_exp])
    ws.append(["Net Gain/(Loss) on Disposal", total_disp])
    for cell in ("B5", "B6", "B7", "B8", "B9"):
        ws[cell].number_format = ACCOUNTING_FMT
    _autosize(ws)

    # Depreciation Detail
    cols = ["Asset ID", "Description", "Class", "Method", "Acquired", "Cost", "Salvage",
            "Life (yrs)", "Months in Service", "Accumulated (Book)", "Period Expense",
            "NBV (Book)", "Fully Dep?", "Disposed?", "Accumulated (Tax)", "Book-Tax Diff", "Notes"]
    ws_d = wb.create_sheet("DepreciationDetail")
    ws_d.append(cols)
    _style_header(ws_d)
    for a in assets:
        months = _months_in_service(a.acquisition_date, report_date, "half-year")
        ws_d.append([
            a.asset_id, a.description, a.asset_class, a.method,
            a.acquisition_date.strftime("%Y-%m-%d"), a.cost, a.salvage, a.useful_life_years,
            round(months, 1), a.accumulated_book, a.period_expense_book, a.nbv,
            "Yes" if a.fully_depreciated else "", "Yes" if a.is_disposed else "",
            a.accumulated_tax, a.book_tax_diff, "; ".join(a.notes),
        ])
    for letter in ("F", "G", "J", "K", "L", "O", "P"):
        for cell in ws_d[letter][1:]:
            cell.number_format = ACCOUNTING_FMT
    _autosize(ws_d)
    ws_d.freeze_panes = "A2"

    # Depreciation JE
    ws_je = wb.create_sheet("DepreciationJE")
    ws_je.append(["Date", "Account", "Description", "Debit", "Credit", "Memo"])
    _style_header(ws_je)
    ws_je.append([report_date.strftime("%Y-%m-%d"), dep_account, "Depreciation Expense",
                  round(total_exp, 2), 0, f"Monthly depreciation for {len(assets)} assets"])
    ws_je.append([report_date.strftime("%Y-%m-%d"), accum_account, "Accumulated Depreciation",
                  0, round(total_exp, 2), f"Monthly depreciation for {len(assets)} assets"])
    for cell in ws_je["D"][1:] + ws_je["E"][1:]:
        cell.number_format = ACCOUNTING_FMT
    _autosize(ws_je)

    # Disposals
    ws_dp = wb.create_sheet("Disposals")
    ws_dp.append(["Asset ID", "Description", "Cost", "Accumulated", "NBV", "Proceeds", "Gain/(Loss)", "Disposal Date"])
    _style_header(ws_dp)
    for a in assets:
        if a.is_disposed:
            ws_dp.append([a.asset_id, a.description, a.cost, a.accumulated_book,
                          a.cost - a.accumulated_book, a.disposal_proceeds, a.gain_loss,
                          a.disposal_date.strftime("%Y-%m-%d") if a.disposal_date else ""])
    for letter in ("C", "D", "E", "F", "G"):
        for cell in ws_dp[letter][1:]:
            cell.number_format = ACCOUNTING_FMT
    _autosize(ws_dp)

    # Additions Log
    ws_a = wb.create_sheet("AdditionsLog")
    ws_a.append(["Asset ID", "Description", "Class", "Acquired", "Cost", "Method", "Life"])
    _style_header(ws_a)
    period_start = report_date.replace(day=1)
    for a in assets:
        if a.acquisition_date >= period_start - timedelta(days=1) and a.acquisition_date <= report_date:
            ws_a.append([a.asset_id, a.description, a.asset_class,
                         a.acquisition_date.strftime("%Y-%m-%d"), a.cost, a.method, a.useful_life_years])
    for cell in ws_a["E"][1:]:
        cell.number_format = ACCOUNTING_FMT
    _autosize(ws_a)

    # Roll-forward
    ws_rf = wb.create_sheet("RollForward")
    ws_rf.append(["Roll-Forward", "Cost", "Accumulated Depreciation", "NBV"])
    _style_header(ws_rf)
    if prior_far is not None:
        beg_cost = float(prior_far["cost"].sum())
        beg_accum = float(prior_far.get("accumulated_book", pd.Series([0])).sum())
    else:
        beg_cost = total_cost - sum(a.cost for a in assets if a.acquisition_date >= period_start)
        beg_accum = total_accum - total_exp
    additions = sum(a.cost for a in assets if a.acquisition_date >= period_start - timedelta(days=1) and a.acquisition_date <= report_date)
    disp_cost = sum(a.cost for a in assets if a.is_disposed)
    disp_accum = sum(a.accumulated_book for a in assets if a.is_disposed)
    end_cost = beg_cost + additions - disp_cost
    end_accum = beg_accum + total_exp - disp_accum
    ws_rf.append(["Beginning", beg_cost, beg_accum, beg_cost - beg_accum])
    ws_rf.append(["Additions", additions, 0, additions])
    ws_rf.append(["Disposals", -disp_cost, -disp_accum, -(disp_cost - disp_accum)])
    ws_rf.append(["Period Expense", 0, total_exp, -total_exp])
    ws_rf.append(["Ending", end_cost, end_accum, end_cost - end_accum])
    for letter in ("B", "C", "D"):
        for cell in ws_rf[letter][1:]:
            cell.number_format = ACCOUNTING_FMT
    _autosize(ws_rf)

    # Impairment Flags
    ws_i = wb.create_sheet("ImpairmentFlags")
    ws_i.append(["Asset ID", "Description", "Flag", "Notes"])
    _style_header(ws_i)
    any_flag = False
    for a in assets:
        if a.notes and any("impairment" in n.lower() for n in a.notes):
            ws_i.append([a.asset_id, a.description, "Recoverability test recommended", "; ".join(a.notes)])
            any_flag = True
    if not any_flag:
        ws_i.append(["(no triggering events detected)"])
    _autosize(ws_i)

    # Capitalization Check
    ws_c = wb.create_sheet("CapitalizationCheck")
    ws_c.append(["Asset ID", "Description", "Class", "Cost", "Threshold", "Compliant?"])
    _style_header(ws_c)
    for a in assets:
        thr = cap_policy.get(a.asset_class, cap_policy.get("__default__", DEFAULT_CAP_THRESHOLD))
        ok = a.cost >= thr
        ws_c.append([a.asset_id, a.description, a.asset_class, a.cost, thr,
                     "Yes" if ok else "No (consider expensing)"])
    for letter in ("D", "E"):
        for cell in ws_c[letter][1:]:
            cell.number_format = ACCOUNTING_FMT
    # Color non-compliant
    for row in ws_c.iter_rows(min_row=2):
        if row[5].value and row[5].value.startswith("No"):
            for c in row:
                c.fill = RED_FILL
    _autosize(ws_c)

    # Book-Tax
    ws_bt = wb.create_sheet("BookTax")
    ws_bt.append(["Asset ID", "Description", "Cost", "Book Accum", "Tax Accum", "Book-Tax Diff", "DTA/(DTL) @ 21%"])
    _style_header(ws_bt)
    total_diff = 0.0
    for a in assets:
        dt = a.book_tax_diff * 0.21
        total_diff += a.book_tax_diff
        ws_bt.append([a.asset_id, a.description, a.cost, a.accumulated_book, a.accumulated_tax,
                      a.book_tax_diff, round(dt, 2)])
    ws_bt.append(["TOTAL", "", "", "", "", total_diff, round(total_diff * 0.21, 2)])
    for letter in ("C", "D", "E", "F", "G"):
        for cell in ws_bt[letter][1:]:
            cell.number_format = ACCOUNTING_FMT
    _autosize(ws_bt)

    # Audit Trail
    ws_at = wb.create_sheet("AuditTrail")
    ws_at.append(["Key", "Value"])
    _style_header(ws_at)
    ws_at.append(["FAR Input", str(far_path)])
    ws_at.append(["Report Date", report_date.strftime("%Y-%m-%d")])
    ws_at.append(["Asset Count", len(assets)])
    ws_at.append(["Depreciation Expense Account", dep_account])
    ws_at.append(["Accumulated Depreciation Account", accum_account])
    ws_at.append(["Disposal Account", disposal_account])
    ws_at.append(["Generated", datetime.now().isoformat(timespec="seconds")])
    _autosize(ws_at)

    # Sign-Off
    ws_s = wb.create_sheet("SignOff")
    ws_s.append(["Role", "Name", "Title", "Date", "Notes"])
    _style_header(ws_s)
    ws_s.append(["Preparer (Fixed Asset Accountant)", "", "", "", ""])
    ws_s.append(["Reviewer (Senior Accountant)", "", "", "", ""])
    ws_s.append(["Approver (Controller)", "", "", "", ""])
    _autosize(ws_s)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    LOG.info("Workpaper written: %s", output)


# --- CLI ------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--far", required=True, type=Path)
    parser.add_argument("--cap-policy", type=Path, default=None)
    parser.add_argument("--report-date", required=True, help="YYYY-MM-DD period-end")
    parser.add_argument("--prior-far", type=Path, default=None)
    parser.add_argument("--depreciation-account", default="7100")
    parser.add_argument("--accumulated-account", default="1599")
    parser.add_argument("--disposal-account", default="8200")
    parser.add_argument("--method-default", default="SL")
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--quiet", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.WARNING if args.quiet else logging.INFO,
        format="%(levelname)s %(name)s: %(message)s",
    )

    try:
        far = _validate_far(_read_table(args.far))
    except ValueError as e:
        LOG.error("FAR validation failed: %s", e)
        return 1
    far["method"] = far["method"].fillna(args.method_default)

    cap_policy: dict[str, float] = {"__default__": DEFAULT_CAP_THRESHOLD}
    if args.cap_policy:
        cp = _read_table(args.cap_policy)
        for _, r in cp.iterrows():
            cap_policy[str(r["asset_class"])] = float(r["threshold_usd"])

    report_date = datetime.strptime(args.report_date, "%Y-%m-%d")
    assets = [compute_asset(r, report_date) for _, r in far.iterrows()]
    prior_far = _read_table(args.prior_far) if args.prior_far else None

    write_workpaper(
        assets=assets, far_df=far, report_date=report_date, cap_policy=cap_policy,
        dep_account=args.depreciation_account, accum_account=args.accumulated_account,
        disposal_account=args.disposal_account, prior_far=prior_far,
        output=args.output, far_path=args.far,
    )
    total_exp = sum(a.period_expense_book for a in assets)
    total_nbv = sum(a.nbv for a in assets)
    print(f"Assets:            {len(assets)}")
    print(f"Period Expense:    {total_exp:>15,.2f}")
    print(f"Total NBV:         {total_nbv:>15,.2f}")
    print(f"Workpaper:         {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
