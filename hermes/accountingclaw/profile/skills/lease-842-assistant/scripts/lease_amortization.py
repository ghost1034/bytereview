"""ASC 842 Lease Assistant — classification, PV, schedules, modifications, JE.

Implements the workflow in ../SKILL.md.
"""
from __future__ import annotations

import argparse
import logging
import sys
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

LOG = logging.getLogger("lease_842")

ACCOUNTING_FMT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN = Side(border_style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

OWNERSHIP_THRESHOLD = 0.75
FV_THRESHOLD = 0.90


REQUIRED = (
    "lease_id", "description", "commencement_date", "term_months",
    "monthly_payment", "ibr_annual", "payment_timing",
)


@dataclass
class LeaseClassification:
    c1_transfer_ownership: bool
    c2_purchase_option_certain: bool
    c3_term_major_part: bool
    c4_pv_substantially_all: bool
    c5_specialized: bool

    @property
    def is_finance(self) -> bool:
        return any([self.c1_transfer_ownership, self.c2_purchase_option_certain,
                    self.c3_term_major_part, self.c4_pv_substantially_all, self.c5_specialized])

    @property
    def label(self) -> str:
        return "Finance" if self.is_finance else "Operating"


@dataclass
class Lease:
    lease_id: str
    description: str
    commencement_date: datetime
    term_months: int
    monthly_payment: float
    ibr_annual: float
    payment_timing: str  # advance / arrears
    prepaid_rent: float
    idc: float
    incentives: float
    economic_life_months: int
    fair_value: float
    purchase_option_exercise_price: float
    reasonably_certain_purchase: bool
    specialized_nature: bool
    transfer_ownership: bool
    classification: LeaseClassification | None = None
    initial_liability: float = 0.0
    initial_rou: float = 0.0
    schedule: pd.DataFrame = field(default_factory=pd.DataFrame)
    modification_effective_date: datetime | None = None
    modification_new_term_months: int | None = None
    modification_new_payment: float | None = None
    modification_new_ibr: float | None = None
    period_interest: float = 0.0
    period_amort: float = 0.0
    period_lease_cost: float = 0.0
    period_principal: float = 0.0
    liability_at_report: float = 0.0
    rou_at_report: float = 0.0


def _read_table(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path)
    else:
        df = pd.read_excel(path)
    df.columns = [c.strip().lower() for c in df.columns]
    return df


def _validate(df: pd.DataFrame) -> pd.DataFrame:
    missing = [c for c in REQUIRED if c not in df.columns]
    if missing:
        raise ValueError(f"Lease inventory missing required columns: {missing}")
    df = df.copy()
    df["commencement_date"] = pd.to_datetime(df["commencement_date"], errors="coerce")
    if "modification_effective_date" in df.columns:
        df["modification_effective_date"] = pd.to_datetime(df["modification_effective_date"], errors="coerce")
    for c in ("term_months", "economic_life_months"):
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0).astype(int)
    for c in ("monthly_payment", "ibr_annual", "prepaid_rent", "idc", "incentives",
              "fair_value", "purchase_option_exercise_price"):
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    for c in ("reasonably_certain_purchase", "specialized_nature", "transfer_ownership"):
        if c in df.columns:
            df[c] = df[c].astype(bool)
        else:
            df[c] = False
    for c in ("prepaid_rent", "idc", "incentives"):
        if c not in df.columns:
            df[c] = 0.0
    for c in ("economic_life_months",):
        if c not in df.columns:
            df[c] = 0
    for c in ("fair_value", "purchase_option_exercise_price"):
        if c not in df.columns:
            df[c] = 0.0
    df["payment_timing"] = df["payment_timing"].astype(str).str.lower().str.strip()
    return df.reset_index(drop=True)


# --- PV calculation ------------------------------------------------------

def pv_annuity(payment: float, n: int, monthly_rate: float, advance: bool) -> float:
    if monthly_rate == 0:
        return payment * n
    pv_arrears = payment * (1 - (1 + monthly_rate) ** -n) / monthly_rate
    return pv_arrears * (1 + monthly_rate) if advance else pv_arrears


def classify(lease: Lease, ownership_thresh: float, fv_thresh: float, pv: float) -> LeaseClassification:
    c1 = lease.transfer_ownership
    c2 = lease.reasonably_certain_purchase
    c3 = lease.economic_life_months > 0 and lease.term_months / lease.economic_life_months >= ownership_thresh
    c4 = lease.fair_value > 0 and pv / lease.fair_value >= fv_thresh
    c5 = lease.specialized_nature
    return LeaseClassification(c1, c2, c3, c4, c5)


def build_schedule(lease: Lease) -> pd.DataFrame:
    """Build the month-by-month amortization schedule."""
    r = lease.ibr_annual / 12.0
    n = lease.term_months
    advance = lease.payment_timing == "advance"
    pv = pv_annuity(lease.monthly_payment, n, r, advance)
    lease.initial_liability = round(pv, 2)
    lease.initial_rou = round(pv + lease.prepaid_rent + lease.idc - lease.incentives, 2)

    # Straight-line for operating
    total_cost = lease.monthly_payment * n + lease.idc + lease.prepaid_rent - lease.incentives
    sl_monthly = total_cost / n if n else 0.0

    rows: list[dict] = []
    liability = pv
    rou = lease.initial_rou
    for i in range(n):
        period_date = lease.commencement_date + relativedelta(months=i)
        # interest accretion
        if advance:
            # advance: payment at beginning, then interest on remaining
            principal = lease.monthly_payment - 0 if i == 0 else 0
            # simpler: for advance, treat payment then interest accretes for the month on (liab - payment)
            interest = (liability - lease.monthly_payment) * r if i == 0 else (liability) * r
            # we'll use the canonical schedule: advance reduces liability immediately, accretes after
            if i == 0:
                liability_after_payment = liability - lease.monthly_payment
                interest = liability_after_payment * r
                liability_end = liability_after_payment + interest
                principal_actual = lease.monthly_payment
            else:
                # pay at beginning
                liability_after_payment = liability - lease.monthly_payment
                interest = liability_after_payment * r
                liability_end = liability_after_payment + interest
                principal_actual = lease.monthly_payment
            principal = principal_actual
        else:
            # arrears
            interest = liability * r
            principal = lease.monthly_payment - interest
            liability_end = liability - principal

        # operating lease cost = single line, finance lease cost = interest + ROU amort
        if lease.classification and lease.classification.is_finance:
            # finance: ROU amortized SL over shorter of lease term and economic life (or to purchase option)
            amort_period = lease.economic_life_months if (
                lease.reasonably_certain_purchase or lease.transfer_ownership
            ) and lease.economic_life_months > 0 else n
            rou_amort = lease.initial_rou / max(amort_period, 1)
            lease_cost = interest + rou_amort
        else:
            # operating: single-cost; ROU amort = SL cost - interest
            rou_amort = sl_monthly - interest
            lease_cost = sl_monthly

        rou_end = max(rou - rou_amort, 0.0)

        rows.append({
            "Month #": i + 1,
            "Period": period_date.strftime("%Y-%m"),
            "Beg Liability": round(liability, 2),
            "Payment": round(lease.monthly_payment, 2),
            "Interest": round(interest, 2),
            "Principal": round(principal, 2),
            "End Liability": round(max(liability_end, 0), 2),
            "Beg ROU": round(rou, 2),
            "ROU Amort": round(rou_amort, 2),
            "End ROU": round(rou_end, 2),
            "Period Lease Cost": round(lease_cost, 2),
        })
        liability = max(liability_end, 0)
        rou = rou_end

    return pd.DataFrame(rows)


def report_date_state(lease: Lease, schedule: pd.DataFrame, report_date: datetime) -> None:
    if schedule.empty:
        return
    # Determine which month is the report_date
    target = report_date.strftime("%Y-%m")
    matched = schedule[schedule["Period"] == target]
    if matched.empty:
        # report date is after lease end
        lease.liability_at_report = 0.0
        lease.rou_at_report = 0.0
        return
    row = matched.iloc[0]
    lease.period_interest = float(row["Interest"])
    lease.period_amort = float(row["ROU Amort"])
    lease.period_principal = float(row["Principal"])
    lease.period_lease_cost = float(row["Period Lease Cost"])
    lease.liability_at_report = float(row["End Liability"])
    lease.rou_at_report = float(row["End ROU"])


# --- Writer --------------------------------------------------------------

def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX


def _autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(40, max_len + 2)


def write_workpaper(leases: list[Lease], report_date: datetime, output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["ASC 842 Lease Workpaper"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Report Date", report_date.strftime("%Y-%m-%d")])
    ws.append(["Total Leases", len(leases)])
    finance_count = sum(1 for l in leases if l.classification and l.classification.is_finance)
    operating_count = len(leases) - finance_count
    ws.append(["Operating Leases", operating_count])
    ws.append(["Finance Leases", finance_count])
    ws.append(["Total ROU (initial)", sum(l.initial_rou for l in leases)])
    ws.append(["Total Liability (initial)", sum(l.initial_liability for l in leases)])
    ws.append(["Total Liability (at report)", sum(l.liability_at_report for l in leases)])
    ws.append(["Total ROU (at report)", sum(l.rou_at_report for l in leases)])
    ws.append(["Period Interest (Finance)", sum(l.period_interest for l in leases if l.classification and l.classification.is_finance)])
    ws.append(["Period ROU Amort (Finance)", sum(l.period_amort for l in leases if l.classification and l.classification.is_finance)])
    ws.append(["Period Operating Lease Cost", sum(l.period_lease_cost for l in leases if l.classification and not l.classification.is_finance)])
    for cell in ("B7", "B8", "B9", "B10", "B11", "B12", "B13"):
        ws[cell].number_format = ACCOUNTING_FMT
    _autosize(ws)

    # Classification
    ws_c = wb.create_sheet("Classification")
    ws_c.append([
        "Lease ID", "Description", "C1 Transfer", "C2 Purchase",
        "C3 Term Major Part", "C4 PV Substantially All", "C5 Specialized",
        "Classification",
    ])
    _style_header(ws_c)
    for l in leases:
        c = l.classification
        ws_c.append([
            l.lease_id, l.description,
            "Y" if c and c.c1_transfer_ownership else "N",
            "Y" if c and c.c2_purchase_option_certain else "N",
            "Y" if c and c.c3_term_major_part else "N",
            "Y" if c and c.c4_pv_substantially_all else "N",
            "Y" if c and c.c5_specialized else "N",
            c.label if c else "Unknown",
        ])
    _autosize(ws_c)

    # Per-lease schedules
    for l in leases:
        if l.schedule.empty:
            continue
        ws_s = wb.create_sheet(f"Schedule_{l.lease_id}"[:31])
        ws_s.append(list(l.schedule.columns))
        _style_header(ws_s)
        for _, r in l.schedule.iterrows():
            ws_s.append([r[c] for c in l.schedule.columns])
        money_cols = ["Beg Liability", "Payment", "Interest", "Principal",
                       "End Liability", "Beg ROU", "ROU Amort", "End ROU", "Period Lease Cost"]
        for col_idx, col in enumerate(l.schedule.columns, start=1):
            if col in money_cols:
                for cell in ws_s[get_column_letter(col_idx)][1:]:
                    cell.number_format = ACCOUNTING_FMT
        _autosize(ws_s)

    # IBR Sensitivity
    ws_ibr = wb.create_sheet("IBR_Sensitivity")
    ws_ibr.append(["Lease ID", "Base Liability"] +
                  [f"IBR {bp:+d} bps" for bp in (-200, -100, -50, 50, 100, 200)])
    _style_header(ws_ibr)
    for l in leases:
        base = l.initial_liability
        row = [l.lease_id, base]
        for bp in (-200, -100, -50, 50, 100, 200):
            ibr = l.ibr_annual + bp / 10000.0
            pv = pv_annuity(l.monthly_payment, l.term_months, ibr / 12,
                            l.payment_timing == "advance")
            row.append(round(pv, 2))
        ws_ibr.append(row)
    for col_idx in range(2, 9):
        for cell in ws_ibr[get_column_letter(col_idx)][1:]:
            cell.number_format = ACCOUNTING_FMT
    _autosize(ws_ibr)

    # JE
    ws_je = wb.create_sheet("JE")
    ws_je.append(["Period", "Lease ID", "Account", "Description", "Debit", "Credit"])
    _style_header(ws_je)
    period_label = report_date.strftime("%Y-%m")
    for l in leases:
        if not l.classification:
            continue
        if l.classification.is_finance:
            ws_je.append([period_label, l.lease_id, "7140 Interest Expense (Lease)", l.description, round(l.period_interest, 2), 0])
            ws_je.append([period_label, l.lease_id, "1610 ROU Asset", l.description, 0, round(l.period_amort, 2)])
            ws_je.append([period_label, l.lease_id, "7150 ROU Amortization Expense", l.description, round(l.period_amort, 2), 0])
            ws_je.append([period_label, l.lease_id, "2510 Lease Liability", l.description, round(l.period_principal, 2), 0])
            ws_je.append([period_label, l.lease_id, "1000 Cash", l.description, 0, round(l.monthly_payment, 2)])
        else:
            ws_je.append([period_label, l.lease_id, "6420 Operating Lease Cost", l.description, round(l.period_lease_cost, 2), 0])
            ws_je.append([period_label, l.lease_id, "2510 Lease Liability", l.description, round(l.period_principal, 2), 0])
            ws_je.append([period_label, l.lease_id, "1610 ROU Asset", l.description, 0, round(l.period_amort, 2)])
            ws_je.append([period_label, l.lease_id, "1000 Cash", l.description, 0, round(l.monthly_payment, 2)])
    for letter in ("E", "F"):
        for cell in ws_je[letter][1:]:
            cell.number_format = ACCOUNTING_FMT
    _autosize(ws_je)

    # Modifications (placeholder)
    ws_m = wb.create_sheet("Modifications")
    ws_m.append(["Lease ID", "Mod Date", "Original Term", "New Term", "Original Payment", "New Payment", "Original IBR", "New IBR"])
    _style_header(ws_m)
    any_mod = False
    for l in leases:
        if l.modification_effective_date is not None and pd.notna(l.modification_effective_date):
            any_mod = True
            ws_m.append([l.lease_id, l.modification_effective_date.strftime("%Y-%m-%d"),
                         l.term_months, l.modification_new_term_months,
                         l.monthly_payment, l.modification_new_payment,
                         f"{l.ibr_annual*100:.2f}%", f"{(l.modification_new_ibr or 0)*100:.2f}%"])
    if not any_mod:
        ws_m.append(["(no modifications)"])
    _autosize(ws_m)

    # AuditTrail
    ws_a = wb.create_sheet("AuditTrail")
    ws_a.append(["Key", "Value"])
    _style_header(ws_a)
    ws_a.append(["Report Date", report_date.strftime("%Y-%m-%d")])
    ws_a.append(["Leases Processed", len(leases)])
    ws_a.append(["Generated", datetime.now().isoformat(timespec="seconds")])
    _autosize(ws_a)

    # SignOff
    ws_s = wb.create_sheet("SignOff")
    ws_s.append(["Role", "Name", "Title", "Date", "Notes"])
    _style_header(ws_s)
    ws_s.append(["Preparer (Lease Accountant)", "", "", "", ""])
    ws_s.append(["Reviewer (Tech Accounting)", "", "", "", ""])
    ws_s.append(["Approver (Controller)", "", "", "", ""])
    _autosize(ws_s)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    LOG.info("Workpaper written: %s", output)


# --- CLI ------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--leases", required=True, type=Path)
    parser.add_argument("--report-date", required=True)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--ownership-criteria-threshold", type=float, default=OWNERSHIP_THRESHOLD)
    parser.add_argument("--fair-value-criteria-threshold", type=float, default=FV_THRESHOLD)
    parser.add_argument("--quiet", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.WARNING if args.quiet else logging.INFO,
        format="%(levelname)s %(name)s: %(message)s",
    )

    try:
        df = _validate(_read_table(args.leases))
    except ValueError as e:
        LOG.error("Validation failed: %s", e)
        return 1

    report_date = datetime.strptime(args.report_date, "%Y-%m-%d")
    leases: list[Lease] = []
    for _, r in df.iterrows():
        l = Lease(
            lease_id=str(r["lease_id"]), description=str(r["description"]),
            commencement_date=r["commencement_date"], term_months=int(r["term_months"]),
            monthly_payment=float(r["monthly_payment"]), ibr_annual=float(r["ibr_annual"]),
            payment_timing=str(r["payment_timing"]),
            prepaid_rent=float(r.get("prepaid_rent", 0)),
            idc=float(r.get("idc", 0)),
            incentives=float(r.get("incentives", 0)),
            economic_life_months=int(r.get("economic_life_months", 0)),
            fair_value=float(r.get("fair_value", 0)),
            purchase_option_exercise_price=float(r.get("purchase_option_exercise_price", 0)),
            reasonably_certain_purchase=bool(r.get("reasonably_certain_purchase", False)),
            specialized_nature=bool(r.get("specialized_nature", False)),
            transfer_ownership=bool(r.get("transfer_ownership", False)),
        )
        # Classify
        pv = pv_annuity(l.monthly_payment, l.term_months, l.ibr_annual / 12,
                        l.payment_timing == "advance")
        l.classification = classify(l, args.ownership_criteria_threshold,
                                     args.fair_value_criteria_threshold, pv)
        l.schedule = build_schedule(l)
        report_date_state(l, l.schedule, report_date)
        leases.append(l)

    write_workpaper(leases, report_date, args.output)
    print(f"Leases:          {len(leases)}")
    print(f"Total ROU:       {sum(l.rou_at_report for l in leases):>15,.2f}")
    print(f"Total Liability: {sum(l.liability_at_report for l in leases):>15,.2f}")
    print(f"Period Cost:     {sum(l.period_lease_cost for l in leases):>15,.2f}")
    print(f"Workpaper:       {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
