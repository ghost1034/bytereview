"""Flux Variance Analyst — material variance detection with top-driver extraction.

Implements the workflow in ../SKILL.md.
"""
from __future__ import annotations

import argparse
import logging
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

LOG = logging.getLogger("flux_analyst")

ACCOUNTING_FMT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
PCT_FMT = "0.00%"
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN = Side(border_style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

REQUIRED = ("account_code", "account_name", "balance")
MATERIALITY_PCT_DEFAULT = 0.10
MATERIALITY_ABS_DEFAULT = 5000.0


def _read_tb(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path)
    else:
        df = pd.read_excel(path)
    df.columns = [c.strip().lower() for c in df.columns]
    missing = [c for c in REQUIRED if c not in df.columns]
    if missing:
        raise ValueError(f"Trial balance missing columns: {missing}")
    df = df.copy()
    df["account_code"] = df["account_code"].astype(str)
    df["balance"] = pd.to_numeric(df["balance"], errors="coerce").fillna(0)
    return df


def analyze(cp: pd.DataFrame, pp: pd.DataFrame, pct_thresh: float, abs_thresh: float) -> pd.DataFrame:
    merged = cp.merge(pp, on=["account_code", "account_name"], suffixes=("_cp", "_pp"), how="outer")
    merged["balance_cp"] = merged["balance_cp"].fillna(0)
    merged["balance_pp"] = merged["balance_pp"].fillna(0)
    merged["delta_abs"] = merged["balance_cp"] - merged["balance_pp"]
    merged["delta_pct"] = merged.apply(
        lambda r: r["delta_abs"] / r["balance_pp"] if r["balance_pp"] != 0 else (1.0 if r["delta_abs"] else 0.0),
        axis=1,
    )
    merged["material"] = (
        (merged["delta_abs"].abs() >= abs_thresh)
        & ((merged["delta_pct"].abs() >= pct_thresh) | (merged["balance_pp"].abs() == 0))
    )
    merged["driver_rank"] = merged["delta_abs"].abs().rank(ascending=False, method="dense")
    return merged.sort_values("driver_rank")


def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX


def _autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(45, max_len + 2)


def write_report(merged: pd.DataFrame, pct_thresh: float, abs_thresh: float,
                 cp_label: str, pp_label: str, output: Path) -> None:
    material = merged[merged["material"]].copy()
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Flux Variance Analysis"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Current Period", cp_label])
    ws.append(["Prior Period", pp_label])
    ws.append(["Pct Threshold", pct_thresh])
    ws.append(["Abs Threshold", abs_thresh])
    ws.append(["Accounts Analyzed", len(merged)])
    ws.append(["Material Variances", len(material)])
    ws.append(["Total Absolute Flux", float(merged["delta_abs"].abs().sum())])
    ws["B8"].number_format = ACCOUNTING_FMT
    _autosize(ws)

    ws_d = wb.create_sheet("Detail")
    cols = ["account_code", "account_name", "balance_pp", "balance_cp", "delta_abs", "delta_pct", "material", "driver_rank"]
    ws_d.append(cols)
    _style_header(ws_d)
    for _, r in merged.iterrows():
        ws_d.append([r[c] for c in cols])
    for r_idx, r in enumerate(merged.itertuples(index=False), start=2):
        if getattr(r, "material"):
            for cell in ws_d[r_idx]:
                cell.fill = RED_FILL
    for letter in ("C", "D", "E"):
        for cell in ws_d[letter][1:]:
            cell.number_format = ACCOUNTING_FMT
    for cell in ws_d["F"][1:]:
        cell.number_format = PCT_FMT
    _autosize(ws_d)

    ws_top = wb.create_sheet("TopDrivers")
    ws_top.append(["Rank", "Account", "Delta", "Pct Change", "Suggested Review"])
    _style_header(ws_top)
    for _, r in material.head(10).iterrows():
        review = "Investigate — new account" if r["balance_pp"] == 0 else "Investigate — material flux"
        ws_top.append([int(r["driver_rank"]), r["account_name"], r["delta_abs"], r["delta_pct"], review])
    for letter in ("C",):
        for cell in ws_top[letter][1:]:
            cell.number_format = ACCOUNTING_FMT
    for letter in ("D",):
        for cell in ws_top[letter][1:]:
            cell.number_format = PCT_FMT
    _autosize(ws_top)

    ws_at = wb.create_sheet("AuditTrail")
    ws_at.append(["Key", "Value"])
    _style_header(ws_at)
    ws_at.append(["Generated", datetime.now().isoformat(timespec="seconds")])
    _autosize(ws_at)

    ws_so = wb.create_sheet("SignOff")
    ws_so.append(["Role", "Name", "Title", "Date", "Notes"])
    _style_header(ws_so)
    for role in ("Preparer", "Reviewer", "Controller"):
        ws_so.append([role, "", "", "", ""])
    _autosize(ws_so)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    LOG.info("Report written: %s", output)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--cp", required=True, type=Path, help="Current period TB")
    parser.add_argument("--pp", required=True, type=Path, help="Prior period TB")
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--pct-threshold", type=float, default=MATERIALITY_PCT_DEFAULT)
    parser.add_argument("--abs-threshold", type=float, default=MATERIALITY_ABS_DEFAULT)
    parser.add_argument("--quiet", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.WARNING if args.quiet else logging.INFO,
        format="%(levelname)s %(name)s: %(message)s",
    )

    try:
        cp = _read_tb(args.cp)
        pp = _read_tb(args.pp)
    except ValueError as exc:
        LOG.error("%s", exc)
        return 1

    merged = analyze(cp, pp, args.pct_threshold, args.abs_threshold)
    write_report(merged, args.pct_threshold, args.abs_threshold,
                 args.cp.name, args.pp.name, args.output)

    material_count = int(merged["material"].sum())
    print(f"Accounts:        {len(merged)}")
    print(f"Material:        {material_count}")
    print(f"Report:          {args.output}")
    return 2 if material_count else 0


if __name__ == "__main__":
    sys.exit(main())
