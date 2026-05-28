"""Intercompany Elimination Bot — match IC balances with FX difference handling.

Implements the workflow in ../SKILL.md.
"""
from __future__ import annotations

import argparse
import logging
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz

LOG = logging.getLogger("ic_elimination")

ACCOUNTING_FMT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN = Side(border_style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

REQUIRED = ("account_code", "account_name", "balance")
MATCH_THRESHOLD = 70
FX_TOLERANCE_DEFAULT = 100.0


def _read_tb(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path)
    else:
        df = pd.read_excel(path)
    df.columns = [c.strip().lower() for c in df.columns]
    missing = [c for c in REQUIRED if c not in df.columns]
    if missing:
        raise ValueError(f"Trial balance missing columns: {missing}")
    df = df.copy()
    df["balance"] = pd.to_numeric(df["balance"], errors="coerce").fillna(0)
    if "currency" not in df.columns:
        df["currency"] = "USD"
    if "fx_rate" not in df.columns:
        df["fx_rate"] = 1.0
    else:
        df["fx_rate"] = pd.to_numeric(df["fx_rate"], errors="coerce").fillna(1.0)
    df["balance_usd"] = df["balance"] * df["fx_rate"]
    return df


def _find_ic_accounts(df: pd.DataFrame, keyword: str = "ic") -> pd.DataFrame:
    mask = df["account_name"].astype(str).str.lower().str.contains(keyword) | \
           df["account_code"].astype(str).str.lower().str.contains(keyword)
    return df[mask].copy()


def match_pair(a_df: pd.DataFrame, b_df: pd.DataFrame, acct_a: str, acct_b: str,
               fx_tolerance: float) -> dict:
    row_a = a_df[a_df["account_code"] == acct_a]
    row_b = b_df[b_df["account_code"] == acct_b]
    val_a = float(row_a["balance_usd"].sum()) if not row_a.empty else 0.0
    val_b = float(row_b["balance_usd"].sum()) if not row_b.empty else 0.0
    # Due from (asset +) should offset Due to (liability -)
    variance = val_a + val_b
    fx_diff = abs(variance)
    status = "PASS" if fx_diff <= fx_tolerance else "FAIL"
    name_score = 0.0
    if not row_a.empty and not row_b.empty:
        name_score = fuzz.token_set_ratio(str(row_a.iloc[0]["account_name"]),
                                          str(row_b.iloc[0]["account_name"]))
    return {
        "entity_a_account": acct_a,
        "entity_b_account": acct_b,
        "balance_a_usd": val_a,
        "balance_b_usd": val_b,
        "variance": variance,
        "fx_diff": fx_diff,
        "name_similarity": name_score,
        "status": status,
    }


def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX


def _autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(45, max_len + 2)


def write_report(pairs: list[dict], entity_a: str, entity_b: str, output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Intercompany Reconciliation"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Entity A", entity_a])
    ws.append(["Entity B", entity_b])
    ws.append(["Pairs Tested", len(pairs)])
    ws.append(["PASS", sum(1 for p in pairs if p["status"] == "PASS")])
    ws.append(["FAIL", sum(1 for p in pairs if p["status"] == "FAIL")])
    ws.append(["Total Variance", sum(abs(p["variance"]) for p in pairs)])
    ws["B7"].number_format = ACCOUNTING_FMT
    _autosize(ws)

    ws_d = wb.create_sheet("Detail")
    ws_d.append(["Entity A Acct", "Entity B Acct", "Balance A (USD)", "Balance B (USD)",
                 "Variance", "FX Diff", "Name Similarity", "Status"])
    _style_header(ws_d)
    for p in pairs:
        ws_d.append([
            p["entity_a_account"], p["entity_b_account"],
            p["balance_a_usd"], p["balance_b_usd"],
            p["variance"], p["fx_diff"], p["name_similarity"], p["status"],
        ])
    for r_idx, p in enumerate(pairs, start=2):
        if p["status"] == "FAIL":
            for cell in ws_d[r_idx]:
                cell.fill = RED_FILL
    for letter in ("C", "D", "E", "F"):
        for cell in ws_d[letter][1:]:
            cell.number_format = ACCOUNTING_FMT
    _autosize(ws_d)

    ws_je = wb.create_sheet("EliminationJE")
    ws_je.append(["Account", "Debit", "Credit", "Memo"])
    _style_header(ws_je)
    for p in pairs:
        if p["status"] == "PASS" and abs(p["variance"]) > 0.01:
            amt = abs(p["variance"]) / 2
            ws_je.append([p["entity_a_account"], 0, amt, "IC elimination"])
            ws_je.append([p["entity_b_account"], amt, 0, "IC elimination"])
    _autosize(ws_je)

    ws_at = wb.create_sheet("AuditTrail")
    ws_at.append(["Key", "Value"])
    _style_header(ws_at)
    ws_at.append(["Generated", datetime.now().isoformat(timespec="seconds")])
    _autosize(ws_at)

    ws_so = wb.create_sheet("SignOff")
    ws_so.append(["Role", "Name", "Title", "Date", "Notes"])
    _style_header(ws_so)
    for role in ("Preparer", "Consolidation Lead", "Controller"):
        ws_so.append([role, "", "", "", ""])
    _autosize(ws_so)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    LOG.info("Report written: %s", output)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--entity-a", required=True, type=Path)
    parser.add_argument("--entity-b", required=True, type=Path)
    parser.add_argument("--pair", action="append", nargs=2, metavar=("ACCT_A", "ACCT_B"),
                        help="IC account pair (repeatable)")
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--fx-tolerance", type=float, default=FX_TOLERANCE_DEFAULT)
    parser.add_argument("--quiet", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.WARNING if args.quiet else logging.INFO,
        format="%(levelname)s %(name)s: %(message)s",
    )

    try:
        df_a = _read_tb(args.entity_a)
        df_b = _read_tb(args.entity_b)
    except ValueError as exc:
        LOG.error("%s", exc)
        return 1

    pairs_spec = args.pair or [["1200-IC", "2200-IC"]]
    pairs = [match_pair(df_a, df_b, a, b, args.fx_tolerance) for a, b in pairs_spec]
    write_report(pairs, args.entity_a.stem, args.entity_b.stem, args.output)

    fails = sum(1 for p in pairs if p["status"] == "FAIL")
    print(f"Pairs:           {len(pairs)}")
    print(f"FAIL:            {fails}")
    print(f"Report:          {args.output}")
    return 2 if fails else 0


if __name__ == "__main__":
    sys.exit(main())
