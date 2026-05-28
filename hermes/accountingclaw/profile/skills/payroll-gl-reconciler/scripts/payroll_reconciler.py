"""Payroll GL Reconciler — tie payroll register to GL by component & department.

Implements the workflow in ../SKILL.md.
"""
from __future__ import annotations

import argparse
import logging
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

LOG = logging.getLogger("payroll_reconciler")

TOLERANCE_DEFAULT = 1.00
ACCOUNTING_FMT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN = Side(border_style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

REGISTER_REQUIRED = (
    "employee_id", "name", "pay_period_end", "gross_pay",
    "er_taxes", "ee_fica", "ee_fit", "ee_sit", "ee_other_taxes",
    "benefits_pretax", "benefits_posttax", "net_pay", "department",
)
COMPONENTS = (
    "gross_pay", "er_taxes", "er_401k_match", "er_health_contribution",
    "benefits_pretax", "benefits_posttax", "ee_withholdings",
)
# net_pay is validated at the row level (gross - withholdings - benefits = net) in RegisterValidation;
# it is intentionally NOT tied to a GL account because cash is usually consolidated across components.


def _read_table(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path)
    else:
        df = pd.read_excel(path)
    df.columns = [c.strip().lower() for c in df.columns]
    return df


def _validate_register(df: pd.DataFrame, tolerance: float) -> tuple[pd.DataFrame, pd.DataFrame]:
    missing = [c for c in REGISTER_REQUIRED if c not in df.columns]
    if missing:
        raise ValueError(f"Register missing required columns: {missing}")
    df = df.copy()
    for c in REGISTER_REQUIRED[3:-1]:  # all numeric columns
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)
    df["pay_period_end"] = pd.to_datetime(df["pay_period_end"], errors="coerce")
    for opt in ("er_401k_match", "er_health_contribution", "class", "entity", "regular_hours",
                "ot_hours", "bonus", "commission"):
        if opt not in df.columns:
            df[opt] = 0.0 if opt not in ("class", "entity") else ""
        else:
            df[opt] = pd.to_numeric(df[opt], errors="coerce").fillna(0.0) if opt not in ("class", "entity") else df[opt].fillna("").astype(str)
    df["__expected_net__"] = (
        df["gross_pay"] - df["ee_fica"] - df["ee_fit"] - df["ee_sit"]
        - df["ee_other_taxes"] - df["benefits_pretax"] - df["benefits_posttax"]
    )
    df["__net_variance__"] = df["net_pay"] - df["__expected_net__"]
    violations = df[df["__net_variance__"].abs() > tolerance].copy()
    return df, violations


def _validate_mapping(df: pd.DataFrame) -> pd.DataFrame:
    required = ("component", "account", "dr_cr")
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Mapping missing required columns: {missing}")
    df = df.copy()
    df["component"] = df["component"].astype(str).str.strip().str.lower()
    df["account"] = df["account"].astype(str).str.strip()
    df["dr_cr"] = df["dr_cr"].astype(str).str.strip().str.upper()
    invalid = df[~df["dr_cr"].isin({"DR", "CR"})]
    if not invalid.empty:
        raise ValueError(f"Mapping has invalid dr_cr values: {invalid['dr_cr'].unique()}")
    invalid_comp = df[~df["component"].isin(COMPONENTS)]
    if not invalid_comp.empty:
        LOG.warning("Mapping has components not in canonical list: %s", invalid_comp["component"].unique())
    return df


def aggregate_register(reg: pd.DataFrame) -> dict[str, float]:
    return {
        "gross_pay": float(reg["gross_pay"].sum()),
        "er_taxes": float(reg["er_taxes"].sum()),
        "er_401k_match": float(reg.get("er_401k_match", pd.Series([0])).sum()),
        "er_health_contribution": float(reg.get("er_health_contribution", pd.Series([0])).sum()),
        "ee_withholdings": float((reg["ee_fica"] + reg["ee_fit"] + reg["ee_sit"] + reg["ee_other_taxes"]).sum()),
        "benefits_pretax": float(reg["benefits_pretax"].sum()),
        "benefits_posttax": float(reg["benefits_posttax"].sum()),
    }


def aggregate_gl(gl: pd.DataFrame, mapping: pd.DataFrame) -> tuple[dict[str, float], pd.DataFrame]:
    """Aggregate GL by component per the mapping. Returns (component_totals, gl_with_component_tag).

    Note: if multiple components map to the same account, the GL amount on that account
    cannot be split. We sum once per (account, sign) and attribute that sum to the
    *first-listed* component, then group sibling components together for a combined check.
    Callers should configure the mapping with distinct accounts per component for cleanest tie-out.
    """
    gl = gl.copy()
    gl["amount"] = pd.to_numeric(gl["amount"], errors="coerce").fillna(0.0)
    gl["account"] = gl["account"].astype(str).str.strip()
    component_totals: dict[str, float] = {c: 0.0 for c in COMPONENTS}
    # Track which components share each account
    acct_to_components: dict[str, list[tuple[str, str]]] = {}
    for _, r in mapping.iterrows():
        acct_to_components.setdefault(r["account"], []).append((r["component"], r["dr_cr"]))
    gl["component"] = ""
    seen_accounts: set[str] = set()
    for idx, r in gl.iterrows():
        comps = acct_to_components.get(r["account"], [])
        if not comps:
            continue
        # Only attribute the GL amount to the FIRST component for this account
        # (siblings share the account and will be checked as a sibling-sum at validation time).
        first_comp, sign = comps[0]
        signed = r["amount"] if sign == "DR" else -r["amount"]
        component_totals[first_comp] += signed
        gl.at[idx, "component"] = first_comp
        seen_accounts.add(r["account"])
    return component_totals, gl


def reconcile(reg_totals: dict[str, float], gl_totals: dict[str, float],
              mapping: pd.DataFrame, tolerance: float) -> list[dict]:
    rows: list[dict] = []
    # Group components that share an account (siblings) so they're checked as a sum
    sibling_groups: dict[str, list[str]] = {}
    for _, r in mapping.iterrows():
        sibling_groups.setdefault(r["account"], []).append(r["component"])
    component_to_group: dict[str, list[str]] = {}
    for siblings in sibling_groups.values():
        if len(set(siblings)) > 1:
            for s in siblings:
                component_to_group[s] = list(set(siblings))
    seen_groups: set[frozenset] = set()
    for c in COMPONENTS:
        siblings = component_to_group.get(c)
        if siblings:
            key = frozenset(siblings)
            if key in seen_groups:
                continue
            seen_groups.add(key)
            label = " + ".join(sorted(siblings))
            reg = sum(reg_totals.get(s, 0.0) for s in siblings)
            gl = sum(gl_totals.get(s, 0.0) for s in siblings)
        else:
            label = c
            reg = reg_totals.get(c, 0.0)
            gl = gl_totals.get(c, 0.0)
        var = reg - gl
        status = "PASS" if abs(var) <= tolerance else "FAIL"
        rows.append({
            "Component": label, "Register": round(reg, 2), "GL": round(gl, 2),
            "Variance": round(var, 2), "Status": status,
        })
    return rows


def departmental_allocation(reg: pd.DataFrame) -> pd.DataFrame:
    grouped = reg.groupby("department").agg(
        gross_pay=("gross_pay", "sum"),
        er_taxes=("er_taxes", "sum"),
        benefits_pretax=("benefits_pretax", "sum"),
        net_pay=("net_pay", "sum"),
        headcount=("employee_id", "nunique"),
    ).round(2).reset_index()
    return grouped


def build_draft_je(reg_totals: dict[str, float], period_end: datetime,
                   mapping: pd.DataFrame, dept_alloc: pd.DataFrame) -> pd.DataFrame:
    """Build a multi-line draft payroll JE allocated by department where possible."""
    lines: list[dict] = []
    line_no = 0
    # DR Salaries Expense by department
    sal_acct = mapping[(mapping["component"] == "gross_pay") & (mapping["dr_cr"] == "DR")]
    sal_acct_code = sal_acct["account"].iloc[0] if not sal_acct.empty else "6100"
    for _, dept in dept_alloc.iterrows():
        if dept["gross_pay"] == 0:
            continue
        line_no += 1
        lines.append({
            "Line": line_no, "Date": period_end.strftime("%Y-%m-%d"),
            "Account": sal_acct_code, "Description": f"Salaries Expense - {dept['department']}",
            "Department": dept["department"], "Debit": round(float(dept["gross_pay"]), 2), "Credit": 0,
        })
    # DR ER Taxes by department
    tax_acct = mapping[(mapping["component"] == "er_taxes") & (mapping["dr_cr"] == "DR")]
    tax_acct_code = tax_acct["account"].iloc[0] if not tax_acct.empty else "6110"
    for _, dept in dept_alloc.iterrows():
        if dept["er_taxes"] == 0:
            continue
        line_no += 1
        lines.append({
            "Line": line_no, "Date": period_end.strftime("%Y-%m-%d"),
            "Account": tax_acct_code, "Description": f"ER Tax Expense - {dept['department']}",
            "Department": dept["department"], "Debit": round(float(dept["er_taxes"]), 2), "Credit": 0,
        })
    # DR ER 401(k) Match (consolidated)
    if reg_totals.get("er_401k_match", 0) > 0:
        match_acct = mapping[(mapping["component"] == "er_401k_match") & (mapping["dr_cr"] == "DR")]
        match_code = match_acct["account"].iloc[0] if not match_acct.empty else "6120"
        line_no += 1
        lines.append({
            "Line": line_no, "Date": period_end.strftime("%Y-%m-%d"),
            "Account": match_code, "Description": "ER 401(k) Match Expense",
            "Department": "", "Debit": round(reg_totals["er_401k_match"], 2), "Credit": 0,
        })
    # DR ER Health Contribution (consolidated)
    if reg_totals.get("er_health_contribution", 0) > 0:
        h_acct = mapping[(mapping["component"] == "er_health_contribution") & (mapping["dr_cr"] == "DR")]
        h_code = h_acct["account"].iloc[0] if not h_acct.empty else "6130"
        line_no += 1
        lines.append({
            "Line": line_no, "Date": period_end.strftime("%Y-%m-%d"),
            "Account": h_code, "Description": "ER Health Contribution Expense",
            "Department": "", "Debit": round(reg_totals["er_health_contribution"], 2), "Credit": 0,
        })
    # CR EE Withholdings
    eew_acct = mapping[(mapping["component"] == "ee_withholdings") & (mapping["dr_cr"] == "CR")]
    eew_code = eew_acct["account"].iloc[0] if not eew_acct.empty else "2210"
    if reg_totals["ee_withholdings"] > 0:
        line_no += 1
        lines.append({
            "Line": line_no, "Date": period_end.strftime("%Y-%m-%d"),
            "Account": eew_code, "Description": "EE Tax Withholding Liability",
            "Department": "", "Debit": 0, "Credit": round(reg_totals["ee_withholdings"], 2),
        })
    # CR Benefits Liability
    if reg_totals["benefits_pretax"] + reg_totals["benefits_posttax"] > 0:
        ben_acct = mapping[(mapping["component"] == "benefits_pretax") & (mapping["dr_cr"] == "CR")]
        ben_code = ben_acct["account"].iloc[0] if not ben_acct.empty else "2220"
        line_no += 1
        lines.append({
            "Line": line_no, "Date": period_end.strftime("%Y-%m-%d"),
            "Account": ben_code, "Description": "Benefits Withholding Liability",
            "Department": "", "Debit": 0,
            "Credit": round(reg_totals["benefits_pretax"] + reg_totals["benefits_posttax"], 2),
        })
    # CR Cash (balancing line: gross + ER costs - liabilities = cash outflow)
    cash_total = (
        reg_totals["gross_pay"] + reg_totals["er_taxes"]
        + reg_totals["er_401k_match"] + reg_totals["er_health_contribution"]
        - reg_totals["benefits_pretax"] - reg_totals["benefits_posttax"]
        - reg_totals["ee_withholdings"]
    )
    if abs(cash_total) > 0.001:
        line_no += 1
        lines.append({
            "Line": line_no, "Date": period_end.strftime("%Y-%m-%d"),
            "Account": "1000", "Description": "Cash - Payroll Disbursement",
            "Department": "", "Debit": 0, "Credit": round(cash_total, 2),
        })
    return pd.DataFrame(lines)


# --- XLSX writer ----------------------------------------------------------

def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX


def _autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(50, max_len + 2)


def _df_to_sheet(ws, df: pd.DataFrame, money_cols: list[str] | None = None) -> None:
    if df is None or df.empty:
        ws.append(["(no data)"])
        return
    ws.append(list(df.columns))
    _style_header(ws)
    for _, r in df.iterrows():
        ws.append([r[c] if not pd.isna(r[c]) else None for c in df.columns])
    if money_cols:
        for col_idx, col in enumerate(df.columns, start=1):
            if col in money_cols:
                for cell in ws[get_column_letter(col_idx)][1:]:
                    cell.number_format = ACCOUNTING_FMT
    _autosize(ws)


def write_workpaper(
    reg: pd.DataFrame, gl: pd.DataFrame, mapping: pd.DataFrame,
    reg_totals: dict[str, float], gl_totals: dict[str, float],
    tieout_rows: list[dict], dept: pd.DataFrame, draft_je: pd.DataFrame,
    violations: pd.DataFrame, period_end: datetime, output: Path,
) -> str:
    overall_status = "PASS" if all(r["Status"] == "PASS" for r in tieout_rows) and violations.empty else "FAIL"
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Payroll Reconciliation"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Period End", period_end.strftime("%Y-%m-%d")])
    ws.append(["Headcount", reg["employee_id"].nunique()])
    ws.append(["Total Gross Pay (Register)", reg_totals["gross_pay"]])
    ws.append(["Total Net Pay (Register)", float(reg["net_pay"].sum())])
    ws.append(["Status", overall_status])
    for cell in ("B5", "B6"):
        ws[cell].number_format = ACCOUNTING_FMT
    ws["B7"].font = Font(bold=True)
    ws["B7"].fill = GREEN_FILL if overall_status == "PASS" else RED_FILL
    _autosize(ws)

    # Component tie-out
    ws_c = wb.create_sheet("ComponentTieOut")
    ws_c.append(["Component", "Register Total", "GL Total", "Variance", "Status"])
    _style_header(ws_c)
    for r in tieout_rows:
        ws_c.append([r["Component"], r["Register"], r["GL"], r["Variance"], r["Status"]])
    for letter in ("B", "C", "D"):
        for cell in ws_c[letter][1:]:
            cell.number_format = ACCOUNTING_FMT
    for row in ws_c.iter_rows(min_row=2):
        if row[4].value == "FAIL":
            for c in row:
                c.fill = RED_FILL
        elif row[4].value == "PASS":
            row[4].fill = GREEN_FILL
    _autosize(ws_c)

    # Departmental allocation
    _df_to_sheet(
        wb.create_sheet("DepartmentAllocation"), dept,
        money_cols=["gross_pay", "er_taxes", "benefits_pretax", "net_pay"],
    )

    # Draft JE
    _df_to_sheet(wb.create_sheet("DraftJE"), draft_je, money_cols=["Debit", "Credit"])
    ws_je = wb["DraftJE"]
    if not draft_je.empty:
        total_dr = float(draft_je["Debit"].sum())
        total_cr = float(draft_je["Credit"].sum())
        ws_je.append(["", "", "", "TOTAL", "", total_dr, total_cr])
        ws_je[get_column_letter(6)][-1].number_format = ACCOUNTING_FMT
        ws_je[get_column_letter(7)][-1].number_format = ACCOUNTING_FMT
        bal_ok = abs(total_dr - total_cr) < 0.01
        ws_je.append(["", "", "", "BALANCED" if bal_ok else "OUT OF BALANCE", "", "", ""])
        ws_je.cell(row=ws_je.max_row, column=4).font = Font(bold=True)
        ws_je.cell(row=ws_je.max_row, column=4).fill = GREEN_FILL if bal_ok else RED_FILL

    # Register validation
    if not violations.empty:
        v = violations[["employee_id", "name", "department", "gross_pay", "net_pay",
                        "__expected_net__", "__net_variance__"]].rename(
            columns={"__expected_net__": "Expected Net", "__net_variance__": "Variance"})
        _df_to_sheet(wb.create_sheet("RegisterValidation"), v,
                     money_cols=["gross_pay", "net_pay", "Expected Net", "Variance"])
    else:
        ws_v = wb.create_sheet("RegisterValidation")
        ws_v.append(["(all register rows tie: gross - withholdings - benefits = net)"])
        _autosize(ws_v)

    # GL Detail
    _df_to_sheet(wb.create_sheet("GLDetail"), gl[["date", "account", "description",
                                                  "amount", "department", "component"]],
                 money_cols=["amount"])

    # Audit Trail
    ws_a = wb.create_sheet("AuditTrail")
    ws_a.append(["Key", "Value"])
    _style_header(ws_a)
    ws_a.append(["Period End", period_end.strftime("%Y-%m-%d")])
    ws_a.append(["Register Rows", len(reg)])
    ws_a.append(["GL Rows", len(gl)])
    ws_a.append(["Mapping Rows", len(mapping)])
    ws_a.append(["Components Checked", len(tieout_rows)])
    ws_a.append(["Overall Status", overall_status])
    ws_a.append(["Generated", datetime.now().isoformat(timespec="seconds")])
    _autosize(ws_a)

    # Sign-Off
    ws_s = wb.create_sheet("SignOff")
    ws_s.append(["Role", "Name", "Title", "Date", "Notes"])
    _style_header(ws_s)
    ws_s.append(["Preparer (Payroll Accountant)", "", "", "", ""])
    ws_s.append(["Reviewer (Senior Accountant)", "", "", "", ""])
    ws_s.append(["Approver (Controller)", "", "", "", ""])
    _autosize(ws_s)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    LOG.info("Workpaper written: %s", output)
    return overall_status


# --- CLI ------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--register", required=True, type=Path)
    parser.add_argument("--gl", required=True, type=Path)
    parser.add_argument("--mapping", required=True, type=Path)
    parser.add_argument("--period-end", required=True)
    parser.add_argument("--tie-out-tolerance", type=float, default=TOLERANCE_DEFAULT)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--quiet", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.WARNING if args.quiet else logging.INFO,
        format="%(levelname)s %(name)s: %(message)s",
    )

    try:
        reg_raw = _read_table(args.register)
        gl = _read_table(args.gl)
        mapping = _validate_mapping(_read_table(args.mapping))
        reg, violations = _validate_register(reg_raw, args.tie_out_tolerance)
    except ValueError as e:
        LOG.error("Input validation failed: %s", e)
        return 1

    reg_totals = aggregate_register(reg)
    gl_totals, gl_tagged = aggregate_gl(gl, mapping)
    tieout = reconcile(reg_totals, gl_totals, mapping, args.tie_out_tolerance)
    dept = departmental_allocation(reg)
    period_end = datetime.strptime(args.period_end, "%Y-%m-%d")
    draft_je = build_draft_je(reg_totals, period_end, mapping, dept)

    status = write_workpaper(reg, gl_tagged, mapping, reg_totals, gl_totals,
                             tieout, dept, draft_je, violations, period_end, args.output)

    print(f"Headcount:        {reg['employee_id'].nunique()}")
    print(f"Gross Pay (reg):  {reg_totals['gross_pay']:>15,.2f}")
    print(f"Net Pay (reg):    {float(reg['net_pay'].sum()):>15,.2f}")
    print(f"Status:           {status}")
    print(f"Workpaper:        {args.output}")
    return 0 if status == "PASS" else 2


if __name__ == "__main__":
    sys.exit(main())
