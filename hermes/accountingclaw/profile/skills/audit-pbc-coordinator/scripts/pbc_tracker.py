"""Audit PBC Coordinator — request status tracking, aging, and completeness reporting.

Implements the workflow in ../SKILL.md.
"""
from __future__ import annotations

import argparse
import logging
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

LOG = logging.getLogger("pbc_coordinator")

HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN = Side(border_style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

REQUIRED = ("request_id", "description", "owner", "due_date", "status")
PROVIDED_STATUSES = {"provided", "complete", "delivered"}


def _read_tracker(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path)
    else:
        df = pd.read_excel(path)
    df.columns = [c.strip().lower() for c in df.columns]
    missing = [c for c in REQUIRED if c not in df.columns]
    if missing:
        raise ValueError(f"PBC tracker missing columns: {missing}")
    df = df.copy()
    df["due_date"] = pd.to_datetime(df["due_date"], errors="coerce")
    df["status_norm"] = df["status"].astype(str).str.strip().str.lower()
    if "priority" not in df.columns:
        df["priority"] = "Normal"
    return df


def analyze(df: pd.DataFrame, as_of: datetime) -> dict:
    total = len(df)
    provided = int(df["status_norm"].isin(PROVIDED_STATUSES).sum())
    pending = df[~df["status_norm"].isin(PROVIDED_STATUSES)].copy()
    overdue = pending[pending["due_date"] < pd.Timestamp(as_of)].copy()
    pending["days_overdue"] = (pd.Timestamp(as_of) - pending["due_date"]).dt.days.clip(lower=0)
    owner_summary = (
        pending.groupby("owner")
        .agg(count=("request_id", "count"), avg_overdue=("days_overdue", "mean"))
        .reset_index()
        .sort_values("count", ascending=False)
    )
    return {
        "total": total,
        "provided": provided,
        "pending_count": len(pending),
        "overdue_count": len(overdue),
        "pct_complete": round(provided / total * 100, 1) if total else 0,
        "pending": pending,
        "overdue": overdue,
        "owner_summary": owner_summary,
    }


def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX


def _autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(45, max_len + 2)


def write_report(df: pd.DataFrame, stats: dict, as_of: datetime, output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["PBC Request Status Dashboard"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["As Of", as_of.strftime("%Y-%m-%d")])
    ws.append(["Total Requests", stats["total"]])
    ws.append(["Provided", stats["provided"]])
    ws.append(["Pending", stats["pending_count"]])
    ws.append(["Overdue", stats["overdue_count"]])
    ws.append(["Percent Complete", f"{stats['pct_complete']:.1f}%"])
    _autosize(ws)

    ws_p = wb.create_sheet("Pending")
    cols = ["request_id", "description", "owner", "due_date", "status", "days_overdue", "priority"]
    ws_p.append(cols)
    _style_header(ws_p)
    pending = stats["pending"]
    if not pending.empty:
        for _, r in pending.iterrows():
            ws_p.append([
                r["request_id"], r["description"], r["owner"],
                r["due_date"].strftime("%Y-%m-%d") if pd.notna(r["due_date"]) else "",
                r["status"], int(r["days_overdue"]), r.get("priority", "Normal"),
            ])
    else:
        ws_p.append(["(none)"])
    _autosize(ws_p)

    ws_o = wb.create_sheet("Overdue")
    ws_o.append(["request_id", "description", "owner", "due_date", "days_overdue"])
    _style_header(ws_o)
    overdue = stats["overdue"]
    if not overdue.empty:
        for _, r in overdue.iterrows():
            days = (pd.Timestamp(as_of) - r["due_date"]).days
            ws_o.append([r["request_id"], r["description"], r["owner"],
                         r["due_date"].strftime("%Y-%m-%d"), days])
    else:
        ws_o.append(["(none)"])
    _autosize(ws_o)

    ws_own = wb.create_sheet("OwnerSummary")
    ws_own.append(["owner", "pending_count", "avg_days_overdue"])
    _style_header(ws_own)
    for _, r in stats["owner_summary"].iterrows():
        ws_own.append([r["owner"], int(r["count"]), round(float(r["avg_overdue"]), 1)])
    _autosize(ws_own)

    ws_at = wb.create_sheet("AuditTrail")
    ws_at.append(["Key", "Value"])
    _style_header(ws_at)
    ws_at.append(["Generated", datetime.now().isoformat(timespec="seconds")])
    _autosize(ws_at)

    ws_so = wb.create_sheet("SignOff")
    ws_so.append(["Role", "Name", "Title", "Date", "Notes"])
    _style_header(ws_so)
    for role in ("Audit Coordinator", "Controller", "External Auditor"):
        ws_so.append([role, "", "", "", ""])
    _autosize(ws_so)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    LOG.info("Report written: %s", output)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--tracker", required=True, type=Path)
    parser.add_argument("--as-of", required=True)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--quiet", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.WARNING if args.quiet else logging.INFO,
        format="%(levelname)s %(name)s: %(message)s",
    )

    try:
        df = _read_tracker(args.tracker)
    except ValueError as exc:
        LOG.error("%s", exc)
        return 1

    as_of = datetime.strptime(args.as_of, "%Y-%m-%d")
    stats = analyze(df, as_of)
    write_report(df, stats, as_of, args.output)

    print(f"Complete:        {stats['pct_complete']:.1f}%")
    print(f"Pending:         {stats['pending_count']}")
    print(f"Overdue:         {stats['overdue_count']}")
    print(f"Report:          {args.output}")
    return 2 if stats["overdue_count"] else 0


if __name__ == "__main__":
    sys.exit(main())
