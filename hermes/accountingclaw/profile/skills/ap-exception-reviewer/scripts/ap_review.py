"""AP Exception Reviewer — scan an AP ledger for control exceptions.

Implements the workflow described in ../SKILL.md. Detects duplicates,
near-duplicates, split-invoice patterns, coding inconsistencies, three-way-match
failures, approval-limit breaches, weekend payments, round-dollar invoices,
and vendor-master integrity issues; produces a risk-scored XLSX exception report.
"""
from __future__ import annotations

import argparse
import logging
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz

LOG = logging.getLogger("ap_exception_reviewer")

# --- Constants -------------------------------------------------------------
DUPLICATE_WINDOW_DAYS_DEFAULT = 45
AMOUNT_TOLERANCE_PCT_DEFAULT = 0.005
SPLIT_THRESHOLD_USD_DEFAULT = 10_000.0
APPROVAL_LIMIT_USD_DEFAULT = 10_000.0
THREE_WAY_TOLERANCE_PCT_DEFAULT = 0.05
ROUND_DOLLAR_THRESHOLD_DEFAULT = 1_000.0
SPLIT_CLUSTER_WINDOW_DAYS = 14
VENDOR_CLUSTER_SIMILARITY = 90

SEVERITY = {
    "exact_duplicate": 10,
    "near_duplicate": 8,
    "split_invoice": 8,
    "three_way_fail": 7,
    "coding_inconsistency": 5,
    "no_approver": 8,
    "weekend_payment": 2,
    "round_dollar": 3,
    "inactive_vendor": 9,
    "shared_bank_account": 10,
    "missing_tin_1099": 6,
    "terms_anomaly": 4,
}
RISK_BUCKETS = [(20, "Critical"), (10, "High"), (5, "Medium"), (0, "Low")]

ACCOUNTING_FMT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN = Side(border_style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


@dataclass
class Exception_:
    invoice_idx: int
    triggers: list[str] = field(default_factory=list)
    detail: list[str] = field(default_factory=list)
    risk_score: int = 0


def _normalize_vendor(v: str) -> str:
    return re.sub(r"[^a-z0-9 ]", " ", v.lower()).split()


def _vendor_clusters(vendors: Iterable[str], threshold: int = VENDOR_CLUSTER_SIMILARITY) -> dict[str, str]:
    """Return mapping {original_vendor → canonical_vendor} using rapidfuzz token_set_ratio clustering."""
    canonical: dict[str, str] = {}
    seen: list[tuple[str, str]] = []  # (vendor_lower, original)
    for v in vendors:
        if not isinstance(v, str) or not v.strip():
            canonical[v] = v
            continue
        v_l = v.lower().strip()
        merged = None
        for existing_l, existing_orig in seen:
            if fuzz.token_set_ratio(v_l, existing_l) >= threshold:
                merged = existing_orig
                break
        if merged is None:
            canonical[v] = v
            seen.append((v_l, v))
        else:
            canonical[v] = merged
    return canonical


def _read_table(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path)
    else:
        df = pd.read_excel(path)
    df.columns = [c.strip().lower() for c in df.columns]
    return df


def _validate_ledger(df: pd.DataFrame) -> pd.DataFrame:
    required = ("invoice_no", "vendor", "amount", "date", "account")
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Ledger missing required columns: {missing}")
    df = df.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df["amount"] = pd.to_numeric(df["amount"], errors="coerce")
    df = df.dropna(subset=["date", "amount"])
    for opt in ("vendor_id", "po_number", "terms", "due_date", "pay_date",
                "approver", "currency", "category", "payment_method"):
        if opt not in df.columns:
            df[opt] = ""
        else:
            df[opt] = df[opt].fillna("").astype(str).replace({"nan": ""})
    if df["due_date"].astype(bool).any():
        df["due_date"] = pd.to_datetime(df["due_date"], errors="coerce")
    if df["pay_date"].astype(bool).any():
        df["pay_date"] = pd.to_datetime(df["pay_date"], errors="coerce")
    df["invoice_no"] = df["invoice_no"].astype(str).str.strip().str.lstrip("0").str.lower()
    df["vendor"] = df["vendor"].astype(str).str.strip()
    df = df.reset_index(drop=True)
    df["__idx__"] = df.index
    return df


# --- Detectors -------------------------------------------------------------

def _detect_exact_duplicates(df: pd.DataFrame, exceptions: dict[int, Exception_]) -> pd.DataFrame:
    key_cols = ["vendor_canonical", "invoice_no", "amount_round"]
    df = df.copy()
    df["amount_round"] = df["amount"].round(2)
    groups = df.groupby(key_cols).filter(lambda g: len(g) > 1)
    flagged_rows: list[dict] = []
    for _, group in groups.groupby(key_cols):
        for _, row in group.iterrows():
            ex = exceptions.setdefault(int(row["__idx__"]), Exception_(invoice_idx=int(row["__idx__"])))
            ex.triggers.append("exact_duplicate")
            ex.detail.append(
                f"Group of {len(group)}: invoice_no={row['invoice_no']!r} amount={row['amount']:.2f}"
            )
            ex.risk_score += SEVERITY["exact_duplicate"]
            flagged_rows.append(row.to_dict())
    return pd.DataFrame(flagged_rows)


def _detect_near_duplicates(df: pd.DataFrame, exceptions: dict[int, Exception_],
                            window_days: int, amount_tol_pct: float) -> pd.DataFrame:
    pairs: list[dict] = []
    flagged_idx: set[int] = set()
    for vendor, sub in df.groupby("vendor_canonical"):
        rows = sub.sort_values("date").to_dict("records")
        for i in range(len(rows)):
            for j in range(i + 1, len(rows)):
                a, b = rows[i], rows[j]
                # skip if exact-duplicate (handled separately)
                if a["invoice_no"] == b["invoice_no"]:
                    continue
                day_diff = abs((a["date"] - b["date"]).days)
                if day_diff > window_days:
                    break  # sorted; further pairs only grow
                amt_diff_pct = abs(a["amount"] - b["amount"]) / max(abs(a["amount"]), 0.01)
                if amt_diff_pct <= amount_tol_pct:
                    pairs.append({
                        "Vendor": a["vendor"],
                        "Invoice A": a["invoice_no"],
                        "Invoice B": b["invoice_no"],
                        "Date A": a["date"],
                        "Date B": b["date"],
                        "Amount A": a["amount"],
                        "Amount B": b["amount"],
                        "Day Diff": day_diff,
                    })
                    for idx in (int(a["__idx__"]), int(b["__idx__"])):
                        if idx in flagged_idx:
                            continue
                        flagged_idx.add(idx)
                        ex = exceptions.setdefault(idx, Exception_(invoice_idx=idx))
                        ex.triggers.append("near_duplicate")
                        ex.detail.append(
                            f"Near-duplicate with invoice {b['invoice_no'] if idx == int(a['__idx__']) else a['invoice_no']} "
                            f"({day_diff} day gap; {amt_diff_pct*100:.2f}% amount diff)"
                        )
                        ex.risk_score += SEVERITY["near_duplicate"]
    return pd.DataFrame(pairs)


def _detect_splits(df: pd.DataFrame, exceptions: dict[int, Exception_],
                   split_threshold: float, cluster_days: int) -> pd.DataFrame:
    clusters: list[dict] = []
    for vendor, sub in df.groupby("vendor_canonical"):
        sub = sub.sort_values("date")
        rows = sub.to_dict("records")
        n = len(rows)
        for i in range(n):
            cluster = [rows[i]]
            for j in range(i + 1, n):
                if (rows[j]["date"] - rows[i]["date"]).days <= cluster_days:
                    cluster.append(rows[j])
                else:
                    break
            if len(cluster) < 2:
                continue
            if all(r["amount"] < split_threshold for r in cluster) and sum(r["amount"] for r in cluster) > 2 * split_threshold:
                clusters.append({
                    "Vendor": cluster[0]["vendor"],
                    "Date Range": f"{cluster[0]['date'].date()}–{cluster[-1]['date'].date()}",
                    "Invoice Count": len(cluster),
                    "Total Amount": sum(r["amount"] for r in cluster),
                    "Threshold": split_threshold,
                    "Invoices": ", ".join(r["invoice_no"] for r in cluster),
                })
                for r in cluster:
                    ex = exceptions.setdefault(int(r["__idx__"]), Exception_(invoice_idx=int(r["__idx__"])))
                    ex.triggers.append("split_invoice")
                    ex.detail.append(f"Part of split cluster: {len(cluster)} invoices totaling ${sum(r['amount'] for r in cluster):.2f}")
                    ex.risk_score += SEVERITY["split_invoice"]
    return pd.DataFrame(clusters)


def _detect_coding_inconsistency(df: pd.DataFrame, exceptions: dict[int, Exception_]) -> pd.DataFrame:
    rows: list[dict] = []
    for vendor, sub in df.groupby("vendor_canonical"):
        accts = sub["account"].astype(str).str.strip().unique()
        if len(accts) > 1:
            for _, r in sub.iterrows():
                ex = exceptions.setdefault(int(r["__idx__"]), Exception_(invoice_idx=int(r["__idx__"])))
                ex.triggers.append("coding_inconsistency")
                ex.detail.append(f"Vendor coded to {len(accts)} accounts: {sorted(accts)}")
                ex.risk_score += SEVERITY["coding_inconsistency"]
            rows.append({
                "Vendor": sub["vendor"].iloc[0],
                "Account Count": len(accts),
                "Accounts": ", ".join(sorted(accts)),
                "Invoice Count": len(sub),
                "Total Amount": float(sub["amount"].sum()),
            })
    return pd.DataFrame(rows)


def _detect_three_way(df: pd.DataFrame, pos: pd.DataFrame, exceptions: dict[int, Exception_],
                      tolerance_pct: float) -> pd.DataFrame:
    if pos is None or pos.empty:
        return pd.DataFrame()
    po_index = {str(r["po_number"]).strip().lower(): r for _, r in pos.iterrows()}
    rows: list[dict] = []
    for _, r in df.iterrows():
        po_no = str(r.get("po_number", "")).strip().lower()
        if not po_no:
            continue
        po = po_index.get(po_no)
        reasons: list[str] = []
        if po is None:
            reasons.append("PO not found in PO file")
        else:
            if str(po.get("vendor", "")).lower() != str(r["vendor"]).lower():
                reasons.append(f"PO vendor '{po['vendor']}' != invoice vendor '{r['vendor']}'")
            if not bool(po.get("received", False)):
                reasons.append("PO not marked received")
            po_amt = float(po.get("po_amount", 0))
            if po_amt > 0:
                diff = abs(r["amount"] - po_amt) / po_amt
                if diff > tolerance_pct:
                    reasons.append(f"Amount diff {diff*100:.2f}% > tolerance {tolerance_pct*100:.2f}%")
        if reasons:
            rows.append({
                "Invoice": r["invoice_no"], "Vendor": r["vendor"],
                "PO": po_no, "Invoice Amount": r["amount"],
                "Reasons": "; ".join(reasons),
            })
            ex = exceptions.setdefault(int(r["__idx__"]), Exception_(invoice_idx=int(r["__idx__"])))
            ex.triggers.append("three_way_fail")
            ex.detail.append("; ".join(reasons))
            ex.risk_score += SEVERITY["three_way_fail"]
    return pd.DataFrame(rows)


def _detect_approval(df: pd.DataFrame, exceptions: dict[int, Exception_], limit: float) -> None:
    for _, r in df.iterrows():
        if r["amount"] > limit and not str(r.get("approver", "")).strip():
            ex = exceptions.setdefault(int(r["__idx__"]), Exception_(invoice_idx=int(r["__idx__"])))
            ex.triggers.append("no_approver")
            ex.detail.append(f"Amount ${r['amount']:.2f} > limit ${limit:.2f} with no approver listed")
            ex.risk_score += SEVERITY["no_approver"]


def _detect_weekend_payment(df: pd.DataFrame, exceptions: dict[int, Exception_]) -> None:
    if "pay_date" not in df.columns:
        return
    for _, r in df.iterrows():
        if isinstance(r.get("pay_date"), pd.Timestamp) and r["pay_date"].weekday() >= 5:
            ex = exceptions.setdefault(int(r["__idx__"]), Exception_(invoice_idx=int(r["__idx__"])))
            ex.triggers.append("weekend_payment")
            ex.detail.append(f"Pay date {r['pay_date'].date()} is weekend")
            ex.risk_score += SEVERITY["weekend_payment"]


def _detect_round_dollar(df: pd.DataFrame, exceptions: dict[int, Exception_], threshold: float) -> None:
    for _, r in df.iterrows():
        if r["amount"] >= threshold and abs(r["amount"] - round(r["amount"])) < 1e-9 and round(r["amount"]) % 100 == 0:
            ex = exceptions.setdefault(int(r["__idx__"]), Exception_(invoice_idx=int(r["__idx__"])))
            ex.triggers.append("round_dollar")
            ex.detail.append(f"Round-dollar amount ${r['amount']:.2f}")
            ex.risk_score += SEVERITY["round_dollar"]


def _detect_vendor_integrity(df: pd.DataFrame, vendors: pd.DataFrame | None,
                             exceptions: dict[int, Exception_]) -> pd.DataFrame:
    if vendors is None or vendors.empty:
        return pd.DataFrame()
    integrity: list[dict] = []
    v_idx = {str(r["vendor"]).lower().strip(): r for _, r in vendors.iterrows()}
    bank_to_vendors: dict[str, list[str]] = defaultdict(list)
    for _, vr in vendors.iterrows():
        b = str(vr.get("bank_account", "")).strip()
        if b:
            bank_to_vendors[b].append(str(vr["vendor"]))
    shared_banks = {b: vs for b, vs in bank_to_vendors.items() if len(set(vs)) > 1}
    for bank, vs in shared_banks.items():
        integrity.append({"issue": "shared_bank_account", "detail": f"Bank {bank} used by {vs}"})

    # Apply to invoices
    for _, r in df.iterrows():
        v = str(r["vendor"]).lower().strip()
        vrow = v_idx.get(v)
        if vrow is None:
            continue
        if not bool(vrow.get("is_active", True)):
            ex = exceptions.setdefault(int(r["__idx__"]), Exception_(invoice_idx=int(r["__idx__"])))
            ex.triggers.append("inactive_vendor")
            ex.detail.append("Vendor flagged inactive in master")
            ex.risk_score += SEVERITY["inactive_vendor"]
        bank = str(vrow.get("bank_account", "")).strip()
        if bank in shared_banks:
            ex = exceptions.setdefault(int(r["__idx__"]), Exception_(invoice_idx=int(r["__idx__"])))
            ex.triggers.append("shared_bank_account")
            ex.detail.append(f"Vendor shares bank account {bank} with: {shared_banks[bank]}")
            ex.risk_score += SEVERITY["shared_bank_account"]
        if bool(vrow.get("is_1099", False)) and not str(vrow.get("tin", "")).strip():
            ex = exceptions.setdefault(int(r["__idx__"]), Exception_(invoice_idx=int(r["__idx__"])))
            ex.triggers.append("missing_tin_1099")
            ex.detail.append("1099 vendor with no TIN on file")
            ex.risk_score += SEVERITY["missing_tin_1099"]
    return pd.DataFrame(integrity)


def _bucket(score: int) -> str:
    for t, label in RISK_BUCKETS:
        if score >= t:
            return label
    return "Low"


# --- Orchestration --------------------------------------------------------

def review(
    ledger: pd.DataFrame, pos: pd.DataFrame | None, vendors: pd.DataFrame | None,
    duplicate_window_days: int, amount_tolerance_pct: float, split_threshold: float,
    approval_limit: float, three_way_tol: float, round_dollar_threshold: float,
) -> dict:
    LOG.info("AP review starting: %d invoices", len(ledger))
    # Canonical vendor clustering
    mapping = _vendor_clusters(ledger["vendor"].unique())
    ledger = ledger.copy()
    ledger["vendor_canonical"] = ledger["vendor"].map(mapping)

    exceptions: dict[int, Exception_] = {}
    audit: list[dict] = []
    audit.append({"step": "load", "invoice_count": len(ledger)})

    exact_dupes = _detect_exact_duplicates(ledger, exceptions)
    audit.append({"step": "exact_duplicates", "flagged_rows": len(exact_dupes)})

    near_dupes = _detect_near_duplicates(ledger, exceptions, duplicate_window_days, amount_tolerance_pct)
    audit.append({"step": "near_duplicates", "pairs": len(near_dupes)})

    splits = _detect_splits(ledger, exceptions, split_threshold, SPLIT_CLUSTER_WINDOW_DAYS)
    audit.append({"step": "splits", "clusters": len(splits)})

    coding = _detect_coding_inconsistency(ledger, exceptions)
    audit.append({"step": "coding", "inconsistent_vendors": len(coding)})

    three_way = _detect_three_way(ledger, pos, exceptions, three_way_tol)
    audit.append({"step": "three_way", "failures": len(three_way)})

    _detect_approval(ledger, exceptions, approval_limit)
    _detect_weekend_payment(ledger, exceptions)
    _detect_round_dollar(ledger, exceptions, round_dollar_threshold)
    integrity = _detect_vendor_integrity(ledger, vendors, exceptions)
    audit.append({"step": "vendor_integrity", "issues": len(integrity)})

    audit.append({"step": "total_exceptions", "count": len(exceptions)})

    return {
        "ledger": ledger,
        "exceptions": exceptions,
        "exact_dupes": exact_dupes,
        "near_dupes": near_dupes,
        "splits": splits,
        "coding": coding,
        "three_way": three_way,
        "integrity": integrity,
        "audit": audit,
    }


# --- XLSX writer ----------------------------------------------------------

def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX


def _autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(60, max_len + 2)


def write_report(result: dict, output: Path, ledger_path: Path,
                 thresholds: dict) -> None:
    ledger = result["ledger"]
    exceptions = result["exceptions"]
    wb = Workbook()

    # Build exception detail rows
    exc_rows = []
    for idx, ex in sorted(exceptions.items(), key=lambda kv: -kv[1].risk_score):
        r = ledger.loc[idx]
        exc_rows.append({
            "Invoice No": r["invoice_no"],
            "Vendor": r["vendor"],
            "Date": r["date"].strftime("%Y-%m-%d"),
            "Amount": float(r["amount"]),
            "Account": r["account"],
            "Triggers": ", ".join(sorted(set(ex.triggers))),
            "Detail": " | ".join(ex.detail),
            "Risk Score": ex.risk_score,
            "Bucket": _bucket(ex.risk_score),
        })
    exc_df = pd.DataFrame(exc_rows)

    bucket_counts = exc_df["Bucket"].value_counts().to_dict() if not exc_df.empty else {}
    total_dollars_exposed = float(exc_df["Amount"].sum()) if not exc_df.empty else 0.0

    # Summary
    ws = wb.active
    ws.title = "Summary"
    ws.append(["AP Exception Review"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["Total Invoices", len(ledger)])
    ws.append(["Exceptions Flagged", len(exceptions)])
    ws.append(["Dollars Exposed", total_dollars_exposed])
    ws["B6"].number_format = ACCOUNTING_FMT
    ws.append([])
    ws.append(["Severity Bucket", "Count"])
    _style_header(ws, ws.max_row)
    for label in ("Critical", "High", "Medium", "Low"):
        ws.append([label, int(bucket_counts.get(label, 0))])
    # Color the buckets
    fills = {"Critical": RED_FILL, "High": ORANGE_FILL, "Medium": YELLOW_FILL, "Low": GREEN_FILL}
    for row in ws.iter_rows(min_row=ws.max_row - 3, max_row=ws.max_row):
        label = row[0].value
        if label in fills:
            row[0].fill = fills[label]
    ws.append([])
    if not exc_df.empty:
        ws.append(["Top 10 highest-risk invoices"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.append(["Invoice No", "Vendor", "Date", "Amount", "Risk Score", "Triggers"])
        _style_header(ws, ws.max_row)
        for _, r in exc_df.head(10).iterrows():
            ws.append([r["Invoice No"], r["Vendor"], r["Date"], r["Amount"], r["Risk Score"], r["Triggers"]])
        for cell in ws["D"][ws.max_row - 10:]:
            cell.number_format = ACCOUNTING_FMT
    _autosize(ws)

    # Exceptions
    ws_e = wb.create_sheet("Exceptions")
    if exc_df.empty:
        ws_e.append(["(no exceptions detected)"])
    else:
        ws_e.append(list(exc_df.columns))
        _style_header(ws_e)
        for _, r in exc_df.iterrows():
            ws_e.append([r[c] for c in exc_df.columns])
        for cell in ws_e["D"][1:]:
            cell.number_format = ACCOUNTING_FMT
        # color rows by bucket
        for row in ws_e.iter_rows(min_row=2):
            bucket = row[8].value
            if bucket in fills:
                for c in row:
                    c.fill = fills[bucket]
    _autosize(ws_e)
    ws_e.freeze_panes = "A2"

    # Other detail sheets
    for sheet_name, df in [
        ("Duplicates", pd.concat([result["exact_dupes"], result["near_dupes"]], ignore_index=True)
            if not result["exact_dupes"].empty or not result["near_dupes"].empty else pd.DataFrame()),
        ("Splits", result["splits"]),
        ("ThreeWayFail", result["three_way"]),
        ("CodingConsistency", result["coding"]),
        ("VendorIntegrity", result["integrity"]),
    ]:
        ws_x = wb.create_sheet(sheet_name)
        if df is None or df.empty:
            ws_x.append([f"(no findings for {sheet_name})"])
        else:
            ws_x.append(list(df.columns))
            _style_header(ws_x)
            for _, r in df.iterrows():
                ws_x.append([r[c] if not pd.isna(r[c]) else "" for c in df.columns])
        _autosize(ws_x)

    # Audit Trail
    ws_a = wb.create_sheet("AuditTrail")
    ws_a.append(["Step", "Details"])
    _style_header(ws_a)
    for a in result["audit"]:
        step = a.pop("step")
        ws_a.append([step, "; ".join(f"{k}={v}" for k, v in a.items())])
    ws_a.append([])
    ws_a.append(["Ledger Path", str(ledger_path)])
    for k, v in thresholds.items():
        ws_a.append([k, v])
    _autosize(ws_a)

    # SignOff
    ws_s = wb.create_sheet("SignOff")
    ws_s.append(["Role", "Name", "Title", "Date", "Notes"])
    _style_header(ws_s)
    ws_s.append(["Preparer (AP Manager)", "", "", "", ""])
    ws_s.append(["Reviewer (Controller)", "", "", "", ""])
    _autosize(ws_s)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    LOG.info("Exception report written: %s", output)


# --- CLI ------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--ledger", required=True, type=Path)
    parser.add_argument("--pos", type=Path, default=None)
    parser.add_argument("--vendors", type=Path, default=None)
    parser.add_argument("--duplicate-window-days", type=int, default=DUPLICATE_WINDOW_DAYS_DEFAULT)
    parser.add_argument("--amount-tolerance-pct", type=float, default=AMOUNT_TOLERANCE_PCT_DEFAULT)
    parser.add_argument("--split-threshold-usd", type=float, default=SPLIT_THRESHOLD_USD_DEFAULT)
    parser.add_argument("--approval-limit-usd", type=float, default=APPROVAL_LIMIT_USD_DEFAULT)
    parser.add_argument("--three-way-tolerance-pct", type=float, default=THREE_WAY_TOLERANCE_PCT_DEFAULT)
    parser.add_argument("--round-dollar-threshold", type=float, default=ROUND_DOLLAR_THRESHOLD_DEFAULT)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--quiet", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.WARNING if args.quiet else logging.INFO,
        format="%(levelname)s %(name)s: %(message)s",
    )

    try:
        ledger = _validate_ledger(_read_table(args.ledger))
    except ValueError as e:
        LOG.error("Ledger validation failed: %s", e)
        return 1
    pos = _read_table(args.pos) if args.pos else None
    vendors = _read_table(args.vendors) if args.vendors else None

    result = review(
        ledger=ledger, pos=pos, vendors=vendors,
        duplicate_window_days=args.duplicate_window_days,
        amount_tolerance_pct=args.amount_tolerance_pct,
        split_threshold=args.split_threshold_usd,
        approval_limit=args.approval_limit_usd,
        three_way_tol=args.three_way_tolerance_pct,
        round_dollar_threshold=args.round_dollar_threshold,
    )
    thresholds = vars(args).copy()
    thresholds.pop("output", None)
    write_report(result, args.output, args.ledger, {k: str(v) for k, v in thresholds.items()})

    n_critical = sum(1 for e in result["exceptions"].values() if _bucket(e.risk_score) == "Critical")
    n_high = sum(1 for e in result["exceptions"].values() if _bucket(e.risk_score) == "High")
    print(f"Invoices reviewed:    {len(ledger)}")
    print(f"Exceptions flagged:   {len(result['exceptions'])}")
    print(f"  - Critical:         {n_critical}")
    print(f"  - High:             {n_high}")
    print(f"Report:               {args.output}")
    if n_critical or n_high:
        return 2
    return 0


if __name__ == "__main__":
    sys.exit(main())
