"""Expense Policy Reviewer — rule-based T&E audit.

Implements the workflow in ../SKILL.md.
"""
from __future__ import annotations

import argparse
import json
import logging
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

LOG = logging.getLogger("expense_policy_reviewer")

ACCOUNTING_FMT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN = Side(border_style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SEVERITY_WEIGHT = {"FAIL": 10, "WARN": 3}
BUCKETS = [(30, "Critical"), (15, "High"), (5, "Medium"), (0, "Low")]

REQUIRED_COLUMNS = (
    "report_id", "employee_id", "employee_name",
    "expense_date", "category", "amount",
    "vendor", "description", "has_receipt",
)


@dataclass
class Violation:
    report_id: str
    employee_id: str
    employee_name: str
    expense_date: datetime
    category: str
    amount: float
    description: str
    rule: str
    severity: str  # FAIL / WARN
    detail: str


# --- I/O ------------------------------------------------------------------

def _read_table(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path)
    else:
        df = pd.read_excel(path)
    df.columns = [c.strip().lower() for c in df.columns]
    return df


def _read_policy(path: Path) -> dict:
    text = path.read_text()
    if path.suffix.lower() in {".yaml", ".yml"}:
        return yaml.safe_load(text) or {}
    return json.loads(text)


def _validate_expenses(df: pd.DataFrame) -> pd.DataFrame:
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"Expenses missing required columns: {missing}")
    df = df.copy()
    df["expense_date"] = pd.to_datetime(df["expense_date"], errors="coerce")
    df["amount"] = pd.to_numeric(df["amount"], errors="coerce")
    df = df.dropna(subset=["expense_date", "amount"])
    df["has_receipt"] = df["has_receipt"].fillna(False)
    if df["has_receipt"].dtype != bool:
        df["has_receipt"] = df["has_receipt"].astype(str).str.lower().isin({"true", "yes", "1"})
    df["category"] = df["category"].fillna("Other").astype(str).str.strip()
    df["description"] = df["description"].fillna("").astype(str)
    for opt in ("attendees", "currency", "policy_exception_approval"):
        if opt not in df.columns:
            df[opt] = ""
        else:
            df[opt] = df[opt].fillna("").astype(str)
    df["currency"] = df["currency"].replace("", "USD")
    df = df.reset_index(drop=True)
    return df


def _attendee_count(s: str) -> int:
    if not s:
        return 0
    return len([x for x in s.split(";") if x.strip()])


# --- Rule engine ----------------------------------------------------------

def evaluate(expenses: pd.DataFrame, policy: dict) -> list[Violation]:
    violations: list[Violation] = []
    receipt_threshold = float(policy.get("receipt_required_above", 25.0))
    prohibited = set(c.lower() for c in policy.get("prohibited_categories", []))
    requires_attendees = set(c.lower() for c in policy.get("require_attendees_for", []))
    min_purpose = int(policy.get("require_business_purpose_min_chars", 0))
    weekend_rule = str(policy.get("weekend_review", "off")).lower()
    limits = policy.get("limits", {}) or {}

    for _, r in expenses.iterrows():
        cat = r["category"]
        cat_l = cat.lower()
        amt = float(r["amount"])
        attendees = _attendee_count(r["attendees"])
        date_obj = r["expense_date"].to_pydatetime() if hasattr(r["expense_date"], "to_pydatetime") else r["expense_date"]
        is_weekend = date_obj.weekday() >= 5
        exception_approved = bool(str(r.get("policy_exception_approval", "")).strip())

        def add(rule: str, severity: str, detail: str) -> None:
            sev = severity
            if exception_approved and sev == "FAIL":
                sev = "WARN"
                detail += " (downgraded to WARN — manager exception approved)"
            violations.append(Violation(
                report_id=str(r["report_id"]), employee_id=str(r["employee_id"]),
                employee_name=str(r["employee_name"]),
                expense_date=date_obj, category=cat, amount=amt,
                description=r["description"], rule=rule, severity=sev, detail=detail,
            ))

        # Receipt requirement
        if amt > receipt_threshold and not bool(r["has_receipt"]):
            add("Missing Receipt", "FAIL",
                f"Amount ${amt:.2f} > receipt threshold ${receipt_threshold:.2f} and no receipt")
        # Prohibited category
        if cat_l in prohibited:
            add("Prohibited Category", "FAIL", f"Category {cat!r} is prohibited by policy")
        # Attendees requirement
        if cat_l in requires_attendees and attendees == 0:
            add("Missing Attendees", "WARN", f"{cat} requires named attendees")
        # Business purpose
        if min_purpose and len(r["description"].strip()) < min_purpose:
            add("Vague Business Purpose", "WARN",
                f"Description length {len(r['description'].strip())} < {min_purpose} chars")
        # Weekend
        if is_weekend and weekend_rule in {"warn", "fail"}:
            add("Weekend Transaction", "FAIL" if weekend_rule == "fail" else "WARN",
                f"Transaction dated {date_obj.strftime('%A %Y-%m-%d')}")
        # Category limits
        limit_def = None
        for k in limits:
            if k.lower() == cat_l:
                limit_def = limits[k]
                break
        if isinstance(limit_def, dict):
            if "daily_per_attendee" in limit_def:
                per_attendee = amt / max(attendees, 1)
                if per_attendee > float(limit_def["daily_per_attendee"]):
                    add("Limit Exceeded", "FAIL",
                        f"${per_attendee:.2f}/attendee > ${limit_def['daily_per_attendee']:.2f} cap")
                elif "warn_above" in limit_def and per_attendee > float(limit_def["warn_above"]):
                    add("Limit Warning", "WARN",
                        f"${per_attendee:.2f}/attendee > ${limit_def['warn_above']:.2f} warn-threshold")
            if "per_night" in limit_def and amt > float(limit_def["per_night"]):
                add("Limit Exceeded", "FAIL",
                    f"${amt:.2f} > ${limit_def['per_night']:.2f}/night lodging cap")
            if "per_event_per_attendee" in limit_def:
                per_attendee = amt / max(attendees, 1)
                if per_attendee > float(limit_def["per_event_per_attendee"]):
                    add("Limit Exceeded", "FAIL",
                        f"${per_attendee:.2f}/attendee > ${limit_def['per_event_per_attendee']:.2f} entertainment cap")
            if "round_trip_domestic" in limit_def and amt > float(limit_def["round_trip_domestic"]) \
                    and "international" not in r["description"].lower():
                add("Limit Warning", "WARN",
                    f"${amt:.2f} > ${limit_def['round_trip_domestic']:.2f} domestic airfare threshold")
            if "international" in limit_def and amt > float(limit_def["international"]):
                add("Limit Warning", "WARN",
                    f"${amt:.2f} > ${limit_def['international']:.2f} international airfare threshold")
            if "per_day" in limit_def and amt > float(limit_def["per_day"]):
                add("Limit Exceeded", "FAIL",
                    f"${amt:.2f} > ${limit_def['per_day']:.2f}/day cap")
    return violations


# --- Duplicate detection ---------------------------------------------------

def detect_duplicates(expenses: pd.DataFrame, window_days: int, cross_employee: bool) -> list[dict]:
    dupes: list[dict] = []
    # Within-report
    for report_id, sub in expenses.groupby("report_id"):
        seen: dict[tuple, int] = {}
        for idx, r in sub.iterrows():
            key = (r["expense_date"].date(), r["category"], round(float(r["amount"]), 2))
            if key in seen:
                dupes.append({
                    "Type": "Within Report", "Severity": "FAIL",
                    "Report A": report_id, "Report B": report_id,
                    "Employee A": r["employee_name"], "Employee B": r["employee_name"],
                    "Date": r["expense_date"].strftime("%Y-%m-%d"),
                    "Category": r["category"], "Amount": float(r["amount"]),
                    "Vendor": r["vendor"],
                })
            else:
                seen[key] = idx
    # Across reports same employee within window
    for emp, sub in expenses.groupby("employee_id"):
        rows = sub.sort_values("expense_date").to_dict("records")
        for i in range(len(rows)):
            for j in range(i + 1, len(rows)):
                a, b = rows[i], rows[j]
                if a["report_id"] == b["report_id"]:
                    continue
                if a["category"] != b["category"] or round(a["amount"], 2) != round(b["amount"], 2):
                    continue
                diff = abs((a["expense_date"] - b["expense_date"]).days)
                if diff > window_days:
                    break
                dupes.append({
                    "Type": "Across Reports (Same Employee)", "Severity": "WARN",
                    "Report A": a["report_id"], "Report B": b["report_id"],
                    "Employee A": a["employee_name"], "Employee B": b["employee_name"],
                    "Date": a["expense_date"].strftime("%Y-%m-%d") + " / " + b["expense_date"].strftime("%Y-%m-%d"),
                    "Category": a["category"], "Amount": float(a["amount"]),
                    "Vendor": a["vendor"],
                })
    # Across employees (likely same receipt submitted twice)
    if cross_employee:
        key_map: dict[tuple, list[int]] = defaultdict(list)
        for idx, r in expenses.iterrows():
            k = (r["expense_date"].date(), r["category"], round(float(r["amount"]), 2),
                 str(r["vendor"]).lower().strip())
            key_map[k].append(idx)
        for k, idxs in key_map.items():
            if len(idxs) < 2:
                continue
            emps = set(expenses.loc[idxs, "employee_id"])
            if len(emps) > 1:
                for i in range(len(idxs)):
                    for j in range(i + 1, len(idxs)):
                        a, b = expenses.loc[idxs[i]], expenses.loc[idxs[j]]
                        if a["employee_id"] == b["employee_id"]:
                            continue
                        dupes.append({
                            "Type": "Across Employees", "Severity": "FAIL",
                            "Report A": a["report_id"], "Report B": b["report_id"],
                            "Employee A": a["employee_name"], "Employee B": b["employee_name"],
                            "Date": k[0].isoformat(), "Category": k[1],
                            "Amount": float(a["amount"]), "Vendor": a["vendor"],
                        })
    return dupes


def _bucket(score: int) -> str:
    for t, label in BUCKETS:
        if score >= t:
            return label
    return "Low"


# --- Writer ---------------------------------------------------------------

def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX


def _autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(60, max_len + 2)


def write_report(
    expenses: pd.DataFrame, violations: list[Violation], duplicates: list[dict],
    policy_path: Path, output: Path,
) -> None:
    wb = Workbook()
    # Build per-report score
    report_scores: dict[str, int] = defaultdict(int)
    report_line_counts: dict[str, int] = defaultdict(int)
    report_flagged_lines: dict[str, set] = defaultdict(set)
    employee_scores: dict[str, int] = defaultdict(int)
    for v in violations:
        report_scores[v.report_id] += SEVERITY_WEIGHT[v.severity]
        employee_scores[f"{v.employee_id}|{v.employee_name}"] += SEVERITY_WEIGHT[v.severity]
        report_flagged_lines[v.report_id].add((v.expense_date, v.category, v.amount, v.description))
    for d in duplicates:
        report_scores[d["Report A"]] += SEVERITY_WEIGHT[d["Severity"]]
        report_scores[d["Report B"]] += SEVERITY_WEIGHT[d["Severity"]] // 2
    for rid, sub in expenses.groupby("report_id"):
        report_line_counts[rid] = len(sub)

    # Summary
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Expense Policy Review"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    n_reports = expenses["report_id"].nunique()
    n_lines = len(expenses)
    total_dollars = float(expenses["amount"].sum())
    n_fail = sum(1 for v in violations if v.severity == "FAIL")
    n_warn = sum(1 for v in violations if v.severity == "WARN")
    flagged_dollars = sum(v.amount for v in violations)
    ws.append(["Reports Reviewed", n_reports])
    ws.append(["Expense Lines", n_lines])
    ws.append(["Total Dollars", total_dollars])
    ws["B5"].number_format = ACCOUNTING_FMT
    ws.append(["Violations (FAIL)", n_fail])
    ws.append(["Violations (WARN)", n_warn])
    ws.append(["Duplicates Detected", len(duplicates)])
    ws.append(["Flagged Dollars", flagged_dollars])
    ws["B9"].number_format = ACCOUNTING_FMT
    _autosize(ws)

    # Reports
    ws_r = wb.create_sheet("Reports")
    ws_r.append(["Report ID", "Employee(s)", "Lines", "Flagged Lines", "Risk Score", "Bucket"])
    _style_header(ws_r)
    fills = {"Critical": RED_FILL, "High": ORANGE_FILL, "Medium": YELLOW_FILL, "Low": GREEN_FILL}
    for report_id, score in sorted(report_scores.items(), key=lambda kv: -kv[1]):
        sub = expenses[expenses["report_id"] == report_id]
        employees = "; ".join(sub["employee_name"].unique())
        bucket = _bucket(score)
        ws_r.append([report_id, employees, len(sub), len(report_flagged_lines.get(report_id, set())), score, bucket])
    # also include reports with no violations
    no_viol = set(expenses["report_id"]) - set(report_scores)
    for report_id in no_viol:
        sub = expenses[expenses["report_id"] == report_id]
        employees = "; ".join(sub["employee_name"].unique())
        ws_r.append([report_id, employees, len(sub), 0, 0, "Low"])
    for row in ws_r.iter_rows(min_row=2):
        if row[5].value in fills:
            row[5].fill = fills[row[5].value]
    _autosize(ws_r)

    # Exception Lines
    ws_e = wb.create_sheet("ExceptionLines")
    ws_e.append(["Report", "Employee", "Date", "Category", "Amount", "Rule", "Severity", "Detail", "Description"])
    _style_header(ws_e)
    for v in sorted(violations, key=lambda x: (x.report_id, -SEVERITY_WEIGHT[x.severity])):
        ws_e.append([
            v.report_id, v.employee_name, v.expense_date.strftime("%Y-%m-%d"),
            v.category, v.amount, v.rule, v.severity, v.detail, v.description,
        ])
    for cell in ws_e["E"][1:]:
        cell.number_format = ACCOUNTING_FMT
    for row in ws_e.iter_rows(min_row=2):
        if row[6].value == "FAIL":
            row[6].fill = RED_FILL
        elif row[6].value == "WARN":
            row[6].fill = YELLOW_FILL
    _autosize(ws_e)
    ws_e.freeze_panes = "A2"

    # Duplicates
    ws_d = wb.create_sheet("Duplicates")
    if duplicates:
        ws_d.append(list(duplicates[0].keys()))
        _style_header(ws_d)
        for d in duplicates:
            ws_d.append(list(d.values()))
        for cell in ws_d["I"][1:]:
            cell.number_format = ACCOUNTING_FMT
    else:
        ws_d.append(["(no duplicates detected)"])
    _autosize(ws_d)

    # Employee Risk
    ws_er = wb.create_sheet("EmployeeRisk")
    ws_er.append(["Employee ID", "Employee Name", "Aggregate Risk Score", "Bucket"])
    _style_header(ws_er)
    for k, score in sorted(employee_scores.items(), key=lambda kv: -kv[1]):
        emp_id, emp_name = k.split("|", 1)
        ws_er.append([emp_id, emp_name, score, _bucket(score)])
    for row in ws_er.iter_rows(min_row=2):
        if row[3].value in fills:
            row[3].fill = fills[row[3].value]
    _autosize(ws_er)

    # Audit Trail
    ws_a = wb.create_sheet("AuditTrail")
    ws_a.append(["Key", "Value"])
    _style_header(ws_a)
    ws_a.append(["Policy File", str(policy_path)])
    ws_a.append(["Reports", n_reports])
    ws_a.append(["Lines", n_lines])
    ws_a.append(["Generated", datetime.now().isoformat(timespec="seconds")])
    _autosize(ws_a)

    # SignOff
    ws_s = wb.create_sheet("SignOff")
    ws_s.append(["Role", "Name", "Title", "Date", "Notes"])
    _style_header(ws_s)
    ws_s.append(["Preparer (AP)", "", "", "", ""])
    ws_s.append(["Reviewer (Internal Audit)", "", "", "", ""])
    _autosize(ws_s)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    LOG.info("Report written: %s", output)


# --- CLI ------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--expenses", required=True, type=Path)
    parser.add_argument("--policy", required=True, type=Path)
    parser.add_argument("--period-end", required=True)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--no-cross-employee-duplicates", action="store_true")
    parser.add_argument("--quiet", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.WARNING if args.quiet else logging.INFO,
        format="%(levelname)s %(name)s: %(message)s",
    )

    try:
        expenses = _validate_expenses(_read_table(args.expenses))
        policy = _read_policy(args.policy)
    except (ValueError, FileNotFoundError) as e:
        LOG.error("Input validation failed: %s", e)
        return 1

    violations = evaluate(expenses, policy)
    duplicates = detect_duplicates(
        expenses, int(policy.get("duplicate_window_days", 14)),
        cross_employee=not args.no_cross_employee_duplicates,
    )
    write_report(expenses, violations, duplicates, args.policy, args.output)

    n_fail = sum(1 for v in violations if v.severity == "FAIL")
    n_warn = sum(1 for v in violations if v.severity == "WARN")
    print(f"Reports:        {expenses['report_id'].nunique()}")
    print(f"Lines:          {len(expenses)}")
    print(f"FAIL:           {n_fail}")
    print(f"WARN:           {n_warn}")
    print(f"Duplicates:     {len(duplicates)}")
    print(f"Output:         {args.output}")
    return 0 if n_fail == 0 and not duplicates else 2


if __name__ == "__main__":
    sys.exit(main())
