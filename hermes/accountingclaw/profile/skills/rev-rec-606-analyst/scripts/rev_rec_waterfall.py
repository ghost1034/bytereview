"""ASC 606 Revenue Recognition Analyst — 5-step engine + monthly waterfall.

Implements the workflow in ../SKILL.md.
"""
from __future__ import annotations

import argparse
import logging
import sys
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

LOG = logging.getLogger("rev_rec_606")

ACCOUNTING_FMT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN = Side(border_style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

REQUIRED = (
    "contract_id", "po_id", "po_description", "customer",
    "contract_start_date", "service_start_date", "service_end_date",
    "transaction_price", "ssp", "recognition_pattern", "is_distinct",
)


@dataclass
class PO:
    contract_id: str
    po_id: str
    description: str
    customer: str
    contract_start: datetime
    service_start: datetime
    service_end: datetime
    transaction_price: float
    ssp: float
    pattern: str
    is_distinct: bool
    allocated_price: float = 0.0
    recognized_to_date: float = 0.0
    period_recognized: float = 0.0
    modification_date: datetime | None = None
    modification_type: str = ""
    notes: list[str] = field(default_factory=list)


# --- I/O ------------------------------------------------------------------

def _read_table(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path)
    else:
        df = pd.read_excel(path)
    df.columns = [c.strip().lower() for c in df.columns]
    return df


def _validate_contracts(df: pd.DataFrame) -> pd.DataFrame:
    missing = [c for c in REQUIRED if c not in df.columns]
    if missing:
        raise ValueError(f"Contracts missing required columns: {missing}")
    df = df.copy()
    for c in ("contract_start_date", "service_start_date", "service_end_date"):
        df[c] = pd.to_datetime(df[c], errors="coerce")
    if "modification_effective_date" in df.columns:
        df["modification_effective_date"] = pd.to_datetime(df["modification_effective_date"], errors="coerce")
    for c in ("transaction_price", "ssp"):
        df[c] = pd.to_numeric(df[c], errors="coerce")
    df["is_distinct"] = df["is_distinct"].astype(bool)
    df["recognition_pattern"] = df["recognition_pattern"].fillna("ratable").astype(str).str.lower()
    bad = df[df["transaction_price"].isna() | df["service_start_date"].isna() | df["service_end_date"].isna()]
    if not bad.empty:
        LOG.warning("Dropping %d rows with missing TP / dates", len(bad))
        df = df.dropna(subset=["transaction_price", "service_start_date", "service_end_date"])
    return df.reset_index(drop=True)


# --- Allocation (Step 4) --------------------------------------------------

def allocate_transaction_price(rows: list[PO], ssp_method: str, market_discount: float) -> None:
    """Allocate the contract's transaction price across POs in proportion to SSPs."""
    # Group by contract_id
    groups: dict[str, list[PO]] = {}
    for r in rows:
        groups.setdefault(r.contract_id, []).append(r)
    for contract_id, pos in groups.items():
        contract_tp = sum(p.transaction_price for p in pos)
        # If SSP missing, apply --ssp-method
        for p in pos:
            if pd.isna(p.ssp) or p.ssp <= 0:
                if ssp_method == "adjusted_market":
                    p.ssp = p.transaction_price / max(1 - market_discount, 0.01)
                    p.notes.append(f"SSP imputed via adjusted_market (discount={market_discount})")
                elif ssp_method == "expected_cost_plus_margin":
                    p.ssp = p.transaction_price  # no cost data here; user supplies
                    p.notes.append("SSP imputed = TP via cost+margin fallback")
                elif ssp_method == "residual":
                    observed = sum(o.ssp for o in pos if not pd.isna(o.ssp))
                    p.ssp = max(contract_tp - observed, 0.0)
                    p.notes.append("SSP imputed via residual method")
                else:
                    p.ssp = p.transaction_price
                    p.notes.append("SSP missing; defaulted to TP")
        ssp_sum = sum(p.ssp for p in pos)
        if ssp_sum <= 0:
            for p in pos:
                p.allocated_price = p.transaction_price
                p.notes.append("Allocation skipped: SSP sum is zero")
            continue
        for p in pos:
            p.allocated_price = round(contract_tp * (p.ssp / ssp_sum), 2)
            divergence = abs(p.allocated_price - p.transaction_price) / max(p.transaction_price, 0.01)
            if divergence > 0.01:
                p.notes.append(f"Allocation diverges from line TP by {divergence*100:.1f}%")


# --- Recognition (Step 5) -------------------------------------------------

def recognize(po: PO, period_end: datetime) -> tuple[float, float]:
    """Return (recognized_to_date, recognized_in_current_month)."""
    if period_end < po.service_start:
        return 0.0, 0.0
    if po.pattern == "point_in_time":
        recognized = po.allocated_price if period_end >= po.service_start else 0.0
        current_month = po.allocated_price if (
            po.service_start.year == period_end.year and po.service_start.month == period_end.month
        ) else 0.0
        return recognized, current_month
    if po.pattern == "ratable":
        total_days = (po.service_end - po.service_start).days + 1
        if total_days <= 0:
            return 0.0, 0.0
        daily = po.allocated_price / total_days
        clamp_end = min(period_end, po.service_end)
        days_elapsed = (clamp_end - po.service_start).days + 1
        days_elapsed = max(0, min(days_elapsed, total_days))
        recognized = round(daily * days_elapsed, 2)
        # current month
        month_start = period_end.replace(day=1)
        month_end = period_end
        cm_start = max(po.service_start, month_start)
        cm_end = min(po.service_end, month_end)
        if cm_end < cm_start:
            current_month = 0.0
        else:
            cm_days = (cm_end - cm_start).days + 1
            current_month = round(daily * cm_days, 2)
        return recognized, current_month
    if po.pattern == "usage_based":
        po.notes.append("usage_based pattern requires --usage CSV; defaulting to 0 in v1")
        return 0.0, 0.0
    po.notes.append(f"Unknown pattern {po.pattern!r}; defaulting to 0")
    return 0.0, 0.0


def apply_modifications(po: PO, period_end: datetime, original_allocated: float) -> tuple[float, float]:
    """Return (catch_up_adjustment, post_mod_recognized_to_date) — for catch_up modifications."""
    if not po.modification_date or po.modification_date > period_end:
        return 0.0, 0.0
    if po.modification_type.lower() != "catch_up":
        return 0.0, 0.0
    # Cumulative catch-up: recompute as if revised TP applied from inception
    total_days = (po.service_end - po.service_start).days + 1
    clamp_end = min(period_end, po.service_end)
    days_elapsed = max(0, min((clamp_end - po.service_start).days + 1, total_days))
    revised_rec_to_date = round(po.allocated_price * (days_elapsed / max(total_days, 1)), 2)
    original_rec_to_date = round(original_allocated * (days_elapsed / max(total_days, 1)), 2)
    catch_up = revised_rec_to_date - original_rec_to_date
    return catch_up, revised_rec_to_date


# --- Monthly waterfall ----------------------------------------------------

def build_waterfall(pos: list[PO]) -> pd.DataFrame:
    rows: list[dict] = []
    if not pos:
        return pd.DataFrame()
    min_start = min(p.service_start for p in pos)
    max_end = max(p.service_end for p in pos)
    months: list[datetime] = []
    cur = min_start.replace(day=1)
    while cur <= max_end:
        months.append(cur)
        cur = cur + relativedelta(months=1)
    for p in pos:
        per_month: list[float] = []
        for m in months:
            month_last = (m + relativedelta(months=1)) - timedelta(days=1)
            if p.pattern == "point_in_time":
                if p.service_start.year == m.year and p.service_start.month == m.month:
                    per_month.append(p.allocated_price)
                else:
                    per_month.append(0.0)
            elif p.pattern == "ratable":
                total_days = (p.service_end - p.service_start).days + 1
                if total_days <= 0:
                    per_month.append(0.0)
                    continue
                daily = p.allocated_price / total_days
                cm_start = max(p.service_start, m)
                cm_end = min(p.service_end, month_last)
                if cm_end < cm_start:
                    per_month.append(0.0)
                else:
                    days = (cm_end - cm_start).days + 1
                    per_month.append(round(daily * days, 2))
            else:
                per_month.append(0.0)
        row = {
            "Contract": p.contract_id, "PO": p.po_id, "Description": p.description,
            "Pattern": p.pattern, "Allocated TP": p.allocated_price,
        }
        for m, v in zip(months, per_month):
            row[m.strftime("%Y-%m")] = v
        rows.append(row)
    df = pd.DataFrame(rows)
    return df


# --- Writer --------------------------------------------------------------

def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX


def _autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(50, max_len + 2)


def write_workpaper(pos: list[PO], waterfall: pd.DataFrame, allocation_df: pd.DataFrame,
                    deferred_df: pd.DataFrame, period_end: datetime, output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["ASC 606 Revenue Recognition Workpaper"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Period End", period_end.strftime("%Y-%m-%d")])
    ws.append(["Contracts", len({p.contract_id for p in pos})])
    ws.append(["Performance Obligations", len(pos)])
    total_tp = sum(p.allocated_price for p in pos)
    total_rec = sum(p.recognized_to_date for p in pos)
    total_period = sum(p.period_recognized for p in pos)
    ws.append(["Total Allocated TP", total_tp])
    ws.append(["Recognized To-Date", total_rec])
    ws.append(["Period Revenue", total_period])
    ws.append(["Deferred (= TP - Rec)", round(total_tp - total_rec, 2)])
    for cell in ("B6", "B7", "B8", "B9"):
        ws[cell].number_format = ACCOUNTING_FMT
    _autosize(ws)

    # FiveStepWorkpaper
    ws_w = wb.create_sheet("FiveStepWorkpaper")
    ws_w.append([
        "Contract", "PO", "Description", "Customer", "Distinct?",
        "Service Start", "Service End", "TP (Line)", "SSP", "Allocated Price",
        "Pattern", "Recognized To-Date", "Period Revenue", "Notes",
    ])
    _style_header(ws_w)
    for p in pos:
        ws_w.append([
            p.contract_id, p.po_id, p.description, p.customer, "Yes" if p.is_distinct else "No",
            p.service_start.strftime("%Y-%m-%d"), p.service_end.strftime("%Y-%m-%d"),
            p.transaction_price, p.ssp, p.allocated_price, p.pattern,
            p.recognized_to_date, p.period_recognized, "; ".join(p.notes),
        ])
    for letter in ("H", "I", "J", "L", "M"):
        for cell in ws_w[letter][1:]:
            cell.number_format = ACCOUNTING_FMT
    _autosize(ws_w)
    ws_w.freeze_panes = "A2"

    # Allocation
    ws_a = wb.create_sheet("Allocation")
    if not allocation_df.empty:
        ws_a.append(list(allocation_df.columns))
        _style_header(ws_a)
        for _, r in allocation_df.iterrows():
            ws_a.append([r[c] for c in allocation_df.columns])
        for col_idx, col in enumerate(allocation_df.columns, start=1):
            if "Price" in col or "TP" in col or "SSP" in col:
                for cell in ws_a[get_column_letter(col_idx)][1:]:
                    cell.number_format = ACCOUNTING_FMT
    _autosize(ws_a)

    # Waterfall
    ws_wf = wb.create_sheet("Waterfall")
    if not waterfall.empty:
        ws_wf.append(list(waterfall.columns))
        _style_header(ws_wf)
        for _, r in waterfall.iterrows():
            ws_wf.append([r[c] for c in waterfall.columns])
        for col_idx, col in enumerate(waterfall.columns, start=1):
            if "-" in col or "Allocated" in col:
                for cell in ws_wf[get_column_letter(col_idx)][1:]:
                    cell.number_format = ACCOUNTING_FMT
    _autosize(ws_wf)
    ws_wf.freeze_panes = "F2"

    # Deferred RollForward
    ws_d = wb.create_sheet("DeferredRollForward")
    if not deferred_df.empty:
        ws_d.append(list(deferred_df.columns))
        _style_header(ws_d)
        for _, r in deferred_df.iterrows():
            ws_d.append([r[c] for c in deferred_df.columns])
        for col_idx, col in enumerate(deferred_df.columns, start=1):
            if any(k in col for k in ("Beg", "Billings", "Recognized", "End")):
                for cell in ws_d[get_column_letter(col_idx)][1:]:
                    cell.number_format = ACCOUNTING_FMT
    else:
        ws_d.append(["(no billings file supplied)"])
    _autosize(ws_d)

    # Modifications
    ws_m = wb.create_sheet("Modifications")
    ws_m.append(["Contract", "PO", "Mod Date", "Mod Type", "Catch-up Amount", "Notes"])
    _style_header(ws_m)
    any_mod = False
    for p in pos:
        if p.modification_date:
            any_mod = True
            ws_m.append([p.contract_id, p.po_id, p.modification_date.strftime("%Y-%m-%d"),
                         p.modification_type,
                         next((float(n.split("=")[-1]) for n in p.notes if n.startswith("catch_up=")), 0.0),
                         "; ".join(p.notes)])
    if not any_mod:
        ws_m.append(["(no modifications in period)"])
    _autosize(ws_m)

    # ContractCosts (placeholder - based on commission_paid if present)
    ws_c = wb.create_sheet("ContractCosts")
    ws_c.append(["Contract", "Commission Paid", "Amortization Life (months)", "Note"])
    _style_header(ws_c)
    ws_c.append(["(populate via commission_paid column on contracts.csv)"])
    _autosize(ws_c)

    # AuditTrail
    ws_at = wb.create_sheet("AuditTrail")
    ws_at.append(["Key", "Value"])
    _style_header(ws_at)
    ws_at.append(["Period End", period_end.strftime("%Y-%m-%d")])
    ws_at.append(["POs Processed", len(pos)])
    ws_at.append(["Contracts", len({p.contract_id for p in pos})])
    ws_at.append(["POs with Notes", sum(1 for p in pos if p.notes)])
    ws_at.append(["Generated", datetime.now().isoformat(timespec="seconds")])
    _autosize(ws_at)

    # SignOff
    ws_s = wb.create_sheet("SignOff")
    ws_s.append(["Role", "Name", "Title", "Date", "Notes"])
    _style_header(ws_s)
    ws_s.append(["Preparer (Revenue)", "", "", "", ""])
    ws_s.append(["Reviewer (Tech Accounting)", "", "", "", ""])
    ws_s.append(["Approver (Controller)", "", "", "", ""])
    _autosize(ws_s)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    LOG.info("Workpaper written: %s", output)


# --- Orchestration -------------------------------------------------------

def build_pos(df: pd.DataFrame) -> list[PO]:
    pos: list[PO] = []
    for _, r in df.iterrows():
        pos.append(PO(
            contract_id=str(r["contract_id"]), po_id=str(r["po_id"]),
            description=str(r["po_description"]), customer=str(r["customer"]),
            contract_start=r["contract_start_date"], service_start=r["service_start_date"],
            service_end=r["service_end_date"], transaction_price=float(r["transaction_price"]),
            ssp=float(r["ssp"]) if pd.notna(r["ssp"]) else float("nan"),
            pattern=str(r["recognition_pattern"]), is_distinct=bool(r["is_distinct"]),
            modification_date=r.get("modification_effective_date"),
            modification_type=str(r.get("modification_type", "")),
        ))
    return pos


def build_allocation_df(pos: list[PO]) -> pd.DataFrame:
    rows: list[dict] = []
    groups: dict[str, list[PO]] = {}
    for p in pos:
        groups.setdefault(p.contract_id, []).append(p)
    for cid, ps in groups.items():
        ssp_sum = sum(p.ssp for p in ps if not pd.isna(p.ssp))
        tp_sum = sum(p.transaction_price for p in ps)
        for p in ps:
            factor = (p.ssp / ssp_sum) if ssp_sum else 0
            rows.append({
                "Contract": cid, "PO": p.po_id,
                "SSP": p.ssp, "SSP Sum (Contract)": round(ssp_sum, 2),
                "Allocation Factor": round(factor, 4),
                "Contract TP": round(tp_sum, 2),
                "Allocated Price": p.allocated_price,
            })
    return pd.DataFrame(rows)


def build_deferred(pos: list[PO], billings: pd.DataFrame | None, period_end: datetime) -> pd.DataFrame:
    if billings is None or billings.empty:
        return pd.DataFrame()
    billings = billings.copy()
    billings["billing_date"] = pd.to_datetime(billings["billing_date"], errors="coerce")
    billings["amount"] = pd.to_numeric(billings["amount"], errors="coerce")
    rows: list[dict] = []
    contract_pos: dict[str, list[PO]] = {}
    for p in pos:
        contract_pos.setdefault(p.contract_id, []).append(p)
    for cid, ps in contract_pos.items():
        cust = ps[0].customer
        billed_to_date = float(billings[(billings["contract_id"] == cid) & (billings["billing_date"] <= period_end)]["amount"].sum())
        recognized = sum(p.recognized_to_date for p in ps)
        deferred = max(billed_to_date - recognized, 0.0)
        rows.append({
            "Contract": cid, "Customer": cust,
            "Billings To-Date": round(billed_to_date, 2),
            "Recognized To-Date": round(recognized, 2),
            "Deferred Revenue": round(deferred, 2),
        })
    return pd.DataFrame(rows)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--contracts", required=True, type=Path)
    parser.add_argument("--period-end", required=True)
    parser.add_argument("--ssp-method", default="adjusted_market",
                        choices=["observable", "adjusted_market", "expected_cost_plus_margin", "residual"])
    parser.add_argument("--market-discount", type=float, default=0.20)
    parser.add_argument("--billings", type=Path, default=None)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--quiet", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.WARNING if args.quiet else logging.INFO,
        format="%(levelname)s %(name)s: %(message)s",
    )

    try:
        df = _validate_contracts(_read_table(args.contracts))
    except ValueError as e:
        LOG.error("Validation failed: %s", e)
        return 1

    period_end = datetime.strptime(args.period_end, "%Y-%m-%d")
    pos = build_pos(df)
    allocate_transaction_price(pos, args.ssp_method, args.market_discount)

    for p in pos:
        rec_td, period_rec = recognize(p, period_end)
        p.recognized_to_date = rec_td
        p.period_recognized = period_rec
        # Apply catch-up modifications
        if p.modification_date and p.modification_date <= period_end and p.modification_type.lower() == "catch_up":
            catch_up, revised_rec = apply_modifications(p, period_end, p.transaction_price)
            p.recognized_to_date = revised_rec
            p.notes.append(f"catch_up={round(catch_up, 2)}")

    waterfall = build_waterfall(pos)
    allocation_df = build_allocation_df(pos)
    billings = _read_table(args.billings) if args.billings else None
    deferred_df = build_deferred(pos, billings, period_end)

    write_workpaper(pos, waterfall, allocation_df, deferred_df, period_end, args.output)
    total_tp = sum(p.allocated_price for p in pos)
    total_rec = sum(p.recognized_to_date for p in pos)
    period_rev = sum(p.period_recognized for p in pos)
    print(f"Contracts:       {len({p.contract_id for p in pos})}")
    print(f"POs:             {len(pos)}")
    print(f"Total TP:        {total_tp:>15,.2f}")
    print(f"Recognized:      {total_rec:>15,.2f}")
    print(f"Period Revenue:  {period_rev:>15,.2f}")
    print(f"Workpaper:       {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
