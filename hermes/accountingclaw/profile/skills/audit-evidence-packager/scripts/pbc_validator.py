"""Audit Evidence Packager — match PBC requests to evidence files; produce readiness report.

Implements the workflow in ../SKILL.md.
"""
from __future__ import annotations

import argparse
import csv
import fnmatch
import logging
import shutil
import sys
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz, process

LOG = logging.getLogger("audit_evidence_packager")

ACCOUNTING_FMT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN = Side(border_style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

EMPTY_FILE_THRESHOLD_BYTES = 1024


@dataclass
class MatchResult:
    request_id: str
    matched_path: Path | None = None
    similarity: int = 0
    candidates: list[tuple[Path, int]] = None
    status: str = "Not Found"  # Provided / Provided with Issues / Format Mismatch / Not Found
    warnings: list[str] = None


def _read_table(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path)
    else:
        df = pd.read_excel(path)
    df.columns = [c.strip().lower() for c in df.columns]
    return df


def _walk_evidence(directory: Path) -> list[Path]:
    if not directory.exists():
        raise FileNotFoundError(f"Evidence dir not found: {directory}")
    return [p for p in directory.rglob("*") if p.is_file()]


def _match_one(req: dict, files: list[Path], similarity_threshold: int) -> MatchResult:
    expected = str(req.get("expected_filename", "")).strip()
    expected_ext = str(req.get("expected_format", "")).strip().lower().lstrip(".")
    result = MatchResult(request_id=str(req["request_id"]), candidates=[], warnings=[])

    if not expected:
        result.warnings.append("No expected_filename specified")
        return result

    # 1) Try exact / glob match against filename
    if any(c in expected for c in ("*", "?")):
        matches = [p for p in files if fnmatch.fnmatch(p.name.lower(), expected.lower())]
        if matches:
            # pick newest
            best = max(matches, key=lambda p: p.stat().st_mtime)
            result.matched_path = best
            result.similarity = 100
            result.candidates = [(p, 100) for p in matches]
            return result

    exact = [p for p in files if p.name.lower() == expected.lower()]
    if exact:
        result.matched_path = exact[0]
        result.similarity = 100
        return result

    # 2) Fuzzy match on filename
    file_names = [p.name for p in files]
    if not file_names:
        return result
    candidates = process.extract(expected, file_names, scorer=fuzz.token_set_ratio, limit=3)
    result.candidates = [(files[idx], int(score)) for _, score, idx in candidates]
    if candidates:
        top_name, top_score, top_idx = candidates[0]
        result.similarity = int(top_score)
        if top_score >= similarity_threshold:
            result.matched_path = files[top_idx]
    return result


def _validate(match: MatchResult, req: dict, period_end: datetime, freshness_days: int) -> None:
    if match.matched_path is None:
        return
    p = match.matched_path
    expected_ext = str(req.get("expected_format", "")).strip().lower().lstrip(".")
    # Format check
    if expected_ext and p.suffix.lower().lstrip(".") != expected_ext:
        match.warnings.append(f"Format mismatch: expected .{expected_ext}, got {p.suffix}")
        match.status = "Format Mismatch"
        return
    # Size check
    size = p.stat().st_size
    if size == 0:
        match.warnings.append("File is empty (0 bytes)")
        match.status = "Provided with Issues"
        return
    if size < EMPTY_FILE_THRESHOLD_BYTES:
        match.warnings.append(f"File is suspiciously small ({size} bytes)")
    # Freshness
    mtime = datetime.fromtimestamp(p.stat().st_mtime)
    days_old = (period_end - mtime).days
    if days_old > freshness_days:
        match.warnings.append(f"File mtime is {days_old} days before period end (stale)")
    # Multiple candidates
    if match.candidates and len(match.candidates) > 1:
        # Only warn when secondary is close
        primary, secondary = match.candidates[0][1], match.candidates[1][1]
        if secondary >= primary - 5:
            match.warnings.append(
                f"Ambiguous match: alternative {match.candidates[1][0].name} scored {secondary}"
            )
    # Final status
    if match.warnings:
        if match.status != "Format Mismatch":
            match.status = "Provided with Issues"
    else:
        match.status = "Provided"


def _normalize_account(v) -> str:
    """Stringify an account code, stripping a trailing '.0' that pandas adds when reading ints."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def _tie_out(req: dict, tie_source: pd.DataFrame | None) -> tuple[float, float, float, str]:
    """Return (gl_balance_per_request, gl_per_tie_source, variance, status)."""
    if tie_source is None or pd.isna(req.get("gl_account")):
        return float("nan"), float("nan"), float("nan"), "N/A"
    acct = _normalize_account(req["gl_account"])
    tie_source = tie_source.copy()
    tie_source["__acct_norm__"] = tie_source["gl_account"].map(_normalize_account)
    hit = tie_source[tie_source["__acct_norm__"] == acct]
    expected = float(req.get("gl_balance") or 0)
    if hit.empty:
        return expected, float("nan"), float("nan"), "GL account not found in tie source"
    actual = float(hit["total"].iloc[0])
    var = expected - actual
    status = "OK" if abs(var) < 0.01 else "INVESTIGATE"
    return expected, actual, var, status


# --- Writer ---------------------------------------------------------------

def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX


def _autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(60, max_len + 2)


def write_report(
    pbc: pd.DataFrame, matches: list[MatchResult], tie_results: list[dict],
    period_end: datetime, package_dir: Path, output: Path, freshness_days: int,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Audit Evidence Readiness Report"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Period End", period_end.strftime("%Y-%m-%d")])
    n_total = len(matches)
    n_provided = sum(1 for m in matches if m.status == "Provided")
    n_issues = sum(1 for m in matches if m.status == "Provided with Issues")
    n_format = sum(1 for m in matches if m.status == "Format Mismatch")
    n_missing = sum(1 for m in matches if m.status == "Not Found")
    ws.append(["Total Requests", n_total])
    ws.append(["Provided (clean)", n_provided])
    ws.append(["Provided with Issues", n_issues])
    ws.append(["Format Mismatch", n_format])
    ws.append(["Not Found", n_missing])
    ws.append(["% Complete", f"{(n_provided + n_issues) / n_total * 100:.0f}%" if n_total else "n/a"])
    _autosize(ws)

    # Status sheet
    ws_st = wb.create_sheet("Status")
    headers = ["Request ID", "Description", "Category", "Owner", "Expected Filename",
               "Matched File", "Similarity", "File Size (KB)", "Last Modified", "Status"]
    ws_st.append(headers)
    _style_header(ws_st)
    pbc_indexed = pbc.set_index("request_id")
    for m in matches:
        try:
            row = pbc_indexed.loc[m.request_id]
        except KeyError:
            continue
        path = m.matched_path
        size_kb = f"{path.stat().st_size / 1024:.1f}" if path else ""
        mtime = datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d") if path else ""
        ws_st.append([
            m.request_id, row["description"], row.get("category", ""), row.get("owner", ""),
            row.get("expected_filename", ""),
            str(path) if path else "(not found)",
            m.similarity, size_kb, mtime, m.status,
        ])
    for row in ws_st.iter_rows(min_row=2):
        status = row[9].value
        fill = {
            "Provided": GREEN_FILL, "Provided with Issues": YELLOW_FILL,
            "Format Mismatch": RED_FILL, "Not Found": RED_FILL,
        }.get(status)
        if fill:
            row[9].fill = fill
    _autosize(ws_st)
    ws_st.freeze_panes = "A2"

    # Validation
    ws_v = wb.create_sheet("Validation")
    ws_v.append(["Request ID", "Warning / Failure"])
    _style_header(ws_v)
    any_warn = False
    for m in matches:
        for w in (m.warnings or []):
            any_warn = True
            ws_v.append([m.request_id, w])
    if not any_warn:
        ws_v.append(["(no validation issues)"])
    _autosize(ws_v)

    # Tie-out
    ws_to = wb.create_sheet("TieOut")
    ws_to.append(["Request ID", "GL Account", "Per Request", "Per GL Source", "Variance", "Status"])
    _style_header(ws_to)
    for t in tie_results:
        ws_to.append([t["request_id"], t["gl_account"], t["expected"], t["actual"], t["variance"], t["status"]])
    for letter in ("C", "D", "E"):
        for cell in ws_to[letter][1:]:
            cell.number_format = ACCOUNTING_FMT
    for row in ws_to.iter_rows(min_row=2):
        if row[5].value == "OK":
            row[5].fill = GREEN_FILL
        elif row[5].value and row[5].value != "N/A":
            row[5].fill = RED_FILL
    _autosize(ws_to)

    # Missing
    ws_m = wb.create_sheet("Missing")
    ws_m.append(["Request ID", "Description", "Owner", "Due Date", "Expected Filename"])
    _style_header(ws_m)
    for m in matches:
        if m.status == "Not Found":
            try:
                row = pbc_indexed.loc[m.request_id]
                ws_m.append([m.request_id, row["description"], row.get("owner", ""),
                             str(row.get("due_date", "")), row.get("expected_filename", "")])
            except KeyError:
                pass
    _autosize(ws_m)

    # Audit Trail
    ws_a = wb.create_sheet("AuditTrail")
    ws_a.append(["Key", "Value"])
    _style_header(ws_a)
    ws_a.append(["Period End", period_end.strftime("%Y-%m-%d")])
    ws_a.append(["Freshness Window (days)", freshness_days])
    ws_a.append(["Package Dir", str(package_dir)])
    ws_a.append(["Generated", datetime.now().isoformat(timespec="seconds")])
    _autosize(ws_a)

    # SignOff
    ws_s = wb.create_sheet("SignOff")
    ws_s.append(["Role", "Name", "Title", "Date", "Notes"])
    _style_header(ws_s)
    ws_s.append(["Preparer (Audit Coordinator)", "", "", "", ""])
    ws_s.append(["Reviewer (Controller)", "", "", "", ""])
    _autosize(ws_s)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    LOG.info("Report written: %s", output)


def write_package(matches: list[MatchResult], pbc: pd.DataFrame,
                  package_dir: Path, do_copy: bool) -> None:
    package_dir.mkdir(parents=True, exist_ok=True)
    pbc_indexed = pbc.set_index("request_id")
    index_rows: list[dict] = []
    for m in matches:
        if m.matched_path is None:
            index_rows.append({
                "request_id": m.request_id,
                "expected_filename": pbc_indexed.loc[m.request_id].get("expected_filename", "") if m.request_id in pbc_indexed.index else "",
                "matched_path": "",
                "status": m.status,
                "package_path": "",
            })
            continue
        row = pbc_indexed.loc[m.request_id]
        category = str(row.get("category", "Other")).strip() or "Other"
        target_dir = package_dir / category
        target_dir.mkdir(parents=True, exist_ok=True)
        new_name = f"{m.request_id}_{m.matched_path.name}"
        target_path = target_dir / new_name
        if do_copy:
            shutil.copy2(m.matched_path, target_path)
        index_rows.append({
            "request_id": m.request_id,
            "expected_filename": row.get("expected_filename", ""),
            "matched_path": str(m.matched_path),
            "status": m.status,
            "package_path": str(target_path) if do_copy else "",
        })
    with open(package_dir / "index.csv", "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["request_id", "expected_filename", "matched_path", "status", "package_path"])
        writer.writeheader()
        writer.writerows(index_rows)


# --- CLI ------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--pbc", required=True, type=Path)
    parser.add_argument("--evidence-dir", required=True, type=Path)
    parser.add_argument("--period-end", required=True)
    parser.add_argument("--tie-source", type=Path, default=None)
    parser.add_argument("--package-dir", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--similarity-threshold", type=int, default=80)
    parser.add_argument("--freshness-days", type=int, default=45)
    parser.add_argument("--copy", action="store_true")
    parser.add_argument("--quiet", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.WARNING if args.quiet else logging.INFO,
        format="%(levelname)s %(name)s: %(message)s",
    )

    pbc = _read_table(args.pbc)
    required = ("request_id", "description", "expected_filename")
    missing = [c for c in required if c not in pbc.columns]
    if missing:
        LOG.error("PBC list missing required columns: %s", missing)
        return 1
    for c in ("category", "owner", "due_date", "expected_format", "gl_account", "gl_balance"):
        if c not in pbc.columns:
            pbc[c] = ""

    try:
        files = _walk_evidence(args.evidence_dir)
    except FileNotFoundError as e:
        LOG.error("%s", e)
        return 1
    LOG.info("Scanning %d files in %s", len(files), args.evidence_dir)

    period_end = datetime.strptime(args.period_end, "%Y-%m-%d")
    tie_source = None
    if args.tie_source:
        tie_source = _read_table(args.tie_source)

    matches: list[MatchResult] = []
    tie_results: list[dict] = []
    for _, req in pbc.iterrows():
        m = _match_one(req.to_dict(), files, args.similarity_threshold)
        _validate(m, req.to_dict(), period_end, args.freshness_days)
        matches.append(m)
        if pd.notna(req.get("gl_account")) and req.get("gl_account"):
            exp, act, var, status = _tie_out(req.to_dict(), tie_source)
            tie_results.append({
                "request_id": req["request_id"], "gl_account": req["gl_account"],
                "expected": exp, "actual": act, "variance": var, "status": status,
            })

    write_package(matches, pbc, args.package_dir, args.copy)
    write_report(pbc, matches, tie_results, period_end, args.package_dir, args.output, args.freshness_days)

    n_total = len(matches)
    n_provided = sum(1 for m in matches if m.status == "Provided")
    n_missing = sum(1 for m in matches if m.status == "Not Found")
    print(f"Requests:          {n_total}")
    print(f"Provided (clean):  {n_provided}")
    print(f"Missing:           {n_missing}")
    print(f"% Complete:        {(n_total - n_missing) / n_total * 100:.0f}%" if n_total else "n/a")
    print(f"Report:            {args.output}")
    print(f"Package dir:       {args.package_dir}")
    return 0 if n_missing == 0 else 2


if __name__ == "__main__":
    sys.exit(main())
