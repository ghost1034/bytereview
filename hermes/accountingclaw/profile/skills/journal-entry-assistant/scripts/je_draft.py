"""Journal Entry Assistant — Draft balanced, ERP-ready journal entries.

Implements the workflow described in ../SKILL.md. Reads a source-transactions
file plus a Chart of Accounts, performs fuzzy account matching, applies
type-driven Dr/Cr signs, balances the entry, validates against COA rules,
and writes (a) an ERP upload CSV and (b) a preparer workpaper XLSX.
"""
from __future__ import annotations

import argparse
import csv
import logging
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz, process

LOG = logging.getLogger("journal_entry_assistant")

# --- Constants -------------------------------------------------------------
ACCOUNT_TYPES = {"Asset", "Liability", "Equity", "Revenue", "Expense", "Other"}
NATURAL_DEBIT = {"Asset", "Expense"}
NATURAL_CREDIT = {"Liability", "Equity", "Revenue"}
BALANCE_TOLERANCE = 0.01
SIMILARITY_DEFAULT = 70
AMBIGUITY_MARGIN = 2  # if top two fuzzy matches are within this many points → ambiguous

ACCOUNTING_FMT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN = Side(border_style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ERP column templates
ERP_TEMPLATES = {
    "generic": ["Date", "Account Code", "Account Name", "Description", "Debit", "Credit",
                "Memo", "Department", "Class", "Project"],
    "netsuite": ["ExternalId", "TranDate", "Account", "Memo", "Debit", "Credit",
                 "Department", "Class", "Location"],
    "workday": ["JE_Number", "Line_No", "Date", "Ledger_Account", "Cost_Center",
                "Debit", "Credit", "Memo"],
    "sage_intacct": ["BATCH_NO", "RECORDNO", "ENTRY_DATE", "ACCOUNTNO", "MEMO",
                     "DEBIT", "CREDIT", "DEPARTMENTID", "CLASSID"],
}


@dataclass
class JELine:
    source_row: int
    account_code: str
    account_name: str
    account_type: str
    date: datetime
    description: str
    debit: float = 0.0
    credit: float = 0.0
    memo: str = ""
    department: str = ""
    class_: str = ""
    project: str = ""
    entity: str = ""
    similarity: int = 100
    validation_status: str = "ok"
    validation_notes: list[str] = field(default_factory=list)


@dataclass
class JEResult:
    je_date: datetime
    memo: str
    erp: str
    lines: list[JELine]
    suspense_line: JELine | None = None
    total_debit: float = 0.0
    total_credit: float = 0.0
    status: str = "PASS"
    validation: list[dict] = field(default_factory=list)
    audit: list[dict] = field(default_factory=list)


# --- I/O helpers -----------------------------------------------------------

def _read_table(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path)
    else:
        df = pd.read_excel(path)
    df.columns = [c.strip().lower() for c in df.columns]
    return df


def _validate_coa(df: pd.DataFrame) -> pd.DataFrame:
    required = ("account_code", "account_name", "type")
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"COA missing required columns: {missing}")
    df = df.copy()
    df["account_code"] = df["account_code"].astype(str).str.strip()
    df["account_name"] = df["account_name"].astype(str).str.strip()
    df["type"] = df["type"].astype(str).str.strip().str.title()
    bad_types = set(df["type"]) - ACCOUNT_TYPES
    if bad_types:
        raise ValueError(f"COA contains invalid type values: {bad_types}; allowed: {sorted(ACCOUNT_TYPES)}")
    if "is_active" not in df.columns:
        df["is_active"] = True
    if "period_open_through" not in df.columns:
        df["period_open_through"] = pd.NaT
    else:
        df["period_open_through"] = pd.to_datetime(df["period_open_through"], errors="coerce")
    if "requires_class" not in df.columns:
        df["requires_class"] = False
    return df


def _validate_source(df: pd.DataFrame) -> pd.DataFrame:
    required = ("date", "description", "amount")
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Source missing required columns: {missing}")
    df = df.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df["amount"] = pd.to_numeric(df["amount"], errors="coerce")
    bad = df[df["date"].isna() | df["amount"].isna()]
    if not bad.empty:
        LOG.warning("Dropping %d source rows with invalid date/amount", len(bad))
        df = df.dropna(subset=["date", "amount"])
    df["description"] = df["description"].fillna("").astype(str)
    for opt in ("account", "memo", "department", "class", "project", "entity", "dr_cr"):
        if opt not in df.columns:
            df[opt] = ""
        else:
            df[opt] = df[opt].fillna("").astype(str).replace({"nan": ""})
    df = df.reset_index(drop=True)
    return df


# --- Matching --------------------------------------------------------------

def _match_account(row: pd.Series, coa: pd.DataFrame, similarity_threshold: int) -> tuple[pd.Series | None, int, str]:
    """Return (matched_account_row, similarity_0_100, match_method)."""
    explicit = str(row.get("account", "")).strip()
    if explicit:
        hit = coa[coa["account_code"].str.lower() == explicit.lower()]
        if not hit.empty:
            return hit.iloc[0], 100, "by_code"
        hit = coa[coa["account_name"].str.lower() == explicit.lower()]
        if not hit.empty:
            return hit.iloc[0], 100, "by_name_exact"
        # fuzzy on account_name using the explicit hint
        candidates = process.extract(explicit, coa["account_name"].tolist(), scorer=fuzz.token_set_ratio, limit=2)
        if candidates and candidates[0][1] >= similarity_threshold:
            top, top_score, idx = candidates[0]
            if len(candidates) > 1 and (top_score - candidates[1][1]) <= AMBIGUITY_MARGIN:
                return None, top_score, "ambiguous_by_name_hint"
            return coa.iloc[idx], int(top_score), "fuzzy_by_name_hint"
        return None, candidates[0][1] if candidates else 0, "no_match_by_name_hint"
    desc = row["description"]
    if not desc.strip():
        return None, 0, "empty_description"
    candidates = process.extract(desc, coa["account_name"].tolist(), scorer=fuzz.token_set_ratio, limit=2)
    if not candidates or candidates[0][1] < similarity_threshold:
        return None, int(candidates[0][1]) if candidates else 0, "below_threshold"
    top, top_score, idx = candidates[0]
    if len(candidates) > 1 and (top_score - candidates[1][1]) <= AMBIGUITY_MARGIN:
        return None, int(top_score), "ambiguous"
    return coa.iloc[idx], int(top_score), "fuzzy_by_description"


# --- Dr/Cr assignment ------------------------------------------------------

def _assign_dr_cr(amount: float, account_type: str, explicit: str = "") -> tuple[float, float]:
    explicit = (explicit or "").strip().upper()
    if explicit in {"DR", "DEBIT"}:
        return abs(amount), 0.0
    if explicit in {"CR", "CREDIT"}:
        return 0.0, abs(amount)
    if account_type in NATURAL_DEBIT:
        return (abs(amount), 0.0) if amount >= 0 else (0.0, abs(amount))
    if account_type in NATURAL_CREDIT:
        return (0.0, abs(amount)) if amount >= 0 else (abs(amount), 0.0)
    return (abs(amount), 0.0) if amount >= 0 else (0.0, abs(amount))


# --- Core orchestration ---------------------------------------------------

def build_je(
    source: pd.DataFrame,
    coa: pd.DataFrame,
    je_date: datetime,
    memo: str,
    erp: str,
    cash_account: str,
    suspense_account: str,
    similarity_threshold: int,
    strict: bool = False,
) -> JEResult:
    lines: list[JELine] = []
    validation: list[dict] = []
    audit: list[dict] = []
    suspense_amount = 0.0
    auto_offsets: list[tuple[float, str]] = []  # (offset_amount, source_description)

    coa_index_by_code = {c.lower(): r for c, r in zip(coa["account_code"], coa.to_dict("records"))}
    suspense_row = coa_index_by_code.get(suspense_account.lower())
    if suspense_row is None:
        raise ValueError(f"--suspense-account {suspense_account} not present in COA")
    cash_row = coa_index_by_code.get(cash_account.lower())
    if cash_row is None:
        raise ValueError(f"--cash-account {cash_account} not present in COA")

    seen_signed = 0.0
    for idx, row in source.iterrows():
        acct_row, sim, method = _match_account(row, coa, similarity_threshold)
        notes: list[str] = []
        if acct_row is None:
            # Route to suspense
            line = JELine(
                source_row=int(idx),
                account_code=str(suspense_row["account_code"]),
                account_name=str(suspense_row["account_name"]),
                account_type=str(suspense_row["type"]),
                date=row["date"],
                description=row["description"],
                memo=row.get("memo", ""),
                department=row.get("department", ""),
                class_=row.get("class", ""),
                project=row.get("project", ""),
                entity=row.get("entity", ""),
                similarity=int(sim),
                validation_status="suspense",
                validation_notes=[f"Account resolution failed via method='{method}' (sim={sim})"],
            )
            d, c = _assign_dr_cr(row["amount"], "Other", row.get("dr_cr", ""))
            line.debit, line.credit = d, c
            lines.append(line)
            suspense_amount += abs(row["amount"])
            validation.append({"source_row": int(idx), "status": "suspense",
                               "detail": f"method={method}; sim={sim}; description={row['description']!r}"})
            continue

        # Validation rules
        if not bool(acct_row.get("is_active", True)):
            notes.append("account inactive")
        if pd.notna(acct_row.get("period_open_through")):
            if row["date"] > acct_row["period_open_through"]:
                notes.append(
                    f"period closed (open through {acct_row['period_open_through'].date()})"
                )
        if bool(acct_row.get("requires_class", False)) and not str(row.get("class", "")).strip():
            notes.append("class dimension required but missing")
        if abs(row["amount"]) < 1e-9:
            notes.append("zero-amount line rejected")
            validation.append({"source_row": int(idx), "status": "rejected", "detail": "; ".join(notes)})
            continue

        d, c = _assign_dr_cr(row["amount"], acct_row["type"], row.get("dr_cr", ""))
        line = JELine(
            source_row=int(idx),
            account_code=str(acct_row["account_code"]),
            account_name=str(acct_row["account_name"]),
            account_type=str(acct_row["type"]),
            date=row["date"],
            description=row["description"],
            debit=d,
            credit=c,
            memo=row.get("memo", ""),
            department=row.get("department", ""),
            class_=row.get("class", ""),
            project=row.get("project", ""),
            entity=row.get("entity", ""),
            similarity=int(sim),
            validation_status="warn" if notes else "ok",
            validation_notes=notes,
        )
        lines.append(line)
        if notes:
            validation.append({"source_row": int(idx), "status": "warn", "detail": "; ".join(notes)})
        seen_signed += d - c
        audit.append({"step": "line", "source_row": int(idx),
                      "account": f"{acct_row['account_code']} {acct_row['account_name']}",
                      "type": acct_row["type"], "method": method, "similarity": sim,
                      "debit": d, "credit": c})

    total_debit = sum(l.debit for l in lines)
    total_credit = sum(l.credit for l in lines)
    imbalance = round(total_debit - total_credit, 2)

    # Auto-offset to cash if the source is one-sided
    suspense_line: JELine | None = None
    if abs(imbalance) > BALANCE_TOLERANCE:
        # Decide: auto-cash-offset or suspense?
        # If imbalance corresponds exactly to net non-cash debits, balance to cash.
        # We use cash auto-offset by default unless suspense lines exist (then suspense routing already FAILS the JE).
        if suspense_amount > 0:
            method = "suspense"
        else:
            method = "auto_cash_offset"
        if method == "auto_cash_offset":
            offset = JELine(
                source_row=-1,
                account_code=str(cash_row["account_code"]),
                account_name=str(cash_row["account_name"]),
                account_type=str(cash_row["type"]),
                date=je_date,
                description="Auto-generated cash offset",
                debit=0.0 if imbalance > 0 else abs(imbalance),
                credit=abs(imbalance) if imbalance > 0 else 0.0,
                memo=memo,
                validation_status="auto_offset",
                validation_notes=[f"Imbalance of {imbalance:.2f} auto-offset to cash account {cash_row['account_code']}"],
            )
            lines.append(offset)
            audit.append({"step": "auto_offset", "imbalance": imbalance,
                          "account": f"{cash_row['account_code']} {cash_row['account_name']}"})
        else:
            sus = JELine(
                source_row=-1,
                account_code=str(suspense_row["account_code"]),
                account_name=str(suspense_row["account_name"]),
                account_type="Other",
                date=je_date,
                description="Suspense plug — JE out of balance",
                debit=0.0 if imbalance > 0 else abs(imbalance),
                credit=abs(imbalance) if imbalance > 0 else 0.0,
                memo=memo,
                validation_status="suspense",
                validation_notes=[f"JE imbalance of {imbalance:.2f} routed to Suspense"],
            )
            lines.append(sus)
            suspense_line = sus
            audit.append({"step": "suspense_plug", "imbalance": imbalance})

    # Recompute totals
    total_debit = round(sum(l.debit for l in lines), 2)
    total_credit = round(sum(l.credit for l in lines), 2)

    # Determine status
    in_balance = abs(total_debit - total_credit) <= BALANCE_TOLERANCE
    has_suspense = any(l.validation_status == "suspense" for l in lines)
    has_warn = any(l.validation_status == "warn" for l in lines)
    if not in_balance:
        status = "FAIL"
    elif has_suspense:
        status = "FAIL"
    elif has_warn and strict:
        status = "FAIL"
    else:
        status = "PASS"
    audit.append({"step": "status", "in_balance": in_balance, "has_suspense": has_suspense,
                  "has_warn": has_warn, "total_debit": total_debit, "total_credit": total_credit,
                  "status": status})

    return JEResult(je_date=je_date, memo=memo, erp=erp, lines=lines,
                    suspense_line=suspense_line, total_debit=total_debit,
                    total_credit=total_credit, status=status,
                    validation=validation, audit=audit)


# --- ERP CSV writer -------------------------------------------------------

def write_erp_csv(result: JEResult, output: Path) -> None:
    template = ERP_TEMPLATES.get(result.erp, ERP_TEMPLATES["generic"])
    rows: list[dict[str, str]] = []
    je_number = f"JE-{result.je_date.strftime('%Y%m%d')}"
    for i, l in enumerate(result.lines, start=1):
        if result.erp == "netsuite":
            rows.append({
                "ExternalId": f"{je_number}-{i:03d}",
                "TranDate": l.date.strftime("%Y-%m-%d"),
                "Account": l.account_code,
                "Memo": l.memo or l.description,
                "Debit": f"{l.debit:.2f}" if l.debit else "",
                "Credit": f"{l.credit:.2f}" if l.credit else "",
                "Department": l.department,
                "Class": l.class_,
                "Location": l.entity,
            })
        elif result.erp == "workday":
            rows.append({
                "JE_Number": je_number,
                "Line_No": str(i),
                "Date": l.date.strftime("%Y-%m-%d"),
                "Ledger_Account": l.account_code,
                "Cost_Center": l.department,
                "Debit": f"{l.debit:.2f}" if l.debit else "",
                "Credit": f"{l.credit:.2f}" if l.credit else "",
                "Memo": l.memo or l.description,
            })
        elif result.erp == "sage_intacct":
            rows.append({
                "BATCH_NO": je_number,
                "RECORDNO": str(i),
                "ENTRY_DATE": l.date.strftime("%Y-%m-%d"),
                "ACCOUNTNO": l.account_code,
                "MEMO": l.memo or l.description,
                "DEBIT": f"{l.debit:.2f}" if l.debit else "",
                "CREDIT": f"{l.credit:.2f}" if l.credit else "",
                "DEPARTMENTID": l.department,
                "CLASSID": l.class_,
            })
        else:
            rows.append({
                "Date": l.date.strftime("%Y-%m-%d"),
                "Account Code": l.account_code,
                "Account Name": l.account_name,
                "Description": l.description,
                "Debit": f"{l.debit:.2f}" if l.debit else "",
                "Credit": f"{l.credit:.2f}" if l.credit else "",
                "Memo": l.memo,
                "Department": l.department,
                "Class": l.class_,
                "Project": l.project,
            })
    output.parent.mkdir(parents=True, exist_ok=True)
    with open(output, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=template)
        writer.writeheader()
        for r in rows:
            writer.writerow({k: r.get(k, "") for k in template})
    LOG.info("ERP CSV written: %s (%d lines)", output, len(rows))


# --- Workpaper XLSX writer ------------------------------------------------

def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX


def _autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(50, max_len + 2)


def write_workpaper(result: JEResult, output: Path, source_path: Path, coa_path: Path,
                    erp_csv_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Journal Entry"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["JE Date", result.je_date.strftime("%Y-%m-%d")])
    ws.append(["Memo", result.memo])
    ws.append(["ERP Format", result.erp])
    ws.append(["Total Debits", result.total_debit])
    ws.append(["Total Credits", result.total_credit])
    ws.append(["Status", result.status])
    ws.append(["Line Count", len(result.lines)])
    for cell in ("B6", "B7"):
        ws[cell].number_format = ACCOUNTING_FMT
    ws["B8"].font = Font(bold=True)
    ws["B8"].fill = GREEN_FILL if result.status == "PASS" else RED_FILL
    _autosize(ws)

    # Lines
    ws_l = wb.create_sheet("Lines")
    header = ["Source Row", "Date", "Account Code", "Account Name", "Type", "Description",
              "Debit", "Credit", "Memo", "Department", "Class", "Project", "Entity",
              "Similarity", "Validation"]
    ws_l.append(header)
    _style_header(ws_l)
    for l in result.lines:
        ws_l.append([
            l.source_row if l.source_row >= 0 else "(auto)",
            l.date.strftime("%Y-%m-%d"),
            l.account_code, l.account_name, l.account_type, l.description,
            l.debit if l.debit else None, l.credit if l.credit else None,
            l.memo, l.department, l.class_, l.project, l.entity,
            l.similarity, l.validation_status,
        ])
    for cell in ws_l["G"][1:] + ws_l["H"][1:]:
        cell.number_format = ACCOUNTING_FMT
    _autosize(ws_l)
    ws_l.freeze_panes = "A2"

    # Validation
    ws_v = wb.create_sheet("Validation")
    ws_v.append(["Source Row", "Status", "Detail"])
    _style_header(ws_v)
    if not result.validation:
        ws_v.append(["", "ok", "(no validation issues)"])
    else:
        for v in result.validation:
            ws_v.append([v["source_row"], v["status"], v["detail"]])
    _autosize(ws_v)

    # BS / PL impact
    bs_rows = {}
    pl_rows = {}
    for l in result.lines:
        impact = l.debit - l.credit
        target = bs_rows if l.account_type in {"Asset", "Liability", "Equity", "Other"} else pl_rows
        key = (l.account_code, l.account_name, l.account_type)
        target[key] = target.get(key, 0.0) + impact

    for sheet_name, rows in [("BSImpact", bs_rows), ("PLImpact", pl_rows)]:
        ws_imp = wb.create_sheet(sheet_name)
        ws_imp.append(["Account Code", "Account Name", "Type", "Net Impact (Dr - Cr)"])
        _style_header(ws_imp)
        if not rows:
            ws_imp.append(["", "", "", 0])
        else:
            for (code, name, t), v in sorted(rows.items()):
                ws_imp.append([code, name, t, round(v, 2)])
        for cell in ws_imp["D"][1:]:
            cell.number_format = ACCOUNTING_FMT
        _autosize(ws_imp)

    # Audit Trail
    ws_a = wb.create_sheet("AuditTrail")
    ws_a.append(["Step", "Details"])
    _style_header(ws_a)
    for a in result.audit:
        step = a.pop("step")
        ws_a.append([step, "; ".join(f"{k}={v}" for k, v in a.items())])
    ws_a.append([])
    ws_a.append(["Input - Source", str(source_path)])
    ws_a.append(["Input - COA", str(coa_path)])
    ws_a.append(["Output - ERP CSV", str(erp_csv_path)])
    ws_a.append(["Generated", datetime.now().isoformat(timespec="seconds")])
    _autosize(ws_a)

    # SignOff
    ws_s = wb.create_sheet("SignOff")
    ws_s.append(["Role", "Name", "Title", "Date", "Notes"])
    _style_header(ws_s)
    ws_s.append(["Preparer", "", "", "", ""])
    ws_s.append(["Reviewer", "", "", "", ""])
    ws_s.append(["Approver", "", "", "", ""])
    _autosize(ws_s)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    LOG.info("Workpaper written: %s", output)


# --- CLI -------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--source", required=True, type=Path)
    parser.add_argument("--coa", required=True, type=Path)
    parser.add_argument("--je-date", required=True, help="Effective date YYYY-MM-DD")
    parser.add_argument("--memo", default="")
    parser.add_argument("--erp", default="generic", choices=sorted(ERP_TEMPLATES))
    parser.add_argument("--cash-account", default="1000")
    parser.add_argument("--suspense-account", default="9999")
    parser.add_argument("--similarity-threshold", type=int, default=SIMILARITY_DEFAULT)
    parser.add_argument("--output-csv", required=True, type=Path)
    parser.add_argument("--output-xlsx", required=True, type=Path)
    parser.add_argument("--strict", action="store_true",
                        help="Treat any validation warning as FAIL")
    parser.add_argument("--quiet", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.WARNING if args.quiet else logging.INFO,
        format="%(levelname)s %(name)s: %(message)s",
    )

    try:
        source = _validate_source(_read_table(args.source))
        coa = _validate_coa(_read_table(args.coa))
    except ValueError as e:
        LOG.error("Input validation failed: %s", e)
        return 1

    je_date = datetime.strptime(args.je_date, "%Y-%m-%d")
    try:
        result = build_je(
            source=source, coa=coa, je_date=je_date, memo=args.memo, erp=args.erp,
            cash_account=args.cash_account, suspense_account=args.suspense_account,
            similarity_threshold=args.similarity_threshold, strict=args.strict,
        )
    except ValueError as e:
        LOG.error("JE build failed: %s", e)
        return 1

    write_erp_csv(result, args.output_csv)
    write_workpaper(result, args.output_xlsx, args.source, args.coa, args.output_csv)

    print(f"Total Debits:    {result.total_debit:>15,.2f}")
    print(f"Total Credits:   {result.total_credit:>15,.2f}")
    print(f"Lines:           {len(result.lines)}")
    print(f"Status:          {result.status}")
    print(f"ERP CSV:         {args.output_csv}")
    print(f"Workpaper:       {args.output_xlsx}")
    return 0 if result.status == "PASS" else 2


if __name__ == "__main__":
    sys.exit(main())
