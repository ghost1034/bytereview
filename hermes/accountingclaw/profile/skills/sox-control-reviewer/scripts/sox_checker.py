"""SOX Control Reviewer — narrative vs. evidence assessment.

Implements the workflow in ../SKILL.md. Reads a control specification (YAML) and
one or more evidence sample files; produces a multi-sheet XLSX workpaper with
Design and Operating effectiveness assessments, gap log, and IPE validation.
"""
from __future__ import annotations

import argparse
import logging
import sys
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

LOG = logging.getLogger("sox_control_reviewer")

HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN = Side(border_style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# AICPA Audit Sampling Guide — non-statistical sample sizes for tests of operating effectiveness
# (high frequency → larger sample for control testing)
EXPECTED_SAMPLE_BY_FREQUENCY = {
    "daily": 25, "weekly": 5, "monthly": 2,
    "quarterly": 1, "annual": 1, "event-driven": 1,
}

# Design criteria the skill evaluates
DESIGN_CRITERIA = [
    ("objective_present", "Control objective is stated and aligned to a risk"),
    ("frequency_declared", "Control frequency is explicitly declared"),
    ("roles_named", "Preparer and reviewer roles are named in the narrative"),
    ("evidence_type_named", "Evidence type / source is identified"),
    ("required_attributes_listed", "Required evidence attributes are enumerated"),
    ("ipe_addressed", "IPE / IUC reports identified with completeness & accuracy evidence"),
    ("automation_declared", "Automation level (manual / IT-dep / automated) is declared"),
]


@dataclass
class DesignFinding:
    criterion: str
    description: str
    pass_: bool
    note: str = ""


@dataclass
class OperatingFinding:
    evidence_id: str
    attribute: str
    expected: str
    observed: str
    severity: str  # PASS / WARN / FAIL
    note: str = ""


# --- I/O ------------------------------------------------------------------

def _read_yaml(path: Path) -> dict:
    with open(path) as f:
        return yaml.safe_load(f) or {}


def _to_date(v: Any) -> datetime | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    try:
        return datetime.fromisoformat(str(v))
    except ValueError:
        try:
            return datetime.strptime(str(v), "%Y-%m-%d")
        except ValueError:
            return None


# --- Design assessment ----------------------------------------------------

def assess_design(control: dict) -> list[DesignFinding]:
    findings: list[DesignFinding] = []
    findings.append(DesignFinding(
        "objective_present", DESIGN_CRITERIA[0][1],
        bool(control.get("control_objective")) and bool(control.get("risk_addressed")),
        note=f"objective_len={len(str(control.get('control_objective', '')))}",
    ))
    findings.append(DesignFinding(
        "frequency_declared", DESIGN_CRITERIA[1][1],
        str(control.get("frequency", "")).lower() in EXPECTED_SAMPLE_BY_FREQUENCY,
        note=f"frequency={control.get('frequency')!r}",
    ))
    narr = str(control.get("narrative", "")).lower()
    has_preparer = any(w in narr for w in ["prepar", "clerk", "analyst", "accountant"])
    has_reviewer = any(w in narr for w in ["review", "approv", "manager", "controller", "supervis"])
    findings.append(DesignFinding(
        "roles_named", DESIGN_CRITERIA[2][1],
        has_preparer and has_reviewer,
        note=f"preparer_term_found={has_preparer}; reviewer_term_found={has_reviewer}",
    ))
    findings.append(DesignFinding(
        "evidence_type_named", DESIGN_CRITERIA[3][1],
        any(w in narr for w in ["log", "screenshot", "report", "system", "audit trail", "signature", "email"]),
        note="",
    ))
    findings.append(DesignFinding(
        "required_attributes_listed", DESIGN_CRITERIA[4][1],
        bool(control.get("required_attributes")) and len(control["required_attributes"]) >= 3,
        note=f"count={len(control.get('required_attributes', []))}",
    ))
    findings.append(DesignFinding(
        "ipe_addressed", DESIGN_CRITERIA[5][1],
        bool(control.get("ipe")) and all("completeness" in str(i).lower() or "accuracy" in str(i).lower()
                                          for i in control.get("ipe", [{}])[0].keys()
                                          if control.get("ipe")) if control.get("ipe") else False,
        note=f"ipe_count={len(control.get('ipe', []) or [])}",
    ))
    findings.append(DesignFinding(
        "automation_declared", DESIGN_CRITERIA[6][1],
        bool(control.get("automation")),
        note=f"automation={control.get('automation')!r}",
    ))
    return findings


def design_verdict(findings: list[DesignFinding]) -> str:
    fails = sum(1 for f in findings if not f.pass_)
    if fails == 0:
        return "Effective"
    if fails <= 2:
        return "Effective with Observations"
    return "Deficient"


# --- Operating assessment -------------------------------------------------

def assess_evidence(control: dict, evidence: dict) -> list[OperatingFinding]:
    """Return per-attribute operating findings for one evidence sample."""
    ev_id = str(evidence.get("evidence_id", "(no id)"))
    required = control.get("required_attributes", []) or []
    attributes = evidence.get("attributes", {}) or {}
    findings: list[OperatingFinding] = []
    # Required attributes present
    for attr in required:
        val = attributes.get(attr)
        if val is None or val == "":
            findings.append(OperatingFinding(ev_id, attr, "non-blank", "blank/missing", "FAIL",
                                             "Required attribute is missing"))
        else:
            findings.append(OperatingFinding(ev_id, attr, "non-blank", str(val), "PASS"))
    # Segregation of duties
    preparer = str(attributes.get("preparer_id", "")).strip().lower()
    approver = str(attributes.get("approver_id", "")).strip().lower()
    if preparer and approver:
        if preparer == approver:
            findings.append(OperatingFinding(
                ev_id, "segregation_of_duties", "preparer != approver",
                f"preparer={preparer!r}; approver={approver!r}", "FAIL",
                "Same person prepared and approved — segregation-of-duties violation",
            ))
        else:
            findings.append(OperatingFinding(
                ev_id, "segregation_of_duties", "preparer != approver",
                f"preparer={preparer!r}; approver={approver!r}", "PASS",
            ))
    # Date logic
    prepared = _to_date(attributes.get("prepared_date"))
    approved = _to_date(attributes.get("approved_date"))
    if prepared and approved:
        if approved < prepared:
            findings.append(OperatingFinding(
                ev_id, "approval_date_logic", "approved_date >= prepared_date",
                f"prepared={prepared.date()}; approved={approved.date()}", "FAIL",
                "Approval predates preparation — date integrity violation",
            ))
        elif (approved - prepared).days > 30:
            findings.append(OperatingFinding(
                ev_id, "approval_timeliness", "approval within 30 days",
                f"lag={(approved - prepared).days} days", "WARN",
                "Approval lag exceeds 30 days — timeliness concern",
            ))
        else:
            findings.append(OperatingFinding(
                ev_id, "approval_date_logic", "approved_date >= prepared_date",
                f"lag={(approved - prepared).days} days", "PASS",
            ))
    # IPE validation
    ipe_val = evidence.get("ipe_validation", {}) or {}
    if control.get("ipe"):
        if not ipe_val.get("report_complete"):
            findings.append(OperatingFinding(
                ev_id, "ipe_completeness", "report_complete=True",
                str(ipe_val.get("report_complete")), "FAIL",
                "IPE completeness not evidenced — auditor cannot rely on underlying data",
            ))
        else:
            findings.append(OperatingFinding(
                ev_id, "ipe_completeness", "report_complete=True", "True", "PASS",
            ))
        if not ipe_val.get("ties_to_sub_ledger"):
            findings.append(OperatingFinding(
                ev_id, "ipe_accuracy", "ties_to_sub_ledger=True",
                str(ipe_val.get("ties_to_sub_ledger")), "FAIL",
                "IPE accuracy not evidenced — does not tie to sub-ledger",
            ))
        else:
            variance = ipe_val.get("variance_to_sub_ledger", 0)
            if abs(float(variance or 0)) > 0.01:
                findings.append(OperatingFinding(
                    ev_id, "ipe_accuracy", "variance == 0", f"{variance}", "WARN",
                    "IPE ties but with rounding variance",
                ))
            else:
                findings.append(OperatingFinding(
                    ev_id, "ipe_accuracy", "ties_to_sub_ledger=True", "True", "PASS",
                ))
    return findings


def sample_verdict(findings: list[OperatingFinding]) -> str:
    if any(f.severity == "FAIL" for f in findings):
        return "FAIL"
    if any(f.severity == "WARN" for f in findings):
        return "WARN"
    return "PASS"


# --- Writer --------------------------------------------------------------

def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX


def _autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(80, max_len + 2)


def write_workpaper(control: dict, evidence_list: list[dict],
                    design: list[DesignFinding],
                    operating_by_sample: dict[str, list[OperatingFinding]],
                    expected_sample_size: int, output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["SOX Control Review"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Control ID", control.get("control_id", "")])
    ws.append(["Frequency", control.get("frequency", "")])
    ws.append(["Type", control.get("type", "")])
    ws.append(["Automation", control.get("automation", "")])
    ws.append([])
    ws.append(["Design Verdict", design_verdict(design)])
    sample_verdicts = {ev_id: sample_verdict(f) for ev_id, f in operating_by_sample.items()}
    n_fail = sum(1 for v in sample_verdicts.values() if v == "FAIL")
    n_warn = sum(1 for v in sample_verdicts.values() if v == "WARN")
    n_pass = sum(1 for v in sample_verdicts.values() if v == "PASS")
    op_verdict = "FAIL" if n_fail else ("WARN" if n_warn else "PASS")
    ws.append(["Operating Verdict", op_verdict])
    ws.append(["Samples Tested", len(evidence_list)])
    ws.append(["Samples Expected", expected_sample_size])
    if len(evidence_list) < expected_sample_size:
        ws.cell(row=ws.max_row, column=2).fill = YELLOW_FILL
    ws.append(["Samples PASS", n_pass])
    ws.append(["Samples WARN", n_warn])
    ws.append(["Samples FAIL", n_fail])
    # color the verdicts
    for cell in (ws["B8"], ws["B9"]):
        cell.font = Font(bold=True)
        if cell.value in ("Effective", "PASS"):
            cell.fill = GREEN_FILL
        elif cell.value in ("Effective with Observations", "WARN"):
            cell.fill = YELLOW_FILL
        else:
            cell.fill = RED_FILL
    _autosize(ws)

    # Design Assessment
    ws_d = wb.create_sheet("DesignAssessment")
    ws_d.append(["Criterion", "Description", "Result", "Note"])
    _style_header(ws_d)
    for f in design:
        ws_d.append([f.criterion, f.description, "PASS" if f.pass_ else "FAIL", f.note])
    for row in ws_d.iter_rows(min_row=2):
        row[2].fill = GREEN_FILL if row[2].value == "PASS" else RED_FILL
    _autosize(ws_d)

    # Operating Results
    ws_o = wb.create_sheet("OperatingResults")
    ws_o.append(["Evidence ID", "Sample Verdict", "Attribute", "Expected", "Observed", "Severity", "Note"])
    _style_header(ws_o)
    for ev_id, findings in operating_by_sample.items():
        verdict = sample_verdict(findings)
        for f in findings:
            ws_o.append([ev_id, verdict, f.attribute, f.expected, f.observed, f.severity, f.note])
    for row in ws_o.iter_rows(min_row=2):
        sev = row[5].value
        if sev == "FAIL":
            row[5].fill = RED_FILL
        elif sev == "WARN":
            row[5].fill = YELLOW_FILL
        elif sev == "PASS":
            row[5].fill = GREEN_FILL
    _autosize(ws_o)
    ws_o.freeze_panes = "A2"

    # Gap Log
    ws_g = wb.create_sheet("GapLog")
    ws_g.append(["Evidence ID", "Gap Type", "Severity", "Description", "Recommendation"])
    _style_header(ws_g)
    has_gap = False
    for ev_id, findings in operating_by_sample.items():
        for f in findings:
            if f.severity in ("WARN", "FAIL"):
                has_gap = True
                rec = {
                    "FAIL": "Treat as control deficiency — escalate per AS 2201 §61 for SD/MW evaluation",
                    "WARN": "Document compensating evidence or strengthen the control",
                }.get(f.severity, "Review")
                ws_g.append([ev_id, f.attribute, f.severity, f.note or f.observed, rec])
    if not has_gap:
        ws_g.append(["(no gaps detected)"])
    for row in ws_g.iter_rows(min_row=2):
        if row[2].value == "FAIL":
            for c in row:
                c.fill = RED_FILL
        elif row[2].value == "WARN":
            for c in row:
                c.fill = YELLOW_FILL
    _autosize(ws_g)

    # IPE Validation
    ws_i = wb.create_sheet("IPEValidation")
    ws_i.append(["Report Name", "Source", "Completeness Evidence", "Accuracy Evidence"])
    _style_header(ws_i)
    for ipe in control.get("ipe", []) or []:
        ws_i.append([
            ipe.get("report_name", ""), ipe.get("source", ""),
            ipe.get("completeness_evidence", ""), ipe.get("accuracy_evidence", ""),
        ])
    _autosize(ws_i)

    # Audit Trail
    ws_a = wb.create_sheet("AuditTrail")
    ws_a.append(["Key", "Value"])
    _style_header(ws_a)
    ws_a.append(["Control ID", control.get("control_id", "")])
    ws_a.append(["Evidence Samples", len(evidence_list)])
    ws_a.append(["Required Attributes", ", ".join(control.get("required_attributes", []) or [])])
    ws_a.append(["Generated", datetime.now().isoformat(timespec="seconds")])
    _autosize(ws_a)

    # SignOff
    ws_s = wb.create_sheet("SignOff")
    ws_s.append(["Role", "Name", "Title", "Date", "Notes"])
    _style_header(ws_s)
    ws_s.append(["Preparer (SOX Analyst)", "", "", "", ""])
    ws_s.append(["Reviewer (SOX Manager)", "", "", "", ""])
    ws_s.append(["Approver (Control Owner)", "", "", "", ""])
    _autosize(ws_s)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    LOG.info("Workpaper written: %s", output)


# --- CLI ------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--control", required=True, type=Path)
    parser.add_argument("--evidence", required=True, type=Path, nargs="+")
    parser.add_argument("--required-sample-size", default="auto")
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--quiet", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.WARNING if args.quiet else logging.INFO,
        format="%(levelname)s %(name)s: %(message)s",
    )

    try:
        control = _read_yaml(args.control)
    except Exception as e:
        LOG.error("Failed to parse control: %s", e)
        return 1

    evidence_list: list[dict] = []
    for ep in args.evidence:
        try:
            ev = _read_yaml(ep)
            evidence_list.append(ev)
        except Exception as e:
            LOG.error("Failed to parse evidence %s: %s", ep, e)
            return 1

    expected = EXPECTED_SAMPLE_BY_FREQUENCY.get(
        str(control.get("frequency", "")).lower(), 5,
    )
    if args.required_sample_size != "auto":
        try:
            expected = int(args.required_sample_size)
        except ValueError:
            LOG.error("--required-sample-size must be 'auto' or an integer")
            return 1

    design = assess_design(control)
    op_by_sample: dict[str, list[OperatingFinding]] = {}
    for ev in evidence_list:
        ev_id = str(ev.get("evidence_id", f"sample-{len(op_by_sample) + 1}"))
        op_by_sample[ev_id] = assess_evidence(control, ev)

    write_workpaper(control, evidence_list, design, op_by_sample, expected, args.output)

    n_fail = sum(1 for f in op_by_sample.values() if sample_verdict(f) == "FAIL")
    n_warn = sum(1 for f in op_by_sample.values() if sample_verdict(f) == "WARN")
    print(f"Control:          {control.get('control_id')}")
    print(f"Design Verdict:   {design_verdict(design)}")
    print(f"Samples Tested:   {len(evidence_list)} (expected {expected})")
    print(f"  PASS:           {len(op_by_sample) - n_fail - n_warn}")
    print(f"  WARN:           {n_warn}")
    print(f"  FAIL:           {n_fail}")
    print(f"Workpaper:        {args.output}")
    return 0 if (n_fail == 0 and design_verdict(design) != "Deficient") else 2


if __name__ == "__main__":
    sys.exit(main())
