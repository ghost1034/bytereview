"""Balance Sheet Reconciler — GL vs. Support reconciliation with roll-forward + aging.

Implements the workflow described in ../SKILL.md. Produces a multi-sheet XLSX
workpaper that ties the General Ledger ending balance to an external support
document, identifies and ages reconciling items, and emits a PASS/FAIL status.
"""
from __future__ import annotations

import argparse
import logging
import sys
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz

LOG = logging.getLogger("balance_sheet_reconciler")

# --- Domain constants ---
MATERIALITY_DEFAULT_USD = 5_000.0
MATCH_TOLERANCE_USD_DEFAULT = 0.01
MATCH_TOLERANCE_DAYS_DEFAULT = 3
DESCRIPTION_SIMILARITY_DEFAULT = 60  # 0-100 scale (rapidfuzz). Used as tie-breaker when multiple candidates exist; a unique amount+date candidate always matches regardless of similarity.
AGING_BUCKETS = ("0-30", "31-60", "61-90", ">90")
REQUIRED_COLUMNS = ("date", "description", "amount")

ACCOUNTING_FMT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN = Side(border_style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


@dataclass
class ReconResult:
    gl_balance: float
    support_balance: float
    variance: float
    status: str
    materiality: float
    matches_exact: int = 0
    matches_fuzzy: int = 0
    unmatched_gl: pd.DataFrame = field(default_factory=pd.DataFrame)
    unmatched_support: pd.DataFrame = field(default_factory=pd.DataFrame)
    aging_summary: pd.DataFrame = field(default_factory=pd.DataFrame)
    roll_forward: pd.DataFrame | None = None
    detail: pd.DataFrame = field(default_factory=pd.DataFrame)
    audit_trail: list[dict] = field(default_factory=list)


# --- I/O helpers -----------------------------------------------------------

def _read_table(path: Path) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        df = pd.read_csv(path)
    elif suffix in {".xlsx", ".xls"}:
        df = pd.read_excel(path)
    else:
        raise ValueError(f"Unsupported file type for {path}: {suffix}")
    df.columns = [c.strip().lower() for c in df.columns]
    return df


def _validate(df: pd.DataFrame, label: str, required: Iterable[str] = REQUIRED_COLUMNS) -> pd.DataFrame:
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"{label} missing required columns: {missing} (got {list(df.columns)})")
    df = df.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df["amount"] = pd.to_numeric(df["amount"], errors="coerce")
    bad = df[df["date"].isna() | df["amount"].isna()]
    if not bad.empty:
        LOG.warning("%s: dropping %d rows with invalid date or amount", label, len(bad))
        df = df.dropna(subset=["date", "amount"])
    df["description"] = df["description"].fillna("").astype(str)
    df["__source__"] = label
    df = df.reset_index(drop=True)
    df["__row_id__"] = df.index
    return df


# --- Matching --------------------------------------------------------------

def _exact_match(gl: pd.DataFrame, sup: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Inner-merge on rounded amount + exact date. Returns matched rows and the residual unmatched halves."""
    gl_key = gl.assign(__amt_key__=gl["amount"].round(2))
    sup_key = sup.assign(__amt_key__=sup["amount"].round(2))
    merged = gl_key.merge(
        sup_key,
        on=["__amt_key__", "date"],
        suffixes=("_gl", "_sup"),
        how="inner",
    )
    # de-duplicate so each row only matches once
    merged = merged.drop_duplicates(subset=["__row_id___gl"]).drop_duplicates(subset=["__row_id___sup"])
    matched_gl_ids = set(merged["__row_id___gl"])
    matched_sup_ids = set(merged["__row_id___sup"])
    return (
        merged,
        gl[~gl["__row_id__"].isin(matched_gl_ids)].copy(),
        sup[~sup["__row_id__"].isin(matched_sup_ids)].copy(),
    )


def _description_score(a: str, b: str) -> float:
    """Combined similarity score (0-100): max of token_set_ratio and partial_ratio.

    Bank-statement vs GL descriptions often share the key vendor/payee token but have
    different supplemental words ('wire', 'ACH', 'remittance'), so token_set captures
    common-vocabulary matches well and partial_ratio captures substring matches.
    """
    return max(fuzz.token_set_ratio(a, b), fuzz.partial_ratio(a, b))


def _fuzzy_match(
    gl: pd.DataFrame,
    sup: pd.DataFrame,
    amt_tol: float,
    day_tol: int,
    sim_threshold: int,
) -> tuple[list[tuple[int, int, float]], pd.DataFrame, pd.DataFrame]:
    """Greedy 1:1 fuzzy match. Returns list of (gl_idx, sup_idx, similarity) plus residual unmatched.

    Matching rules:
    1. Candidates must satisfy amount-tolerance AND date-tolerance.
    2. If a GL row has exactly one amount+date candidate on the support side, accept it
       (description differences are normal between systems).
    3. If multiple candidates exist, pick the highest description-similarity score, and
       only accept if it clears the `sim_threshold`.
    """
    if gl.empty or sup.empty:
        return [], gl.copy(), sup.copy()
    pairs: list[tuple[int, int, float]] = []
    sup_remaining = set(sup["__row_id__"].tolist())
    sup_indexed = sup.set_index("__row_id__")
    for _, gl_row in gl.iterrows():
        candidates: list[tuple[int, float]] = []
        for sup_id in sup_remaining:
            sup_row = sup_indexed.loc[sup_id]
            if abs(gl_row["amount"] - sup_row["amount"]) > amt_tol:
                continue
            if abs((gl_row["date"] - sup_row["date"]).days) > day_tol:
                continue
            sim = _description_score(gl_row["description"], sup_row["description"])
            candidates.append((int(sup_id), float(sim)))
        if not candidates:
            continue
        # Rule 2: unique candidate auto-matches
        if len(candidates) == 1:
            sup_id, sim = candidates[0]
        else:
            candidates.sort(key=lambda x: -x[1])
            sup_id, sim = candidates[0]
            if sim < sim_threshold:
                continue
        pairs.append((int(gl_row["__row_id__"]), sup_id, sim))
        sup_remaining.discard(sup_id)
    matched_gl_ids = {g for g, _, _ in pairs}
    matched_sup_ids = {s for _, s, _ in pairs}
    return (
        pairs,
        gl[~gl["__row_id__"].isin(matched_gl_ids)].copy(),
        sup[~sup["__row_id__"].isin(matched_sup_ids)].copy(),
    )


def _age(df: pd.DataFrame, as_of: datetime) -> pd.DataFrame:
    if df.empty:
        out = df.copy()
        out["age_days"] = pd.Series(dtype="int64")
        out["aging_bucket"] = pd.Series(dtype="object")
        return out
    out = df.copy()
    out["age_days"] = (as_of - out["date"]).dt.days
    def bucket(d: int) -> str:
        if d <= 30:
            return "0-30"
        if d <= 60:
            return "31-60"
        if d <= 90:
            return "61-90"
        return ">90"
    out["aging_bucket"] = out["age_days"].astype(int).map(bucket)
    return out


def _aging_summary(unmatched_gl: pd.DataFrame, unmatched_sup: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for label, df in [("Unmatched GL", unmatched_gl), ("Unmatched Support", unmatched_sup)]:
        for bucket in AGING_BUCKETS:
            sub = df[df.get("aging_bucket") == bucket] if not df.empty else df
            rows.append({
                "Side": label,
                "Aging Bucket": bucket,
                "Count": len(sub),
                "Total Amount": float(sub["amount"].sum()) if not sub.empty else 0.0,
            })
    return pd.DataFrame(rows)


def _roll_forward(prior_open: pd.DataFrame | None, current_open: pd.DataFrame) -> pd.DataFrame | None:
    if prior_open is None:
        return None
    key_cols = ["date", "amount", "description"]
    prior_keys = set(prior_open[key_cols].apply(tuple, axis=1)) if not prior_open.empty else set()
    cur_keys = set(current_open[key_cols].apply(tuple, axis=1)) if not current_open.empty else set()
    cleared = prior_keys - cur_keys
    still_open = prior_keys & cur_keys
    newly_open = cur_keys - prior_keys
    rows = []
    for label, keys in [("Cleared", cleared), ("Still Open", still_open), ("Newly Open", newly_open)]:
        for k in keys:
            rows.append({"Status": label, "Date": k[0], "Amount": k[1], "Description": k[2]})
    return pd.DataFrame(rows)


# --- Reconciliation orchestration -----------------------------------------

def reconcile(
    gl: pd.DataFrame,
    support: pd.DataFrame,
    as_of: datetime,
    materiality: float,
    amt_tol: float,
    day_tol: int,
    sim_threshold: int,
    prior_open: pd.DataFrame | None = None,
) -> ReconResult:
    gl_total = float(gl["amount"].sum())
    sup_total = float(support["amount"].sum())
    variance = gl_total - sup_total

    audit: list[dict] = []
    audit.append({"step": "load", "gl_rows": len(gl), "support_rows": len(support),
                  "gl_balance": gl_total, "support_balance": sup_total, "variance": variance})

    exact, gl_rem, sup_rem = _exact_match(gl, support)
    LOG.info("Exact matches: %d (of %d GL rows)", len(exact), len(gl))
    audit.append({"step": "exact_match", "matches": len(exact),
                  "gl_remaining": len(gl_rem), "support_remaining": len(sup_rem)})

    fuzzy, gl_rem2, sup_rem2 = _fuzzy_match(gl_rem, sup_rem, amt_tol, day_tol, sim_threshold)
    LOG.info("Fuzzy matches: %d", len(fuzzy))
    audit.append({"step": "fuzzy_match", "matches": len(fuzzy),
                  "gl_remaining": len(gl_rem2), "support_remaining": len(sup_rem2)})

    unmatched_gl = _age(gl_rem2, as_of)
    unmatched_sup = _age(sup_rem2, as_of)
    aging = _aging_summary(unmatched_gl, unmatched_sup)

    over_90 = (unmatched_gl["aging_bucket"] == ">90").sum() + (unmatched_sup["aging_bucket"] == ">90").sum() \
        if not (unmatched_gl.empty and unmatched_sup.empty) else 0
    max_unmatched = max(
        unmatched_gl["amount"].abs().max() if not unmatched_gl.empty else 0.0,
        unmatched_sup["amount"].abs().max() if not unmatched_sup.empty else 0.0,
    )
    if abs(variance) <= materiality and max_unmatched <= materiality and over_90 == 0:
        status = "PASS"
    else:
        status = "FAIL"
    audit.append({"step": "status", "variance": variance, "max_unmatched": max_unmatched,
                  "over_90_count": int(over_90), "status": status})

    # Build detail sheet
    detail_rows: list[dict] = []
    for _, r in exact.iterrows():
        detail_rows.append({
            "Match Status": "matched-exact", "Similarity": 100,
            "Date": r["date"], "GL Description": r["description_gl"], "GL Amount": r["amount_gl"],
            "Support Description": r["description_sup"], "Support Amount": r["amount_sup"],
        })
    sup_indexed = support.set_index("__row_id__")
    gl_indexed = gl.set_index("__row_id__")
    for g, s, sim in fuzzy:
        gl_r = gl_indexed.loc[g]
        sup_r = sup_indexed.loc[s]
        detail_rows.append({
            "Match Status": "matched-fuzzy", "Similarity": int(sim),
            "Date": gl_r["date"], "GL Description": gl_r["description"], "GL Amount": gl_r["amount"],
            "Support Description": sup_r["description"], "Support Amount": sup_r["amount"],
        })
    for _, r in unmatched_gl.iterrows():
        detail_rows.append({
            "Match Status": "unmatched (GL only)", "Similarity": 0,
            "Date": r["date"], "GL Description": r["description"], "GL Amount": r["amount"],
            "Support Description": "", "Support Amount": None,
        })
    for _, r in unmatched_sup.iterrows():
        detail_rows.append({
            "Match Status": "unmatched (Support only)", "Similarity": 0,
            "Date": r["date"], "GL Description": "", "GL Amount": None,
            "Support Description": r["description"], "Support Amount": r["amount"],
        })
    detail = pd.DataFrame(detail_rows).sort_values("Date").reset_index(drop=True) if detail_rows else pd.DataFrame()

    # Roll-forward (combine both sides of open items into a single canonical list)
    current_open = pd.concat(
        [unmatched_gl[["date", "amount", "description"]], unmatched_sup[["date", "amount", "description"]]],
        ignore_index=True,
    ) if not (unmatched_gl.empty and unmatched_sup.empty) else pd.DataFrame(columns=["date", "amount", "description"])
    roll = _roll_forward(prior_open, current_open)

    return ReconResult(
        gl_balance=gl_total,
        support_balance=sup_total,
        variance=variance,
        status=status,
        materiality=materiality,
        matches_exact=len(exact),
        matches_fuzzy=len(fuzzy),
        unmatched_gl=unmatched_gl.drop(columns=["__source__", "__row_id__"], errors="ignore"),
        unmatched_support=unmatched_sup.drop(columns=["__source__", "__row_id__"], errors="ignore"),
        aging_summary=aging,
        roll_forward=roll,
        detail=detail,
        audit_trail=audit,
    )


# --- Workpaper writer -----------------------------------------------------

def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX


def _autosize(ws) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col_letter].width = min(50, max_len + 2)


def _df_to_ws(ws, df: pd.DataFrame, money_cols: Iterable[str] = ()) -> None:
    if df is None or df.empty:
        ws.append(["(no data)"])
        return
    ws.append(list(df.columns))
    _style_header(ws)
    for _, row in df.iterrows():
        ws.append([row[c] if not pd.isna(row[c]) else None for c in df.columns])
    money_cols = set(money_cols)
    for col_idx, col in enumerate(df.columns, start=1):
        if col in money_cols:
            letter = get_column_letter(col_idx)
            for cell in ws[letter][1:]:
                cell.number_format = ACCOUNTING_FMT
    _autosize(ws)


def write_workpaper(result: ReconResult, account: str, as_of: datetime, output: Path,
                    gl_path: Path, support_path: Path) -> None:
    wb = Workbook()
    # --- Summary ---
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Balance Sheet Reconciliation"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Account", account])
    ws.append(["As-of Date", as_of.strftime("%Y-%m-%d")])
    ws.append(["Prepared By", "<auto>"])
    ws.append(["Prepared At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append([])
    ws.append(["GL Balance", result.gl_balance])
    ws.append(["Support Balance", result.support_balance])
    ws.append(["Variance (GL - Support)", result.variance])
    ws.append(["Materiality Threshold", result.materiality])
    ws.append(["Status", result.status])
    for cell in ("B8", "B9", "B10", "B11"):
        ws[cell].number_format = ACCOUNTING_FMT
    status_cell = ws["B12"]
    status_cell.font = Font(bold=True)
    status_cell.fill = GREEN_FILL if result.status == "PASS" else RED_FILL
    if abs(result.variance) > result.materiality:
        ws["B10"].fill = RED_FILL
    ws.append([])
    ws.append(["Exact Matches", result.matches_exact])
    ws.append(["Fuzzy Matches", result.matches_fuzzy])
    ws.append(["Unmatched (GL side)", len(result.unmatched_gl)])
    ws.append(["Unmatched (Support side)", len(result.unmatched_support)])
    _autosize(ws)
    ws.freeze_panes = "A7"

    # --- Detail ---
    _df_to_ws(wb.create_sheet("Detail"), result.detail,
              money_cols=["GL Amount", "Support Amount"])

    # --- Unmatched ---
    ws_un = wb.create_sheet("Unmatched")
    ws_un.append(["UNMATCHED — In GL not in Support"])
    ws_un["A1"].font = Font(bold=True, size=12)
    if result.unmatched_gl.empty:
        ws_un.append(["(none)"])
    else:
        cols = list(result.unmatched_gl.columns)
        ws_un.append(cols)
        _style_header(ws_un, ws_un.max_row)
        for _, r in result.unmatched_gl.iterrows():
            ws_un.append([r[c] if not pd.isna(r[c]) else None for c in cols])
    ws_un.append([])
    ws_un.append(["UNMATCHED — In Support not in GL"])
    ws_un.cell(row=ws_un.max_row, column=1).font = Font(bold=True, size=12)
    if result.unmatched_support.empty:
        ws_un.append(["(none)"])
    else:
        cols = list(result.unmatched_support.columns)
        ws_un.append(cols)
        _style_header(ws_un, ws_un.max_row)
        for _, r in result.unmatched_support.iterrows():
            ws_un.append([r[c] if not pd.isna(r[c]) else None for c in cols])
    _autosize(ws_un)

    # --- Aging ---
    _df_to_ws(wb.create_sheet("Aging"), result.aging_summary, money_cols=["Total Amount"])
    # Highlight >90
    ws_ag = wb["Aging"]
    for row in ws_ag.iter_rows(min_row=2):
        if any(c.value == ">90" for c in row) and (row[2].value or 0) > 0:
            for c in row:
                c.fill = RED_FILL

    # --- RollForward ---
    rf_ws = wb.create_sheet("RollForward")
    if result.roll_forward is None:
        rf_ws.append(["(no prior reconciliation supplied)"])
    else:
        _df_to_ws(rf_ws, result.roll_forward, money_cols=["Amount"])

    # --- Audit Trail ---
    at_ws = wb.create_sheet("AuditTrail")
    at_ws.append(["Step", "Details"])
    _style_header(at_ws)
    for entry in result.audit_trail:
        step = entry.pop("step")
        details = "; ".join(f"{k}={v}" for k, v in entry.items())
        at_ws.append([step, details])
    at_ws.append([])
    at_ws.append(["Input - GL", str(gl_path)])
    at_ws.append(["Input - Support", str(support_path)])
    at_ws.append(["Generated", datetime.now().isoformat(timespec="seconds")])
    _autosize(at_ws)

    # --- Sign-Off ---
    so_ws = wb.create_sheet("SignOff")
    so_ws.append(["Role", "Name", "Title", "Date", "Notes"])
    _style_header(so_ws)
    so_ws.append(["Preparer", "", "", "", ""])
    so_ws.append(["Reviewer", "", "", "", ""])
    so_ws.append(["Approver (Controller)", "", "", "", ""])
    _autosize(so_ws)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    LOG.info("Workpaper written: %s", output)


# --- CLI -------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--gl", required=True, type=Path, help="GL export (CSV or XLSX)")
    parser.add_argument("--support", required=True, type=Path, help="Support schedule (CSV or XLSX)")
    parser.add_argument("--prior", type=Path, default=None, help="Prior period reconciliation XLSX (optional)")
    parser.add_argument("--as-of", required=True, help="As-of date YYYY-MM-DD")
    parser.add_argument("--account", default="", help="Account label (e.g., '1000 - Cash')")
    parser.add_argument("--materiality", type=float, default=MATERIALITY_DEFAULT_USD)
    parser.add_argument("--match-tolerance-usd", type=float, default=MATCH_TOLERANCE_USD_DEFAULT)
    parser.add_argument("--match-tolerance-days", type=int, default=MATCH_TOLERANCE_DAYS_DEFAULT)
    parser.add_argument("--description-similarity", type=int, default=DESCRIPTION_SIMILARITY_DEFAULT)
    parser.add_argument("--output", required=True, type=Path, help="Workpaper XLSX output path")
    parser.add_argument("--quiet", action="store_true")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.WARNING if args.quiet else logging.INFO,
        format="%(levelname)s %(name)s: %(message)s",
    )

    try:
        gl = _validate(_read_table(args.gl), "GL")
        support = _validate(_read_table(args.support), "Support")
    except ValueError as e:
        LOG.error("Input validation failed: %s", e)
        return 1

    prior_open = None
    if args.prior:
        try:
            prior_open = pd.read_excel(args.prior, sheet_name="Unmatched")
        except Exception as e:
            LOG.warning("Could not read prior reconciliation: %s", e)
            prior_open = None

    as_of = datetime.strptime(args.as_of, "%Y-%m-%d")
    result = reconcile(
        gl=gl,
        support=support,
        as_of=as_of,
        materiality=args.materiality,
        amt_tol=args.match_tolerance_usd,
        day_tol=args.match_tolerance_days,
        sim_threshold=args.description_similarity,
        prior_open=prior_open,
    )
    write_workpaper(result, args.account, as_of, args.output, args.gl, args.support)

    print(f"GL Balance:        {result.gl_balance:>15,.2f}")
    print(f"Support Balance:   {result.support_balance:>15,.2f}")
    print(f"Variance:          {result.variance:>15,.2f}")
    print(f"Status:            {result.status}")
    print(f"Workpaper:         {args.output}")
    return 0 if result.status == "PASS" else 2


if __name__ == "__main__":
    sys.exit(main())
